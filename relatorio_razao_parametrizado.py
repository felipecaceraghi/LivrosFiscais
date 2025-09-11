import json
import pyodbc
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pyperclip

# --- PARÂMETROS DE CONFIGURAÇÃO ---
DEBUG = False
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

#
# >>>>> CONSULTA SQL PRINCIPAL (RAZÃO) PARAMETRIZADA <<<<<
#
SQL_RAZAO_TEMPLATE = """
SELECT 
    TD_DADOS_RAZAO.CODI_EMP, 
    TD_DADOS_RAZAO.CLASC, 
    TD_DADOS_RAZAO.NOMEC, 
    TD_DADOS_RAZAO.CODIC, 
    TD_DADOS_RAZAO.TIPO, 
    TD_DADOS_RAZAO.DATALAN, 
    TD_DADOS_RAZAO.NUMELAN, 
    TD_DADOS_RAZAO.SALDOANT, 
    TD_DADOS_RAZAO.CONTRAP, 
    TD_DADOS_RAZAO.ORDEM_NAT_CTA, 
    TD_DADOS_RAZAO.ORIGEM, 
    TD_DADOS_RAZAO.VALDEB, 
    TD_DADOS_RAZAO.VALCRE, 
    ROUND(TD_DADOS_RAZAO.VALDEB - TD_DADOS_RAZAO.VALCRE, 2) AS SALDO, 
    ROUND(TD_DADOS_RAZAO.SALDOANT + (TD_DADOS_RAZAO.VALDEB - TD_DADOS_RAZAO.VALCRE), 2) AS SALDO_EXE, 
    TD_DADOS_RAZAO.MASCARA, 
    TD_DADOS_RAZAO.MASCREL, 
    TD_DADOS_RAZAO.HISTORICO AS HISTORICO, 
    TD_DADOS_RAZAO.ZEBRA1 AS ZEBRA1, 
    TD_DADOS_RAZAO.TIPO_LAN, 
    TD_DADOS_RAZAO.EMISSAO, 
    TD_DADOS_RAZAO.CODI_LOTE, 
    TD_DADOS_RAZAO.NOME_FANTASIA_INCORPORACAO, 
    TD_DADOS_RAZAO.NRO_QUEBRA_INCORPORACAO, 
    TD_DADOS_RAZAO.NATUREZA, 
    CASE 
        WHEN TD_VALIDA_SCP.TEM_SCP = 'S' THEN 
            CASE WHEN TD_DADOS_RAZAO.FILIAL = 0 THEN {codi_emp} ELSE TD_DADOS_RAZAO.FILIAL END 
        ELSE TD_DADOS_RAZAO.FILIAL 
    END AS FILIAL, 
    TD_DADOS_RAZAO.CODIGO_SCP, 
    CASE 
        WHEN TD_VALIDA_SCP.TEM_SCP = 'S' THEN COALESCE(TD_DADOS_RAZAO.DESCRICAO_SCP, TD_NOME_SCP.RAZAO_EMP, '') 
        ELSE TD_DADOS_RAZAO.DESCRICAO_SCP 
    END AS DESCRICAO_SCP, 
    CASE 
        WHEN TD_CONTA_REDUTORA.TOTAL > 0 THEN 
            CASE 
                WHEN TD_DADOS_RAZAO.NATUREZA = 'D' THEN 
                    CASE WHEN TD_DADOS_RAZAO.VALCRE > 0 THEN 0 ELSE 1 END 
                ELSE 
                    CASE WHEN TD_DADOS_RAZAO.VALDEB > 0 THEN 0 ELSE 1 END 
            END 
        ELSE TD_DADOS_RAZAO.ORDEM_NAT_CTA 
    END AS ORDEM 
FROM (
    -- UNION 1: Saldos Anteriores (Principal)
    SELECT 
        CTCONTAS.CODI_EMP AS CODI_EMP, 
        CTCONTAS.CLAS_CTA AS CLASC, 
        CTCONTAS.NOME_CTA AS NOMEC, 
        CTCONTAS.CODI_CTA AS CODIC, 
        1 AS TIPO, 
        DATE ({data_inicial_sql}) AS DATALAN, 
        0 AS NUMELAN, 
        TDDEBITO.VALOR - TDCREDITO.VALOR AS SALDOANT, 
        0 AS CONTRAP, 
        0 AS ORDEM_NAT_CTA, 
        0 AS ORIGEM, 
        CAST(0 AS DECIMAL (13, 2)) AS VALDEB, 
        CAST(0 AS DECIMAL (13, 2)) AS VALCRE, 
        CAST(0 AS DECIMAL (13, 2)) AS SALDO, 
        CAST(0 AS DECIMAL (13, 2)) AS SALDO_EXE, 
        CTPARMTO.MASC_PAR AS MASCARA, 
        CTPARMTO.MASC_REL AS MASCREL, 
        CTCONTAS.NOME_CTA AS HISTORICO, 
        CTCONTAS.NOME_CTA AS ZEBRA1, 
        SPACE (1) AS TIPO_LAN, 
        CURRENT TIMESTAMP AS EMISSAO, 
        NULL AS CODI_LOTE, 
        CAST(NULL AS VARCHAR(50)) AS NOME_FANTASIA_INCORPORACAO, 
        1 AS NRO_QUEBRA_INCORPORACAO, 
        TDNATUREZA.NATUREZA AS NATUREZA, 
        (CASE WHEN CTPARMTO.EMAT_PAR = 'S' AND CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR ='S' THEN {codi_emp} ELSE 0 END) AS FILIAL, 
        0 AS CODIGO_SCP, 
        TDEMPRESA.NOME AS DESCRICAO_SCP, 
        0 AS ORDEM 
    FROM BETHADBA.CTCONTAS AS CTCONTAS 
    INNER JOIN BETHADBA.CTPARMTO AS CTPARMTO ON (CTPARMTO.CODI_EMP = {codi_emp}), 
    LATERAL(SELECT COUNT(1) AS TOTAL FROM BETHADBA.CTSELECAO_SCP SELECAO_SCP WHERE SELECAO_SCP.CODI_EMP = CTPARMTO.CODI_EMP AND SELECAO_SCP.USUARIO = CURRENT USER) AS TDEXISTE_SELECAO_SCP, 
    LATERAL(SELECT COALESCE(SUM(CTLANCTO.VLOR_LAN), 0) AS VALOR FROM BETHADBA.CTLANCTO AS CTLANCTO, LATERAL(SELECT (DSDBA.FG_EXISTE_VINCULO_LANCTO_SCP(CTLANCTO.CODI_EMP, CTLANCTO.NUME_LAN, CTLANCTO.CODI_LOTE, CTLANCTO.FILI_LAN, 'D')) AS EXISTE FROM DSDBA.DUMMY) AS TD_SCP WHERE CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTLANCTO.DATA_LAN_BUSCA < {data_inicial_sql} AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP AND CTLANCTO.CDEB_LAN = CTCONTAS.CODI_CTA AND (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'N' OR (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'S' AND TD_SCP.EXISTE = 1)) ) AS TDDEBITO, 
    LATERAL(SELECT COALESCE(SUM(CTLANCTO.VLOR_LAN), 0) AS VALOR FROM BETHADBA.CTLANCTO AS CTLANCTO, LATERAL(SELECT (DSDBA.FG_EXISTE_VINCULO_LANCTO_SCP(CTLANCTO.CODI_EMP, CTLANCTO.NUME_LAN, CTLANCTO.CODI_LOTE, CTLANCTO.FILI_LAN, 'C')) AS EXISTE FROM DSDBA.DUMMY) AS TD_SCP WHERE CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTLANCTO.DATA_LAN_BUSCA < {data_inicial_sql} AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP AND CTLANCTO.CCRE_LAN = CTCONTAS.CODI_CTA AND (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'N' OR (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'S' AND TD_SCP.EXISTE = 1)) ) AS TDCREDITO, 
    LATERAL(SELECT COUNT(*) AS QTDE FROM BETHADBA.CTLANCTO AS CTLANCTO, LATERAL(SELECT (DSDBA.FG_EXISTE_VINCULO_LANCTO_SCP(CTLANCTO.CODI_EMP, CTLANCTO.NUME_LAN, CTLANCTO.CODI_LOTE, CTLANCTO.FILI_LAN, 'D')) AS EXISTE FROM DSDBA.DUMMY) AS TD_SCP WHERE CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP AND CTLANCTO.CDEB_LAN = CTCONTAS.CODI_CTA AND CTLANCTO.DATA_LAN_BUSCA >= {data_inicial_sql} AND CTLANCTO.DATA_LAN_BUSCA <= {data_final_sql} AND (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'N' OR (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'S' AND TD_SCP.EXISTE = 1)) ) AS TDEXISTE_DEBITO, 
    LATERAL(SELECT COUNT(*) AS QTDE FROM BETHADBA.CTLANCTO AS CTLANCTO, LATERAL(SELECT (DSDBA.FG_EXISTE_VINCULO_LANCTO_SCP(CTLANCTO.CODI_EMP, CTLANCTO.NUME_LAN, CTLANCTO.CODI_LOTE, CTLANCTO.FILI_LAN, 'C')) AS EXISTE FROM DSDBA.DUMMY) AS TD_SCP WHERE CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP AND CTLANCTO.CCRE_LAN = CTCONTAS.CODI_CTA AND CTLANCTO.DATA_LAN_BUSCA >= {data_inicial_sql} AND CTLANCTO.DATA_LAN_BUSCA <= {data_final_sql} AND (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'N' OR (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'S' AND TD_SCP.EXISTE = 1)) ) AS TDEXISTE_CREDITO, 
    LATERAL(SELECT COUNT(*) AS TOTAL FROM BETHADBA.CTINCORPORACAO_RET_AUX AS RET_AUX WHERE RET_AUX.CODI_EMP = CTPARMTO.CODI_EMP) AS TDRET_AUX, 
    LATERAL ( SELECT MAX ( R.CLAS_CTA ) AS CLAS_CTA FROM BETHADBA.CTNATUREZA_CONTA R WHERE R.CODI_EMP = CTPARMTO.CODI_EMP AND R.CLAS_CTA = LEFT ( CTCONTAS.CLAS_CTA, LENGTH ( R.CLAS_CTA ) ) ) AS TD_MAX_NATUREZA, 
    LATERAL ( SELECT COALESCE(MIN(R.NATUREZA), '') AS NATUREZA, COALESCE(MIN(R.GRUPO), '') AS GRUPO FROM BETHADBA.CTNATUREZA_CONTA R WHERE R.CODI_EMP = CTPARMTO.CODI_EMP AND R.CLAS_CTA = TD_MAX_NATUREZA.CLAS_CTA ) AS TDNATUREZA, 
    LATERAL(SELECT G.RAZAO_EMP AS NOME FROM BETHADBA.GEEMPRE AS G WHERE G.CODI_EMP = CTPARMTO.CODI_EMP) AS TDEMPRESA 
    WHERE CTCONTAS.CODI_EMP = {codi_emp} 
    AND (0 = 0 OR CTCONTAS.CODI_CTA = 0) 
    AND (0 = 0 OR DSDBA.C_LEFT(CTCONTAS.CLAS_CTA, 0) = '') 
    AND CTCONTAS.TIPO_CTA = DSDBA.C_CHAR (65) 
    AND (('S' = DSDBA.C_CHAR (78) AND TDDEBITO.VALOR - TDCREDITO.VALOR <> 0 ) OR TDEXISTE_DEBITO.QTDE > 0 OR TDEXISTE_CREDITO.QTDE > 0) 
    AND TDRET_AUX.TOTAL = 0 
    AND (TDEXISTE_SELECAO_SCP.TOTAL = 0 OR 'N' = 'N') 
    AND ('S' = 'S' OR ('S' = 'N' AND TDNATUREZA.GRUPO <> 'C'))
    
    UNION 
    
    -- UNION 2: Lançamentos a Débito
    SELECT 
        CTLANCTO.FILI_LAN AS CODI_EMP, 
        CTCONTAS.CLAS_CTA AS CLASC, 
        CTCONTAS.NOME_CTA AS NOMEC, 
        CTCONTAS.CODI_CTA AS CODIC, 
        2 AS TIPO, 
        CTLANCTO.DATA_LAN AS DATALAN, 
        CTLANCTO.NUME_LAN AS NUMELAN, 
        0 AS SALDOANT, 
        COALESCE(TD_LOTE.CONTRA_PARTIDA, CTLANCTO.CCRE_LAN) AS CONTRAP, 
        ( case when left(CTCONTAS.CLAS_CTA, 1) = '1' or left(CTCONTAS.CLAS_CTA, 1) = '4' or left(CTCONTAS.CLAS_CTA, 1) = '5' or left(CTCONTAS.CLAS_CTA, 2) = '62' or left(CTCONTAS.CLAS_CTA, 1) = '9' then case when TD_VALCREDEB.VALCRE = 0 then 0 else 1 end else case when left(CTCONTAS.CLAS_CTA, 1) = '2' or left(CTCONTAS.CLAS_CTA, 1) = '3' or left(CTCONTAS.CLAS_CTA, 2) = '61' or left(CTCONTAS.CLAS_CTA, 1) = '9' then case when TD_VALCREDEB.VALDEB = 0 then 0 else 1 end else 0 end end ) AS ORDEM_NAT_CTA, 
        ( CASE CTLANCTO.ORIG_LAN WHEN 34 THEN 9999 WHEN 2 THEN 9000 ELSE 1 END ) AS ORIGEM, 
        TD_VALCREDEB.VALDEB AS VALDEB, 
        TD_VALCREDEB.VALCRE AS VALCRE, 
        CAST ( 0 AS DECIMAL ( 13, 2 ) ) AS SALDO, 
        CAST ( 0 AS DECIMAL ( 13, 2 ) ) AS SALDO_EXE, 
        CTPARMTO.MASC_PAR AS MASCARA, 
        CTPARMTO.MASC_REL AS MASCREL, 
        TD_CONTA_CONTRA.NOME_CONTA || CTLANCTO.CHIS_LAN AS HISTORICO, 
        TD_CONTA_CONTRA.NOME_CONTA || CTLANCTO.CHIS_LAN AS ZEBRA1, 
        DSDBA.C_CHAR ( 68 ) AS TIPO_LAN, 
        CURRENT TIMESTAMP AS EMISSAO, 
        CTLANCTO.CODI_LOTE AS CODI_LOTE, 
        COALESCE(GEFILIAL.FANTASIA_EMP, '') AS NOME_FANTASIA_INCORPORACAO, 
        COALESCE(RET_AUX.NUMERO_QUEBRA, 1) AS NRO_QUEBRA_INCORPORACAO, 
        TDNATUREZA.NATUREZA AS NATUREZA, 
        TD_SCP_QUEBRA.FILIAL AS FILIAL, 
        TD_SCP_QUEBRA.CODIGO_SCP AS CODIGO_SCP, 
        COALESCE(GESCP.DESCRICAO, TDSEM_SCP.NOME_EMPRESA) AS DESCRICAO_SCP, 
        0 AS ORDEM 
    FROM BETHADBA.CTCONTAS AS CTCONTAS 
    INNER JOIN BETHADBA.CTPARMTO AS CTPARMTO ON (CTPARMTO.CODI_EMP = {codi_emp}) 
    INNER JOIN BETHADBA.CTLANCTO AS CTLANCTO FORCE INDEX(IDX_PLANO_DEB_EMP) ON CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP AND CTLANCTO.CDEB_LAN = CTCONTAS.CODI_CTA, 
    LATERAL(SELECT MAX(RET_AUX.NUMERO_QUEBRA) AS NUMERO_QUEBRA, MAX(RET_AUX.FANTASIA) AS FANTASIA FROM BETHADBA.CTINCORPORACAO_RET_AUX AS RET_AUX WHERE RET_AUX.CODI_EMP = CTLANCTO.CODI_EMP AND RET_AUX.FILIAL = CTLANCTO.FILI_LAN) AS RET_AUX, 
    LATERAL(SELECT MAX(CTLANCTO_SCP.FILIAL) AS FILIAL, MAX(CTLANCTO_SCP.I_SCP) AS I_SCP FROM BETHADBA.CTLANCTO_SCP AS CTLANCTO_SCP WHERE CTLANCTO_SCP.CODI_EMP = CTLANCTO.CODI_EMP AND CTLANCTO_SCP.NUME_LAN = CTLANCTO.NUME_LAN AND CTLANCTO_SCP.CODI_LOTE = CTLANCTO.CODI_LOTE AND CTLANCTO_SCP.FILIAL = CTLANCTO.FILI_LAN AND CTLANCTO_SCP.TIPO = 'D' ) AS CTLANCTO_SCP, 
    LATERAL(SELECT MAX(GESCP.DESCRICAO) AS DESCRICAO FROM BETHADBA.GESCP AS GESCP WHERE GESCP.CODI_EMP = CTLANCTO_SCP.FILIAL AND GESCP.I_SCP = CTLANCTO_SCP.I_SCP) AS GESCP, 
    LATERAL(SELECT MAX(GEFILIAL.FANTASIA_EMP) AS FANTASIA_EMP FROM BETHADBA.GEEMPRE AS GEFILIAL WHERE GEFILIAL.CODI_EMP = RET_AUX.FANTASIA) AS GEFILIAL, 
    LATERAL ( SELECT MAX(LANCTO.CCRE_LAN) AS CONTRA_PARTIDA FROM BETHADBA.CTLANCTO AS LANCTO INNER JOIN BETHADBA.CTLANCTOLOTE AS LOTE ON LANCTO.CODI_EMP = LOTE.CODI_EMP AND LANCTO.CODI_LOTE = LOTE.CODI_LOTE WHERE LANCTO.CODI_EMP = CTLANCTO.CODI_EMP AND LANCTO.CODI_LOTE = CTLANCTO.CODI_LOTE AND LOTE.TIPO = 'C') AS TD_LOTE, 
    LATERAL(SELECT CTLANCTO.VLOR_LAN AS VALDEB, CAST ( 0 AS DECIMAL ( 13, 2 ) ) AS VALCRE FROM DSDBA.DUMMY) AS TD_VALCREDEB, 
    LATERAL ( SELECT MAX ( R.CLAS_CTA ) AS CLAS_CTA FROM BETHADBA.CTNATUREZA_CONTA R WHERE R.CODI_EMP = CTPARMTO.CODI_EMP AND R.CLAS_CTA = LEFT ( CTCONTAS.CLAS_CTA, LENGTH ( R.CLAS_CTA ) ) ) AS TD_MAX_NATUREZA, 
    LATERAL ( SELECT COALESCE(MIN(R.NATUREZA), '') AS NATUREZA, COALESCE(MIN(R.GRUPO), '') AS GRUPO FROM BETHADBA.CTNATUREZA_CONTA R WHERE R.CODI_EMP = CTPARMTO.CODI_EMP AND R.CLAS_CTA = TD_MAX_NATUREZA.CLAS_CTA ) AS TDNATUREZA, 
    LATERAL(SELECT G.IJUC_EMP AS INSC_JUNTA, G.DJUC_EMP AS DATA_JUNTA FROM BETHADBA.GEEMPRE AS G WHERE G.CODI_EMP = CTPARMTO.CODI_EMP) AS TDEMPRESA, 
    LATERAL(SELECT G.RAZAO_EMP AS NOME_EMPRESA FROM BETHADBA.GEEMPRE AS G WHERE G.CODI_EMP = CTLANCTO.FILI_LAN) AS TDSEM_SCP, 
    LATERAL(SELECT (DSDBA.FG_EXISTE_VINCULO_LANCTO_SCP(CTLANCTO.CODI_EMP, CTLANCTO.NUME_LAN, CTLANCTO.CODI_LOTE, CTLANCTO.FILI_LAN, 'D')) AS EXISTE FROM DSDBA.DUMMY) AS TD_SCP, 
    LATERAL(SELECT(CASE WHEN 'N' = 'S' THEN COALESCE(CTLANCTO_SCP.I_SCP, 0) ELSE 0 END) AS CODIGO_SCP, (CASE WHEN 'N' = 'S' AND CTLANCTO.CODI_EMP <> CTLANCTO.FILI_LAN THEN CTLANCTO.FILI_LAN ELSE 0 END) AS FILIAL FROM DSDBA.DUMMY) AS TD_SCP_QUEBRA, 
    LATERAL(SELECT (CASE WHEN 'N' = 'N' THEN '' ELSE (CASE WHEN COALESCE(TD_LOTE.CONTRA_PARTIDA, CTLANCTO.CCRE_LAN) > 0 THEN (CASE WHEN 1 = 1 THEN CAST(MAX(CON.CODI_CTA) AS VARCHAR(30)) ELSE CAST(DSDBA.FG_MONTA_MASCARA_CT( CTPARMTO.MASC_REL, MAX(CON.CLAS_CTA)) AS VARCHAR(30)) END) || ' - ' || MAX(CON.NOME_CTA) || (CASE WHEN LENGTH(CTLANCTO.CHIS_LAN) > 0 THEN CHAR(13) ELSE '' END) ELSE CHAR(13) END) END) AS NOME_CONTA FROM BETHADBA.CTCONTAS AS CON WHERE CON.CODI_EMP = CTCONTAS.CODI_EMP AND CON.CODI_CTA = COALESCE(TD_LOTE.CONTRA_PARTIDA, CTLANCTO.CCRE_LAN)) AS TD_CONTA_CONTRA 
    WHERE CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTCONTAS.CODI_EMP = {codi_emp} 
    AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP 
    AND (0 = 0 OR CTCONTAS.CODI_CTA = 0) 
    AND (0 = 0 OR DSDBA.C_LEFT (CTCONTAS.CLAS_CTA, 0) = '') 
    AND CTCONTAS.TIPO_CTA = DSDBA.C_CHAR (65) 
    AND CTLANCTO.CDEB_LAN = CTCONTAS.CODI_CTA 
    AND CTLANCTO.DATA_LAN_BUSCA >= {data_inicial_sql} 
    AND CTLANCTO.DATA_LAN_BUSCA <= {data_final_sql} 
    AND (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'N' OR (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'S' AND TD_SCP.EXISTE = 1)) 
    AND ('S' = 'S' OR ('S' = 'N' AND TDNATUREZA.GRUPO <> 'C'))
    
    UNION 
    
    -- UNION 3: Lançamentos a Crédito
    SELECT 
        CTLANCTO.FILI_LAN AS CODI_EMP, 
        CTCONTAS.CLAS_CTA AS CLASC, 
        CTCONTAS.NOME_CTA AS NOMEC, 
        CTCONTAS.CODI_CTA AS CODIC, 
        2 AS TIPO, 
        CTLANCTO.DATA_LAN AS DATALAN, 
        CTLANCTO.NUME_LAN AS NUMELAN, 
        0 AS SALDOANT, 
        COALESCE(TD_LOTE.CONTRA_PARTIDA, CTLANCTO.CDEB_LAN) AS CONTRAP, 
        ( case when left(CTCONTAS.CLAS_CTA, 1) = '1' or left(CTCONTAS.CLAS_CTA, 1) = '4' or left(CTCONTAS.CLAS_CTA, 1) = '5' or left(CTCONTAS.CLAS_CTA, 2) = '62' or left(CTCONTAS.CLAS_CTA, 1) = '9' then case when TD_VALCREDEB.VALCRE = 0 then 0 else 1 end else case when left(CTCONTAS.CLAS_CTA, 1) = '2' or left(CTCONTAS.CLAS_CTA, 1) = '3' or left(CTCONTAS.CLAS_CTA, 2) = '61' or left(CTCONTAS.CLAS_CTA, 1) = '9' then case when TD_VALCREDEB.VALDEB = 0 then 0 else 1 end else 0 end end ) AS ORDEM_NAT_CTA, 
        ( CASE CTLANCTO.ORIG_LAN WHEN 34 THEN 9999 WHEN 2 THEN 9000 ELSE 1 END ) AS ORIGEM, 
        TD_VALCREDEB.VALDEB AS VALDEB, 
        TD_VALCREDEB.VALCRE AS VALCRE, 
        CAST ( 0 AS DECIMAL ( 13, 2 ) ) AS SALDO, 
        CAST ( 0 AS DECIMAL ( 13, 2 ) ) AS SALDO_EXE, 
        CTPARMTO.MASC_PAR AS MASCARA, 
        CTPARMTO.MASC_REL AS MASCREL, 
        TD_CONTA_CONTRA.NOME_CONTA || CTLANCTO.CHIS_LAN AS HISTORICO, 
        TD_CONTA_CONTRA.NOME_CONTA || CTLANCTO.CHIS_LAN AS ZEBRA1, 
        DSDBA.C_CHAR ( 67 ) AS TIPO_LAN, 
        CURRENT TIMESTAMP AS EMISSAO, 
        CTLANCTO.CODI_LOTE AS CODI_LOTE, 
        COALESCE(GEFILIAL.FANTASIA_EMP, '') AS NOME_FANTASIA_INCORPORACAO, 
        COALESCE(RET_AUX.NUMERO_QUEBRA, 1) AS NRO_QUEBRA_INCORPORACAO, 
        TDNATUREZA.NATUREZA AS NATUREZA, 
        TD_SCP_QUEBRA.FILIAL AS FILIAL, 
        TD_SCP_QUEBRA.CODIGO_SCP AS CODIGO_SCP, 
        COALESCE(GESCP.DESCRICAO, TDSEM_SCP.NOME_EMPRESA) AS DESCRICAO_SCP, 
        0 AS ORDEM 
    FROM BETHADBA.CTCONTAS AS CTCONTAS 
    INNER JOIN BETHADBA.CTPARMTO AS CTPARMTO ON (CTPARMTO.CODI_EMP = {codi_emp}) 
    INNER JOIN BETHADBA.CTLANCTO AS CTLANCTO FORCE INDEX(IDX_PLANO_CRE_EMP) ON CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP AND CTLANCTO.CCRE_LAN = CTCONTAS.CODI_CTA, 
    LATERAL( SELECT MAX(RET_AUX.NUMERO_QUEBRA) AS NUMERO_QUEBRA, MAX(RET_AUX.FANTASIA) AS FANTASIA FROM BETHADBA.CTINCORPORACAO_RET_AUX AS RET_AUX WHERE RET_AUX.CODI_EMP = CTLANCTO.CODI_EMP AND RET_AUX.FILIAL = CTLANCTO.FILI_LAN) AS RET_AUX, 
    LATERAL( SELECT MAX(CTLANCTO_SCP.FILIAL) AS FILIAL, MAX(CTLANCTO_SCP.I_SCP) AS I_SCP FROM BETHADBA.CTLANCTO_SCP AS CTLANCTO_SCP WHERE CTLANCTO_SCP.CODI_EMP = CTLANCTO.CODI_EMP AND CTLANCTO_SCP.NUME_LAN = CTLANCTO.NUME_LAN AND CTLANCTO_SCP.CODI_LOTE = CTLANCTO.CODI_LOTE AND CTLANCTO_SCP.FILIAL = CTLANCTO.FILI_LAN AND CTLANCTO_SCP.TIPO = 'C' ) AS CTLANCTO_SCP, 
    LATERAL( SELECT MAX(GESCP.DESCRICAO) AS DESCRICAO FROM BETHADBA.GESCP AS GESCP WHERE GESCP.CODI_EMP = CTLANCTO_SCP.FILIAL AND GESCP.I_SCP = CTLANCTO_SCP.I_SCP) AS GESCP, 
    LATERAL( SELECT MAX(GEFILIAL.FANTASIA_EMP) AS FANTASIA_EMP FROM BETHADBA.GEEMPRE AS GEFILIAL WHERE GEFILIAL.CODI_EMP = RET_AUX.FANTASIA) AS GEFILIAL, 
    LATERAL ( SELECT MAX(LANCTO.CDEB_LAN) AS CONTRA_PARTIDA FROM BETHADBA.CTLANCTO AS LANCTO INNER JOIN BETHADBA.CTLANCTOLOTE AS LOTE ON LANCTO.CODI_EMP = LOTE.CODI_EMP AND LANCTO.CODI_LOTE = LOTE.CODI_LOTE WHERE LANCTO.CODI_LOTE = CTLANCTO.CODI_LOTE AND LANCTO.CODI_EMP = CTLANCTO.CODI_EMP AND LOTE.TIPO = 'D') AS TD_LOTE, 
    LATERAL(SELECT CAST ( 0 AS DECIMAL ( 13, 2 ) ) AS VALDEB, CTLANCTO.VLOR_LAN AS VALCRE FROM DSDBA.DUMMY) AS TD_VALCREDEB, 
    LATERAL ( SELECT MAX ( R.CLAS_CTA ) AS CLAS_CTA FROM BETHADBA.CTNATUREZA_CONTA R WHERE R.CODI_EMP = CTPARMTO.CODI_EMP AND R.CLAS_CTA = LEFT ( CTCONTAS.CLAS_CTA, LENGTH ( R.CLAS_CTA ) ) ) AS TD_MAX_NATUREZA, 
    LATERAL ( SELECT COALESCE(MIN(R.NATUREZA), '') AS NATUREZA, COALESCE(MIN(R.GRUPO), '') AS GRUPO FROM BETHADBA.CTNATUREZA_CONTA R WHERE R.CODI_EMP = CTPARMTO.CODI_EMP AND R.CLAS_CTA = TD_MAX_NATUREZA.CLAS_CTA ) AS TDNATUREZA, 
    LATERAL(SELECT G.IJUC_EMP AS INSC_JUNTA, G.DJUC_EMP AS DATA_JUNTA FROM BETHADBA.GEEMPRE AS G WHERE G.CODI_EMP = CTPARMTO.CODI_EMP) AS TDEMPRESA, 
    LATERAL(SELECT G.RAZAO_EMP AS NOME_EMPRESA FROM BETHADBA.GEEMPRE AS G WHERE G.CODI_EMP = CTLANCTO.FILI_LAN) AS TDSEM_SCP, 
    LATERAL(SELECT (DSDBA.FG_EXISTE_VINCULO_LANCTO_SCP(CTLANCTO.CODI_EMP, CTLANCTO.NUME_LAN, CTLANCTO.CODI_LOTE, CTLANCTO.FILI_LAN, 'C')) AS EXISTE FROM DSDBA.DUMMY) AS TD_SCP, 
    LATERAL(SELECT(CASE WHEN 'N' = 'S' THEN COALESCE(CTLANCTO_SCP.I_SCP, 0) ELSE 0 END) AS CODIGO_SCP, (CASE WHEN 'N' = 'S' AND CTLANCTO.CODI_EMP <> CTLANCTO.FILI_LAN THEN CTLANCTO.FILI_LAN ELSE 0 END) AS FILIAL FROM DSDBA.DUMMY) AS TD_SCP_QUEBRA, 
    LATERAL(SELECT (CASE WHEN 'N' = 'N' THEN '' ELSE (CASE WHEN COALESCE(TD_LOTE.CONTRA_PARTIDA, CTLANCTO.CDEB_LAN) > 0 THEN (CASE WHEN 1 = 1 THEN CAST(MAX(CON.CODI_CTA) AS VARCHAR(30)) ELSE CAST(DSDBA.FG_MONTA_MASCARA_CT( CTPARMTO.MASC_REL, MAX(CON.CLAS_CTA)) AS VARCHAR(30)) END) || ' - ' || MAX(CON.NOME_CTA) || (CASE WHEN LENGTH(CTLANCTO.CHIS_LAN) > 0 THEN CHAR(13) ELSE '' END) ELSE CHAR(13) END) END) AS NOME_CONTA FROM BETHADBA.CTCONTAS AS CON WHERE CON.CODI_EMP = CTCONTAS.CODI_EMP AND CON.CODI_CTA = COALESCE(TD_LOTE.CONTRA_PARTIDA, CTLANCTO.CDEB_LAN)) AS TD_CONTA_CONTRA 
    WHERE CTLANCTO.CODI_EMP IN ({codi_emp}) AND {filial_filter} AND CTCONTAS.CODI_EMP = {codi_emp} 
    AND CTLANCTO.CODI_EMP_PLANO = CTCONTAS.CODI_EMP 
    AND (0 = 0 OR CTCONTAS.CODI_CTA = 0) 
    AND (0 = 0 OR DSDBA.C_LEFT ( CTCONTAS.CLAS_CTA, 0) = '') 
    AND CTCONTAS.TIPO_CTA = DSDBA.C_CHAR(65) 
    AND CTLANCTO.CCRE_LAN = CTCONTAS.CODI_CTA 
    AND CTLANCTO.DATA_LAN_BUSCA >= {data_inicial_sql} 
    AND CTLANCTO.DATA_LAN_BUSCA <= {data_final_sql} 
    AND (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'N' OR (CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR = 'S' AND TD_SCP.EXISTE = 1)) 
    AND ('S' = 'S' OR ('S' = 'N' AND TDNATUREZA.GRUPO <> 'C'))
    
) TD_DADOS_RAZAO, 
LATERAL(SELECT FIRST COUNT(1) AS TOTAL FROM BETHADBA.CTSELECAO_CONTA_REDUTORA CONTA_REDUTORA WHERE CONTA_REDUTORA.CODI_EMP = {codi_emp} AND CONTA_REDUTORA.CODI_CTA = TD_DADOS_RAZAO.CODIC) TD_CONTA_REDUTORA, 
LATERAL(SELECT CASE WHEN CTPARMTO.EMAT_PAR = 'S' AND CTPARMTO.SOCIEDADE_EM_CONTA_PARTICIPACAO_PAR ='S' THEN 'S' ELSE 'N' END AS TEM_SCP FROM BETHADBA.CTPARMTO AS CTPARMTO WHERE CTPARMTO.CODI_EMP = {codi_emp}) TD_VALIDA_SCP, 
LATERAL(SELECT G.RAZAO_EMP AS RAZAO_EMP FROM BETHADBA.GEEMPRE AS G WHERE G.CODI_EMP = {codi_emp}) TD_NOME_SCP
"""

def format_cpf_cnpj(value):
    """Formata uma string de números como CPF ou CNPJ."""
    if not value or not isinstance(value, str) or not value.isdigit():
        return value
    value = value.strip()
    if len(value) == 11:
        return f'{value[:3]}.{value[3:6]}.{value[6:9]}-{value[9:]}'
    if len(value) == 14:
        return f'{value[:2]}.{value[2:5]}.{value[5:8]}/{value[8:12]}-{value[12:]}'
    return value

def format_excel_report(df, filename, empresa_info):
    """
    Cria e formata um arquivo Excel replicando o layout do relatório Razão,
    com todos os ajustes de layout solicitados.
    """
    
    # Definição das fontes e estilos
    header_font = Font(bold=True, size=14)
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    normal_font_small = Font(size=7)
    bold_font_small = Font(bold=True, size=7)

    grouped = df.groupby(['CODIC', 'CLASC', 'NOMEC'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Razão Contábil"
    
    ws.sheet_view.showGridLines = False
    
    current_row = 1

    # Cabeçalho do Relatório
    ws.merge_cells('A1:C1'); ws['A1'] = f"Empresa: {empresa_info.get('razao_emp', 'N/A')}"
    ws.merge_cells('G1:H1'); ws['G1'] = f"Folha: {empresa_info.get('folha', '0001')}"
    ws.merge_cells('A2:C2'); ws['A2'] = f"C.N.P.J.: {empresa_info.get('cgce_emp', 'N/A')}"
    ws.merge_cells('G2:H2'); ws['G2'] = f"Emissão: {datetime.now().strftime('%d/%m/%Y')}"
    ws.merge_cells('A3:C3'); ws['A3'] = f"Período: {empresa_info.get('periodo', 'N/A')}"
    ws.merge_cells('G3:H3'); ws['G3'] = f"Hora: {datetime.now().strftime('%H:%M:%S')}"
    
    current_row = 5
    ws.merge_cells(f'A{current_row}:H{current_row}'); cell = ws[f'A{current_row}']
    cell.value = "RAZÃO"; cell.font = header_font; cell.alignment = center_align
    current_row += 2

    # Cabeçalho único das colunas
    col_headers = ['Data', 'Lote Histórico', 'Cta.C.Part.', 'Filial', 'Débito', 'Crédito', 'Saldo-Exercício']
    for col_num, header in enumerate(col_headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = bold_font_small
        cell.border = thin_border
    current_row += 1
    
    ws.freeze_panes = f'A{current_row}'

    for (codic, clasc, nomec), group in grouped:
        ws.cell(row=current_row, column=1, value="Conta:").font = bold_font_small
        ws.cell(row=current_row, column=2, value=f"{codic} - {clasc} {nomec}").font = normal_font_small
        current_row += 1
        
        saldo_anterior_row = group[group['TIPO'] == 1].iloc[0] if not group[group['TIPO'] == 1].empty else group.iloc[0]
        saldo_inicial = saldo_anterior_row.get('SALDOANT', 0)
        natureza_conta_original = saldo_anterior_row.get('NATUREZA', 'D')
        
        running_balance = saldo_inicial

        display_saldo = abs(running_balance)
        natureza_display = ''
        if natureza_conta_original == 'D':
            natureza_display = 'D' if running_balance >= 0 else 'C'
        elif natureza_conta_original == 'C':
            natureza_display = 'C' if running_balance <= 0 else 'D'

        ws.cell(row=current_row, column=2, value="SALDO ANTERIOR").font = normal_font_small
        saldo_cell = ws.cell(row=current_row, column=7, value=f"{display_saldo:,.2f}{natureza_display}")
        saldo_cell.alignment = right_align; saldo_cell.font = normal_font_small
        current_row += 1
        
        movimentos = group[group['TIPO'] == 2].sort_values(by=['DATALAN', 'TIPO_LAN']) if not group[group['TIPO'] == 2].empty else group.sort_values(by=['DATALAN'])
        
        for _, row in movimentos.iterrows():
            valdeb = row.get('VALDEB', 0) or 0
            valcre = row.get('VALCRE', 0) or 0
            running_balance += valdeb - valcre

            display_saldo = abs(running_balance)
            natureza_display = ''
            if natureza_conta_original == 'D':
                natureza_display = 'D' if running_balance >= 0 else 'C'
            elif natureza_conta_original == 'C':
                natureza_display = 'C' if running_balance <= 0 else 'D'
            
            saldo_final_str = f"{display_saldo:,.2f}"
            if abs(running_balance) > 0.001:
                saldo_final_str += natureza_display

            ws.cell(row=current_row, column=1, value=row['DATALAN'].strftime('%d/%m/%Y')).font = normal_font_small
            
            lote = int(row.get('CODI_LOTE', 0)) if pd.notna(row.get('CODI_LOTE')) else ''
            historico = row.get('HISTORICO', '')
            cell_lote_hist = ws.cell(row=current_row, column=2, value=f"{lote} {historico}")
            cell_lote_hist.font = normal_font_small
            cell_lote_hist.alignment = Alignment(wrap_text=True, vertical='top')

            ws.cell(row=current_row, column=3, value=row.get('CONTRAP', '')).font = normal_font_small
            ws.cell(row=current_row, column=4, value=row.get('FILIAL', '')).font = normal_font_small
            
            deb_cell = ws.cell(row=current_row, column=5, value=valdeb if valdeb else '')
            deb_cell.number_format = '#,##0.00'; deb_cell.font = normal_font_small

            cred_cell = ws.cell(row=current_row, column=6, value=valcre if valcre else '')
            cred_cell.number_format = '#,##0.00'; cred_cell.font = normal_font_small
            
            saldo_exe_cell = ws.cell(row=current_row, column=7, value=saldo_final_str)
            saldo_exe_cell.alignment = right_align; saldo_exe_cell.font = normal_font_small
            
            current_row += 1
            
        total_deb = group[group['TIPO'] == 2]['VALDEB'].sum() if not group[group['TIPO'] == 2].empty else 0
        total_cred = group[group['TIPO'] == 2]['VALCRE'].sum() if not group[group['TIPO'] == 2].empty else 0
        
        ws.cell(row=current_row, column=2, value="Total da conta:").font = bold_font_small
        total_deb_cell = ws.cell(row=current_row, column=5, value=total_deb)
        total_deb_cell.font = bold_font_small; total_deb_cell.number_format = '#,##0.00'
        total_cred_cell = ws.cell(row=current_row, column=6, value=total_cred)
        total_cred_cell.font = bold_font_small; total_cred_cell.number_format = '#,##0.00'
        
        current_row += 2

    current_row += 4 
    signature_start_row = current_row

    # Assinatura do Responsável (Esquerda)
    ws.merge_cells(f'A{signature_start_row}:D{signature_start_row}')
    line_cell_resp = ws[f'A{signature_start_row}']
    line_cell_resp.value = '_' * 60  # Linha de assinatura com caracteres
    line_cell_resp.alignment = Alignment(horizontal='center', vertical='bottom')
    line_cell_resp.font = normal_font_small

    ws.merge_cells(f'A{signature_start_row + 1}:D{signature_start_row + 1}'); cell = ws[f'A{signature_start_row + 1}']
    cell.value = empresa_info.get('responsavel_nome', 'N/A'); cell.alignment = center_align; cell.font = normal_font_small
    ws.merge_cells(f'A{signature_start_row + 2}:D{signature_start_row + 2}'); cell = ws[f'A{signature_start_row + 2}']
    cell.value = f"CPF: {empresa_info.get('responsavel_cpf', 'N/A')}"; cell.alignment = center_align; cell.font = normal_font_small
    
    # Assinatura do Contador (Direita)
    ws.merge_cells(f'E{signature_start_row}:H{signature_start_row}')
    line_cell_cont = ws[f'E{signature_start_row}']
    line_cell_cont.value = '_' * 60  # Linha de assinatura com caracteres
    line_cell_cont.alignment = Alignment(horizontal='center', vertical='bottom')
    line_cell_cont.font = normal_font_small
    
    ws.merge_cells(f'E{signature_start_row + 1}:H{signature_start_row + 1}'); cell = ws[f'E{signature_start_row + 1}']
    cell.value = empresa_info.get('contador_nome', 'N/A'); cell.alignment = center_align; cell.font = normal_font_small
    ws.merge_cells(f'E{signature_start_row + 2}:H{signature_start_row + 2}'); cell = ws[f'E{signature_start_row + 2}']
    cell.value = empresa_info.get('contador_crc', 'N/A'); cell.alignment = center_align; cell.font = normal_font_small
    ws.merge_cells(f'E{signature_start_row + 3}:H{signature_start_row + 3}'); cell = ws[f'E{signature_start_row + 3}']
    cell.value = f"CPF: {empresa_info.get('contador_cpf', 'N/A')}"; cell.alignment = center_align; cell.font = normal_font_small

    # Largura das Colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    
    wb.save(filename)

def gerar_relatorio_razao_com_dump(codi_emp: int, data_inicial: str, data_final: str, filiais: bool):
    print(f"\n{'='*60}\nINICIANDO FLUXO PARA EMPRESA: {codi_emp}\n{'='*60}")
    
    conn = None
    try:
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        
        empresa_info = {}
        print("\n[PASSO 1 de 7] Buscando informações da empresa...")
        cursor.execute("SELECT razao_emp, cgce_emp FROM bethadba.geempre WHERE codi_emp = ?", codi_emp)
        emp_data = cursor.fetchone()
        if emp_data:
            empresa_info['razao_emp'] = emp_data.razao_emp
            empresa_info['cgce_emp'] = format_cpf_cnpj(emp_data.cgce_emp)
            cnpj_raiz = emp_data.cgce_emp[:8]

        print("\n[PASSO 2 de 7] Buscando responsável legal...")
        sql_resp = f"""
            SELECT COALESCE(DSDBA.S_BUSCA_ALTERACAO_EMPRESA(GEEMPRE.CODI_EMP, '2050-01-01', 22), GEEMPRE.RLEG_EMP) AS RLEG_EMP,
                   COALESCE(DSDBA.S_BUSCA_ALTERACAO_EMPRESA(GEEMPRE.CODI_EMP, '2050-01-01', 21), GEEMPRE.CPF_LEG_EMP) AS CPF_LEG_EMP
            FROM BETHADBA.GEEMPRE WHERE CODI_EMP = ?
        """
        cursor.execute(sql_resp, codi_emp)
        resp_data = cursor.fetchone()
        if resp_data:
            empresa_info['responsavel_nome'] = resp_data.RLEG_EMP
            empresa_info['responsavel_cpf'] = format_cpf_cnpj(resp_data.CPF_LEG_EMP)

        print("\n[PASSO 3 de 7] Buscando dados do contador...")
        codi_con_fixo = 5 
        sql_contador = "SELECT NOME_CON, RCRC_CON, CPFC_CON, UF_CRC FROM BETHADBA.GECONTADOR WHERE CODI_CON = ?"
        cursor.execute(sql_contador, codi_con_fixo)
        cont_data = cursor.fetchone()
        if cont_data:
            empresa_info['contador_nome'] = cont_data.NOME_CON
            empresa_info['contador_crc'] = f"Reg. no CRC - {cont_data.UF_CRC} sob o No. {cont_data.RCRC_CON}"
            empresa_info['contador_cpf'] = format_cpf_cnpj(cont_data.CPFC_CON)
        
        if filiais:
            print("\n[PASSO 4 de 7] Configurando permissões temporárias para filiais e executando INSERTs auxiliares...")
            # Limpa permissões antigas e insere a da própria empresa
            cursor.execute("DELETE FROM bethadba.ctfilialuser WHERE codi_emp = ? AND usuario = CURRENT USER", codi_emp)
            cursor.execute("INSERT INTO bethadba.ctfilialuser (codi_emp, codi_fil, usuario) VALUES (?, ?, CURRENT USER)", codi_emp, codi_emp)

            sql_filiais = """
                SELECT CODI_EMP, APEL_EMP, CGCE_EMP FROM BETHADBA.GEEMPRE
                WHERE CODI_EMP <> ? AND (UCTA_EMP = 1 OR UCXA_EMP = 1) AND LEFT(CGCE_EMP, 8) = ?
                ORDER BY 2
            """
            cursor.execute(sql_filiais, codi_emp, cnpj_raiz)
            filiais_encontradas = cursor.fetchall()
            
            if filiais_encontradas:
                print(f"  -> {len(filiais_encontradas)} filiais encontradas. Adicionando permissões...")
                for filial in filiais_encontradas:
                    print(f"     - Adicionando filial: {filial.CODI_EMP} ({filial.APEL_EMP})")
                    cursor.execute("INSERT INTO bethadba.ctfilialuser (codi_emp, codi_fil, usuario) VALUES (?, ?, CURRENT USER)", codi_emp, filial.CODI_EMP)
            else:
                print("  -> Nenhuma filial adicional encontrada.")
            
            print("  -> Executando INSERT auxiliar 1 (CTLANCTO_SCP_AUX_EMP)...")
            insert_aux_1 = f"""
            INSERT INTO BETHADBA.CTLANCTO_SCP_AUX_EMP ( CODI_EMP , FILI_LAN ) ( 
                SELECT {codi_emp} , EMP.CODI_EMP 
                FROM BETHADBA.GEEMPRE AS EMP 
                WHERE LEFT ( EMP.CGCE_EMP , 8 ) = '{cnpj_raiz}' 
            )
            """
            cursor.execute(insert_aux_1)
            
            print("  -> Executando INSERT auxiliar 2 (CTLANCTO_SCP_AUX)...")
            insert_aux_2 = f"""
            INSERT INTO BETHADBA.CTLANCTO_SCP_AUX ( CODI_EMP , NUME_LAN , CODI_LOTE , FILI_LAN , TIPO ) ( 
                SELECT CTLANCTO.CODI_EMP , CTLANCTO.NUME_LAN , COALESCE ( CTLANCTO.CODI_LOTE , 0 ) , CTLANCTO.FILI_LAN , 'D' AS TIPO 
                FROM BETHADBA.CTLANCTO AS CTLANCTO 
                INNER JOIN BETHADBA.CTCONTAS AS CTCONTAS ON CTCONTAS.CODI_EMP = {codi_emp} AND CTCONTAS.CODI_CTA = CTLANCTO.CDEB_LAN 
                LEFT OUTER JOIN BETHADBA.CTLANCTO_SCP_AUX_CTA AS CTA ON CTA.CODI_EMP = CTLANCTO.CODI_EMP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE FROM BETHADBA.CTSELECAO_SCP AS SEL WHERE SEL.CODI_EMP = CTLANCTO.CODI_EMP AND SEL.USUARIO = CURRENT USER ) AS TD_SELECAO_SCP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE FROM BETHADBA.CTSELECAO_SCP AS SEL WHERE SEL.CODI_EMP = CTLANCTO.CODI_EMP AND SEL.USUARIO = CURRENT USER AND SEL.FILIAL IS NULL AND SEL.I_SCP IS NULL ) AS TD_SELECAO_SEM_SCP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE , COALESCE ( MAX ( I_SCP ) , 0 ) AS I_SCP FROM BETHADBA.CTLANCTO_SCP AS LAN WHERE LAN.CODI_EMP = CTLANCTO.CODI_EMP AND LAN.NUME_LAN = CTLANCTO.NUME_LAN AND LAN.CODI_LOTE = CTLANCTO.CODI_LOTE AND LAN.FILIAL = CTLANCTO.FILI_LAN AND LAN.TIPO = 'D' ) AS TD_LANCTO_POSSUI_SCP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE FROM BETHADBA.CTLANCTO_SCP AS LAN INNER JOIN BETHADBA.CTSELECAO_SCP AS SEL ON ( SEL.CODI_EMP = LAN.CODI_EMP AND SEL.I_SCP = LAN.I_SCP AND SEL.USUARIO = CURRENT USER ) WHERE LAN.CODI_EMP = CTLANCTO.CODI_EMP AND LAN.NUME_LAN = CTLANCTO.NUME_LAN AND LAN.CODI_LOTE = CTLANCTO.CODI_LOTE AND LAN.FILIAL = CTLANCTO.FILI_LAN AND LAN.TIPO = 'D' ) AS TD_SCP_SELECIONADO , 
                LATERAL ( SELECT COALESCE ( COUNT ( *) , 0 ) AS EXISTE FROM BETHADBA.CTLANCTO_SCP_AUX AS CTA WHERE CTA.CODI_EMP = CTLANCTO.CODI_EMP AND CTA.NUME_LAN = CTLANCTO.NUME_LAN AND CTA.CODI_LOTE = COALESCE ( CTLANCTO.CODI_LOTE , 0 ) AND CTA.FILI_LAN = CTLANCTO.FILI_LAN AND CTA.TIPO = 'D' ) AS TD_EXISTE_LANCTO_AUX , 
                LATERAL ( SELECT COUNT ( *) AS EXISTE FROM BETHADBA.GESCP INNER JOIN BETHADBA.GEEMPRE AS GEEMPRE ON ( GEEMPRE.CODI_EMP = GESCP.CODI_EMP ) WHERE GESCP.CODI_EMP IN ( ( SELECT ( CODI_EMP ) FROM BETHADBA.GEEMPRE WHERE LEFT ( GEEMPRE.CGCE_EMP , 8 ) = '{cnpj_raiz}' ) ) AND ( ( GESCP.SITUACAO = 1 ) OR ( SITUACAO = 2 AND DATA_INATIVO > '{data_inicial}' ) ) AND GESCP.I_SCP = TD_LANCTO_POSSUI_SCP.I_SCP ) AS TDSCP 
                WHERE CTLANCTO.CODI_EMP = {codi_emp} AND ( ( LEFT ( CTCONTAS.CLAS_CTA , LENGTH ( CTA.CLAS_CTA ) ) = ( LEFT ( CTA.CLAS_CTA , LENGTH ( CTA.CLAS_CTA ) ) ) AND 0 = 1 ) OR 0 = 0 ) AND CTLANCTO.DATA_LAN <= '{data_final}' AND TD_EXISTE_LANCTO_AUX.EXISTE = 0 AND ( ( TD_SELECAO_SCP.EXISTE = 0 ) OR ( TD_SELECAO_SCP.EXISTE > 0 AND ( ( ( TD_SELECAO_SEM_SCP.EXISTE = 1 AND TD_LANCTO_POSSUI_SCP.EXISTE = 0 ) OR TD_SCP_SELECIONADO.EXISTE = 1 ) OR ( TD_SELECAO_SEM_SCP.EXISTE = 0 AND TD_LANCTO_POSSUI_SCP.EXISTE = 1 AND TD_SCP_SELECIONADO.EXISTE = 1 ) ) ) ) AND ( TDSCP.EXISTE > 0 OR TD_LANCTO_POSSUI_SCP.EXISTE = 0 ) 
                UNION ALL 
                SELECT CTLANCTO.CODI_EMP , CTLANCTO.NUME_LAN , COALESCE ( CTLANCTO.CODI_LOTE , 0 ) , CTLANCTO.FILI_LAN , 'C' AS TIPO 
                FROM BETHADBA.CTLANCTO AS CTLANCTO 
                INNER JOIN BETHADBA.CTCONTAS AS CTCONTAS ON CTCONTAS.CODI_EMP = {codi_emp} AND CTCONTAS.CODI_CTA = CTLANCTO.CCRE_LAN 
                LEFT OUTER JOIN BETHADBA.CTLANCTO_SCP_AUX_CTA AS CTA ON CTA.CODI_EMP = CTLANCTO.CODI_EMP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE FROM BETHADBA.CTSELECAO_SCP AS SEL WHERE SEL.CODI_EMP = CTLANCTO.CODI_EMP AND SEL.USUARIO = CURRENT USER ) AS TD_SELECAO_SCP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE FROM BETHADBA.CTSELECAO_SCP AS SEL WHERE SEL.CODI_EMP = CTLANCTO.CODI_EMP AND SEL.USUARIO = CURRENT USER AND SEL.FILIAL IS NULL AND SEL.I_SCP IS NULL ) AS TD_SELECAO_SEM_SCP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE , COALESCE ( MAX ( I_SCP ) , 0 ) AS I_SCP FROM BETHADBA.CTLANCTO_SCP AS LAN WHERE LAN.CODI_EMP = CTLANCTO.CODI_EMP AND LAN.NUME_LAN = CTLANCTO.NUME_LAN AND LAN.CODI_LOTE = CTLANCTO.CODI_LOTE AND LAN.FILIAL = CTLANCTO.FILI_LAN AND LAN.TIPO = 'C' ) AS TD_LANCTO_POSSUI_SCP , 
                LATERAL ( SELECT COUNT ( 1 ) AS EXISTE FROM BETHADBA.CTLANCTO_SCP AS LAN INNER JOIN BETHADBA.CTSELECAO_SCP AS SEL ON ( SEL.CODI_EMP = LAN.CODI_EMP AND SEL.I_SCP = LAN.I_SCP AND SEL.USUARIO = CURRENT USER ) WHERE LAN.CODI_EMP = CTLANCTO.CODI_EMP AND LAN.NUME_LAN = CTLANCTO.NUME_LAN AND LAN.CODI_LOTE = CTLANCTO.CODI_LOTE AND LAN.FILIAL = CTLANCTO.FILI_LAN AND LAN.TIPO = 'C' ) AS TD_SCP_SELECIONADO , 
                LATERAL ( SELECT COALESCE ( COUNT ( *) , 0 ) AS EXISTE FROM BETHADBA.CTLANCTO_SCP_AUX AS CTA WHERE CTA.CODI_EMP = CTLANCTO.CODI_EMP AND CTA.NUME_LAN = CTLANCTO.NUME_LAN AND CTA.CODI_LOTE = COALESCE ( CTLANCTO.CODI_LOTE , 0 ) AND CTA.FILI_LAN = CTLANCTO.FILI_LAN AND CTA.TIPO = 'C' ) AS TD_EXISTE_LANCTO_AUX , 
                LATERAL ( SELECT COUNT ( *) AS EXISTE FROM BETHADBA.GESCP INNER JOIN BETHADBA.GEEMPRE AS GEEMPRE ON ( GEEMPRE.CODI_EMP = GESCP.CODI_EMP ) WHERE GESCP.CODI_EMP IN ( ( SELECT ( CODI_EMP ) FROM BETHADBA.GEEMPRE WHERE LEFT ( GEEMPRE.CGCE_EMP , 8 ) = '{cnpj_raiz}' ) ) AND ( ( GESCP.SITUACAO = 1 ) OR ( SITUACAO = 2 AND DATA_INATIVO > '{data_inicial}' ) ) AND GESCP.I_SCP = TD_LANCTO_POSSUI_SCP.I_SCP ) AS TDSCP 
                WHERE CTLANCTO.CODI_EMP = {codi_emp} AND ( ( LEFT ( CTCONTAS.CLAS_CTA , LENGTH ( CTA.CLAS_CTA ) ) = ( LEFT ( CTA.CLAS_CTA , LENGTH ( CTA.CLAS_CTA ) ) ) AND 0 = 1 ) OR 0 = 0 ) AND CTLANCTO.DATA_LAN <= '{data_final}' AND TD_EXISTE_LANCTO_AUX.EXISTE = 0 AND ( TD_SELECAO_SCP.EXISTE = 0 OR ( TD_SELECAO_SCP.EXISTE > 0 AND ( ( ( TD_SELECAO_SEM_SCP.EXISTE = 1 AND TD_LANCTO_POSSUI_SCP.EXISTE = 0 ) OR TD_SCP_SELECIONADO.EXISTE = 1 ) OR ( TD_SELECAO_SEM_SCP.EXISTE = 0 AND TD_LANCTO_POSSUI_SCP.EXISTE = 1 AND TD_SCP_SELECIONADO.EXISTE = 1 ) ) ) ) AND ( TDSCP.EXISTE > 0 OR TD_LANCTO_POSSUI_SCP.EXISTE = 0 ) 
            )
            """
            cursor.execute(insert_aux_2)
            
            conn.commit()
            print("  -> INSERTs auxiliares executados com sucesso.")

        print("\n[PASSO 5 de 7] Limpando tabela auxiliar e configurando datas...")
        cursor.execute("DELETE FROM BETHADBA.CTINCORPORACAO_RET_AUX")
        conn.commit()

        print("\n[PASSO 6 de 7] Executando a consulta principal do Razão...")
        dt_ini_sql = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%Y%m%d")
        dt_fin_sql = datetime.strptime(data_final, "%Y-%m-%d").strftime("%Y%m%d")
        
        # O filtro de filial na query principal deve agora considerar todas as filiais do grupo
        filial_filter_sql = ""
        # Bloco CORRIGIDO para PRESERVAR A ORDEM E DUPLICAR
        if filiais and 'filiais_encontradas' in locals() and filiais_encontradas:
            # 1. Pega a lista de filiais exatamente na ordem que o banco retornou
            codigos_filiais_ordenados = [str(f.CODI_EMP) for f in filiais_encontradas]
            
            # 2. Adiciona a empresa matriz ao final da primeira lista
            codigos_filiais_ordenados.append(str(codi_emp))
            
            # 3. Duplica a lista inteira, preservando a ordem
            lista_duplicada = codigos_filiais_ordenados + codigos_filiais_ordenados
            
            filial_list_str = ", ".join(lista_duplicada)
            filial_filter_sql = f"CTLANCTO.FILI_LAN IN ({filial_list_str})"
            print(f"\n  -> Filtro de filiais gerado (com duplicação e ordem preservada): {filial_filter_sql}") # Linha de debug
        else: # Caso não use filiais ou não encontre nenhuma
            filial_filter_sql = f"CTLANCTO.FILI_LAN = {codi_emp}"
        
        query_razao = SQL_RAZAO_TEMPLATE.format(
            codi_emp=codi_emp,
            data_inicial_sql=dt_ini_sql,
            data_final_sql=dt_fin_sql,
            filial_filter=filial_filter_sql
        )
        
        pyperclip.copy(query_razao)
        df = pd.read_sql(query_razao, conn)
        
        if df.empty:
            print("  -> Consulta não retornou dados. Fim do fluxo.")
            return

        print(f"  -> SUCESSO! {len(df)} registros encontrados.")
        
        dump_filename = f"raw_dump_emp_{codi_emp}_{data_inicial}.json"
        print(f"  -> Salvando dados brutos em '{dump_filename}' para análise...")
        df.to_json(dump_filename, orient='records', indent=4, force_ascii=False, default_handler=str)
        print("  -> Dump JSON salvo com sucesso.")
        
        print("\n[PASSO 7 de 7] Gerando arquivo XLSX...")
        df.loc[df['FILIAL'] == 0, 'FILIAL'] = df['CODI_EMP']
        df['DATALAN'] = pd.to_datetime(df['DATALAN'])
        
        empresa_info['periodo'] = f"{datetime.strptime(data_inicial, '%Y-%m-%d').strftime('%d/%m/%Y')} - {datetime.strptime(data_final, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        
        output_xlsx = f"razao_emp_{codi_emp}_{data_inicial}_a_{data_final}.xlsx"
        format_excel_report(df, output_xlsx, empresa_info)
        print(f"  -> Arquivo XLSX '{output_xlsx}' gerado.")

    except Exception as e:
        print(f"\nERRO INESPERADO: {e}")
    finally:
        if conn:
            if filiais:
                print("\nLimpando permissões temporárias de usuário...")
                try:
                    cursor = conn.cursor()
                    #cursor.execute("DELETE FROM bethadba.ctfilialuser WHERE codi_emp = ? AND usuario = CURRENT USER", codi_emp)
                    conn.commit()
                    print("  -> Permissões limpas com sucesso.")
                except Exception as cleanup_error:
                    print(f"  -> ERRO ao limpar permissões: {cleanup_error}")
            
            conn.close()
            print(f"\nConexão com o banco de dados fechada.\n{'='*60}\n")

if __name__ == "__main__":

    gerar_relatorio_razao_com_dump(
        codi_emp=2026,
        data_inicial="2025-06-01",
        data_final="2025-06-05",
        filiais=True
    )
