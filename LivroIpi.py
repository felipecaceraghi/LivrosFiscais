import pyodbc
from fpdf import FPDF
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# --- CONFIGURAÇÃO PADRÃO PARA EXECUÇÃO DIRETA ---
# Estes valores são usados apenas quando o script é executado diretamente.
# A função `gerarLivroDeIpi` pode ser chamada com outros valores.
# ==============================================================================
CODI_EMP_PADRAO = 2493
DATA_INICIAL_PADRAO = '2025-05-01'
DATA_FINAL_PADRAO = '2025-05-31'
# ==============================================================================

# --- Constantes de Conexão ---
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

# --- Templates das Consultas SQL com Placeholders '?' ---
# COLE SUAS CONSULTAS SQL AQUI DENTRO DAS ASPAS TRIPLAS
# ------------------------------------------------------------------------------
SQL_TEMPLATES = {
    "23": "SELECT E.CGCE_EMP FROM BETHADBA.GEEMPRE AS E WHERE E.CODI_EMP = ?",
    "56": "SELECT COALESCE ( DSDBA.S_BUSCA_ALTERACAO_EMPRESA ( ? , ? , 2 ) , 'GREENV MOBILIDADE ELETRICA BRASIL S.A' ) FROM DSDBA.DUMMY",
    "57": "SELECT COALESCE ( DSDBA.S_BUSCA_ALTERACAO_EMPRESA ( ? , ? , 12 ) , '145.854.808.110' ) FROM DSDBA.DUMMY",
    "53": """
        /* Entradas normal */ SELECT 1 AS TIPO, NOTA.CODI_NAT AS CFOP, SUM(TD_VALORES.VALOR_CONTABIL) AS VCON, SUM(TD_VALORES.BASE_CALCULO) AS BASE, SUM(TD_VALORES.VALOR_IMPOSTO) AS VIMP, SUM(TD_VALORES.ISENTAS) AS VISE, SUM(TD_VALORES.OUTRAS) AS VOUT FROM BETHADBA.EFENTRADAS AS NOTA INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS PARAMETRO_VIGENCIA ON PARAMETRO_VIGENCIA.CODI_EMP = NOTA.CODI_EMP, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS TEM_IMPOSTO, COALESCE(SUM(IMPOSTO_NOTA.BCAL_IEN), 0) AS BASE_CALCULO, COALESCE(SUM(IMPOSTO_NOTA.VLOR_IEN), 0) AS VALOR_IMPOSTO, COALESCE(SUM(IMPOSTO_NOTA.VISE_IEN), 0) AS ISENTAS, COALESCE(SUM(IMPOSTO_NOTA.VOUT_IEN), 0) AS OUTRAS FROM BETHADBA.EFIMPENT AS IMPOSTO_NOTA WHERE IMPOSTO_NOTA.CODI_EMP = NOTA.CODI_EMP AND IMPOSTO_NOTA.CODI_ENT = NOTA.CODI_ENT AND IMPOSTO_NOTA.CODI_IMP = 2 AND (IMPOSTO_NOTA.I_RECOLHIMENTO = 1 OR 1 = 0)) AS TD_IMPOSTO, LATERAL(SELECT (CASE WHEN NOTA.CODI_NAT IN (1111, 1113, 2111, 2113) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE NOTA.VCON_ENT END) AS VALOR_CONTABIL, (CASE WHEN NOTA.CODI_NAT IN (1111, 1113, 2111, 2113) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.BASE_CALCULO END) AS BASE_CALCULO, (CASE WHEN NOTA.CODI_NAT IN (1111, 1113, 2111, 2113) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.VALOR_IMPOSTO END) AS VALOR_IMPOSTO, (CASE WHEN NOTA.CODI_NAT IN (1111, 1113, 2111, 2113) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.ISENTAS END) AS ISENTAS, (CASE WHEN NOTA.CODI_NAT IN (1111, 1113, 2111, 2113) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.OUTRAS END) AS OUTRAS FROM DSDBA.DUMMY) AS TD_VALORES WHERE NOTA.CODI_EMP = ? AND NOTA.DENT_ENT BETWEEN ? AND ? AND PARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(NOTA.CODI_EMP, NOTA.DENT_ENT) AND TD_IMPOSTO.TEM_IMPOSTO = 'S' GROUP BY 1, NOTA.CODI_NAT
        UNION ALL
        /* Entradas Simples Nacional */ SELECT 1 AS TIPO, NOTA.CODI_NAT AS CFOP, SUM(NOTA.VCON_ENT) AS VCON, 0 AS BASE, 0 AS VIMP, 0 AS VISE, SUM(NOTA.VCON_ENT) AS VOUT FROM BETHADBA.EFENTRADAS AS NOTA INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS PARAMETRO_VIGENCIA ON PARAMETRO_VIGENCIA.CODI_EMP = NOTA.CODI_EMP, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS TEM_IMPOSTO FROM BETHADBA.EFIMPENT AS IMPOSTO_NOTA WHERE IMPOSTO_NOTA.CODI_EMP = NOTA.CODI_EMP AND IMPOSTO_NOTA.CODI_ENT = NOTA.CODI_ENT AND IMPOSTO_NOTA.CODI_IMP = 2 AND (IMPOSTO_NOTA.I_RECOLHIMENTO = 1 OR 1 = 0)) AS TD_IMPOSTO, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS TEM_SIMPLES_NACIONAL FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS ACUMULADOR_VIGENCIA_IMPOSTOS WHERE ACUMULADOR_VIGENCIA_IMPOSTOS.CODI_EMP = NOTA.CODI_EMP AND ACUMULADOR_VIGENCIA_IMPOSTOS.CODI_ACU = NOTA.CODI_ACU AND ACUMULADOR_VIGENCIA_IMPOSTOS.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(NOTA.CODI_EMP, NOTA.CODI_ACU, NOTA.DENT_ENT) AND ACUMULADOR_VIGENCIA_IMPOSTOS.CODI_IMP = 44 AND ACUMULADOR_VIGENCIA_IMPOSTOS.SIMPLESN_ANEXO_IAC = 2) AS TD_IMPOSTO_ACUMULADOR WHERE NOTA.CODI_EMP = ? AND NOTA.DENT_ENT BETWEEN ? AND ? AND PARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(NOTA.CODI_EMP, NOTA.DENT_ENT) AND 'N' = 'S' AND TD_IMPOSTO_ACUMULADOR.TEM_SIMPLES_NACIONAL = 'S' AND TD_IMPOSTO.TEM_IMPOSTO = 'N' GROUP BY 1, NOTA.CODI_NAT
        UNION ALL
        /* Saídas normal */ SELECT 2 AS TIPO, NOTA.CODI_NAT AS CFOP, SUM(TD_VALORES.VALOR_CONTABIL) AS VCON, SUM(TD_VALORES.BASE_CALCULO) AS BASE, SUM(TD_VALORES.VALOR_IMPOSTO) AS VIMP, SUM(TD_VALORES.ISENTAS) AS VISE, SUM(TD_VALORES.OUTRAS) AS VOUT FROM BETHADBA.EFSAIDAS AS NOTA INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS PARAMETRO_VIGENCIA ON PARAMETRO_VIGENCIA.CODI_EMP = NOTA.CODI_EMP, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS TEM_IMPOSTO, COALESCE(SUM(IMPOSTO_NOTA.BCAL_ISA), 0) AS BASE_CALCULO, COALESCE(SUM(IMPOSTO_NOTA.VLOR_ISA), 0) AS VALOR_IMPOSTO, COALESCE(SUM(IMPOSTO_NOTA.VISE_ISA), 0) AS ISENTAS, COALESCE(SUM(IMPOSTO_NOTA.VOUT_ISA), 0) AS OUTRAS FROM BETHADBA.EFIMPSAI AS IMPOSTO_NOTA WHERE IMPOSTO_NOTA.CODI_EMP = NOTA.CODI_EMP AND IMPOSTO_NOTA.CODI_SAI = NOTA.CODI_SAI AND IMPOSTO_NOTA.CODI_IMP = 2 AND (IMPOSTO_NOTA.I_RECOLHimento = 1 OR 1 = 0)) AS TD_IMPOSTO, LATERAL(SELECT (CASE WHEN NOTA.CODI_NAT IN (5111, 5112, 5113, 5114, 6111, 6112, 6113, 6114) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE NOTA.VCON_SAI END) AS VALOR_CONTABIL, (CASE WHEN NOTA.CODI_NAT IN (5111, 5112, 5113, 5114, 6111, 6112, 6113, 6114) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.BASE_CALCULO END) AS BASE_CALCULO, (CASE WHEN NOTA.CODI_NAT IN (5111, 5112, 5113, 5114, 6111, 6112, 6113, 6114) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.VALOR_IMPOSTO END) AS VALOR_IMPOSTO, (CASE WHEN NOTA.CODI_NAT IN (5111, 5112, 5113, 5114, 6111, 6112, 6113, 6114) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.ISENTAS END) AS ISENTAS, (CASE WHEN NOTA.CODI_NAT IN (5111, 5112, 5113, 5114, 6111, 6112, 6113, 6114) AND PARAMETRO_VIGENCIA.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S' THEN 0 ELSE TD_IMPOSTO.OUTRAS END) AS OUTRAS FROM DSDBA.DUMMY) AS TD_VALORES WHERE NOTA.CODI_EMP = ? AND NOTA.DSAI_SAI BETWEEN ? AND ? AND PARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(NOTA.CODI_EMP, NOTA.DSAI_SAI) AND TD_IMPOSTO.TEM_IMPOSTO = 'S' GROUP BY 2, NOTA.CODI_NAT
        UNION ALL
        /* Saídas Simples Nacional */ SELECT 2 AS TIPO, NOTA.CODI_NAT AS CFOP, SUM(TDAUX.VALOR) AS VCON, 0 AS BCAL, 0 AS VLOR, 0 AS VISE, SUM(TDAUX.VALOR) AS VOUT FROM BETHADBA.EFSAIDAS AS NOTA INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS PARAMETRO_VIGENCIA ON PARAMETRO_VIGENCIA.CODI_EMP = NOTA.CODI_EMP, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS TEM_IMPOSTO FROM BETHADBA.EFIMPSAI AS IMPOSTO_NOTA WHERE IMPOSTO_NOTA.CODI_EMP = NOTA.CODI_EMP AND IMPOSTO_NOTA.CODI_SAI = NOTA.CODI_SAI AND IMPOSTO_NOTA.CODI_IMP = 2 AND (IMPOSTO_NOTA.I_RECOLHIMENTO = 1 OR 1 = 0)) AS TD_IMPOSTO, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS TEM_SIMPLES_NACIONAL FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS ACUMULADOR_VIGENCIA_IMPOSTOS WHERE ACUMULADOR_VIGENCIA_IMPOSTOS.CODI_EMP = NOTA.CODI_EMP AND ACUMULADOR_VIGENCIA_IMPOSTOS.CODI_ACU = NOTA.CODI_ACU AND ACUMULADOR_VIGENCIA_IMPOSTOS.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(NOTA.CODI_EMP, NOTA.CODI_ACU, NOTA.DSAI_SAI) AND ACUMULADOR_VIGENCIA_IMPOSTOS.CODI_IMP = 44 AND ACUMULADOR_VIGENCIA_IMPOSTOS.SIMPLESN_ANEXO_IAC = 2) AS TD_IMPOSTO_ACUMULADOR, LATERAL(SELECT COALESCE(SUM(ECFA.VLOR_ECA), 0) AS VALOR FROM BETHADBA.EFTABECFM AS ECFM INNER JOIN BETHADBA.EFTABECFA AS ECFA ON ECFA.CODI_EMP = ECFM.CODI_EMP AND ECFA.CODI_ECM = ECFM.CODI_ECM AND ECFA.CODI_MEC = ECFM.CODI_MEC AND ECFA.TIPO_ECM = ECFM.TIPO_ECM WHERE ECFM.CODI_EMP = NOTA.CODI_EMP AND ECFM.CODI_ECM = NOTA.CODI_SAI AND ECFM.TIPO_ECM = 'S' AND ECFM.DEDUZIR_ECM = 'S' AND ECFA.SITU_ECA IN ('CANC', 'DESC')) AS TDECFM_ECFA, LATERAL(SELECT NOTA.VCON_SAI - TDECFM_ECFA.VALOR AS VALOR FROM DSDBA.DUMMY) AS TDAUX WHERE NOTA.CODI_EMP = ? AND NOTA.DSAI_SAI BETWEEN ? AND ? AND PARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(NOTA.CODI_EMP, NOTA.DSAI_SAI) AND 'N' = 'S' AND TD_IMPOSTO_ACUMULADOR.TEM_SIMPLES_NACIONAL = 'S' AND TD_IMPOSTO.TEM_IMPOSTO = 'N' AND NOTA.SITUACAO_SAI NOT IN (2, 3) GROUP BY NOTA.CODI_NAT ORDER BY 1, 2
    """,
    "58": """
        SELECT YEAR ( S.DATA_SIM ) * 100 + MONTH ( S.DATA_SIM ) AS ANOMES, S.VIMS_SIM AS SAIDAS, S.OUTD_SIM AS OUTROSD, S.ESTC_SIM + S.VDI9_SIM AS ESTORNOC, S.VIMS_SIM + S.OUTD_SIM + S.ESTC_SIM + S.VDI9_SIM AS SUBTOTALD, S.VIME_SIM AS ENTRADAS, S.OUTC_SIM + S.VDI7_SIM AS OUTROSC, S.ESTD_SIM AS ESTORNOD, (S.VIME_SIM + S.OUTC_SIM + S.ESTD_SIM + S.VDI7_SIM) AS SUBTOTALC, 0 AS SALDOANT, S.VIME_SIM + S.OUTC_SIM + S.ESTD_SIM + S.VDI7_SIM AS TOTAL, S.SREC_SIM AS DEVEDOR, S.ODED_SIM AS DEDUCOES, S.SDEV_SIM AS RECOLHER, S.SCRE_SIM AS CREDOR, G.ESTA_EMP AS ESTADO FROM BETHADBA.EFSDOIMP AS S INNER JOIN BETHADBA.GEEMPRE AS G ON S.CODI_EMP = G.CODI_EMP, LATERAL(SELECT COUNT(*) AS QUANTIDADE FROM BETHADBA.EFSDOIMP_POR_RECOLHIMENTO AS SALDO_POR_RECOLHIMENTO WHERE SALDO_POR_RECOLHIMENTO.CODI_EMP = S.CODI_EMP AND SALDO_POR_RECOLHIMENTO.DATA_SIM = S.DATA_SIM AND SALDO_POR_RECOLHIMENTO.CODI_IMP = S.CODI_IMP AND SALDO_POR_RECOLHIMENTO.PDIC_SIM = S.PDIC_SIM) AS TD_SALDO_POR_RECOLHIMENTO WHERE S.CODI_EMP = ? AND S.DATA_SIM = ? AND S.CODI_IMP = 2 AND S.PDIC_SIM = 6 AND 'N' = 'N' AND TD_SALDO_POR_RECOLHIMENTO.QUANTIDADE = 0
        UNION ALL
        SELECT YEAR ( S.DATA_SIM ) * 100 + MONTH ( S.DATA_SIM ) AS ANOMES, S.VALOR_IMPOSTO_SAIDA AS SAIDAS, S.OUTROS_DEBITOS + S.RESSARCIMENTO_COMPENSACAO_PIS_COFINS_DEDUCOES AS OUTROSD, S.ESTORNO_CREDITOS + S.VALOR_IPI_DEVOLUCAO_DEBITO AS ESTORNOC, S.VALOR_IMPOSTO_SAIDA + S.OUTROS_DEBITOS + S.ESTORNO_CREDITOS + S.RESSARCIMENTO_COMPENSACAO_PIS_COFINS_DEDUCOES + S.VALOR_IPI_DEVOLUCAO_DEBITO AS SUBTOTALD, S.VALOR_IMPOSTO_ENTRADA AS ENTRADAS, S.OUTROS_CREDITOS + S.VALOR_IPI_DEVOLUCAO AS OUTROSC, S.ESTORNO_DEBITOS AS ESTORNOD, S.VALOR_IMPOSTO_ENTRADA + S.OUTROS_CREDITOS + S.ESTORNO_DEBITOS + S.VALOR_IPI_DEVOLUCAO AS SUBTOTALC, S.SALDO_CREDOR_ANTERIOR AS SALDOANT, (S.VALOR_IMPOSTO_ENTRADA + S.OUTROS_CREDITOS + S.ESTORNO_DEBITOS + S.VALOR_IPI_DEVOLUCAO) AS TOTAL, S.SALDO_DEVEDOR AS DEVEDOR, S.OUTRAS_DEDUCOES + S.RESSARCIMENTO_COMPENSACAO_PIS_COFINS AS DEDUCOES, S.SALDO_RECOLHER AS RECOLHER, S.SALDO_CREDOR AS CREDOR, G.ESTA_EMP AS ESTADO FROM BETHADBA.EFSDOIMP_POR_RECOLHIMENTO AS S INNER JOIN BETHADBA.GEEMPRE AS G ON S.CODI_EMP = G.CODI_EMP WHERE S.CODI_EMP = ? AND S.DATA_SIM = ? AND S.CODI_IMP = 2 AND S.PDIC_SIM = 6 AND 'N' = 'N' AND (S.I_RECOLHIMENTO = 1 OR 1 = 0)
        UNION ALL
        SELECT YEAR ( ? ) * 100 + MONTH ( ? ) AS ANOMES, 0 AS SAIDAS, 0 AS OUTROSD, 0 AS ESTORNOC, 0 AS SUBTOTALD, 0 AS ENTRADAS, 0 AS OUTROSC, 0 AS ESTORNOD, 0 AS SUBTOTALC, 0 AS SALDOANT, 0 AS TOTAL, 0 AS DEVEDOR, 0 AS DEDUCOES, 0 AS RECOLHER, 0 AS CREDOR, G.ESTA_EMP AS ESTADO FROM BETHADBA.GEEMPRE G WHERE G.CODI_EMP = ? AND 'N' = 'S'
    """,
    "65": """
        SELECT /*[1]*/10 AS CODI/*[1]*/, /*[2]*/TDAUX.DESCRICAO AS NOME/*[2]*/, SUM ( M.VALOR ) AS VLOR FROM BETHADBA.EFMOVIMENTO_AJUSTE_FEDERAL M, BETHADBA.EFAJUSTES A, LATERAL(SELECT (CASE WHEN COALESCE(TRIM(M.OBSERVACAO), '') <> '' THEN COALESCE(TRIM(M.OBSERVACAO), '') ELSE COALESCE(TRIM(A.NOME_AJU), '') END) AS DESCRICAO FROM DSDBA.DUMMY) AS TDAUX WHERE M.CODI_EMP = ? AND M.DATA_AJUSTE >= ? AND M.DATA_AJUSTE <= ? AND M.CODI_EMP = A.CODI_EMP AND M.CODI_AJU = A.CODI_AJU AND A.CODI_IMP = 2 AND A.OPER_AJU = 1 AND (M.I_RECOLHIMENTO = 1 OR 1 = 0) GROUP BY /*[1]*/10/*[1]*/, /*[2]*/TDAUX.DESCRICAO/*[2]*/
        UNION ALL
        SELECT /*[1]*/10 AS CODI/*[1]*/, /*[2]*/'IPI devolução' AS NOME/*[2]*/, TDVALOR_IMPOSTO.VALOR_IMPOSTO AS VLOR FROM BETHADBA.EFSDOIMP AS IMPOSTO, LATERAL(SELECT SUM(COALESCE(VALOR_IPI_DEVOLUCAO, 0)) AS TOTAL FROM BETHADBA.EFSDOIMP_POR_RECOLHIMENTO AS IMPOSTO_RECOLHIMENTO WHERE IMPOSTO_RECOLHIMENTO.CODI_EMP = IMPOSTO.CODI_EMP AND IMPOSTO_RECOLHIMENTO.CODI_IMP = IMPOSTO.CODI_IMP AND IMPOSTO_RECOLHIMENTO.DATA_SIM = IMPOSTO.DATA_SIM) AS TDIMPOSTO_RECOLHIDO, LATERAL(SELECT CASE WHEN TDIMPOSTO_RECOLHIDO.TOTAL > 0 THEN TDIMPOSTO_RECOLHIDO.TOTAL ELSE COALESCE(IMPOSTO.VDI7_SIM, 0) END AS VALOR_IMPOSTO FROM DSDBA.DUMMY) AS TDVALOR_IMPOSTO WHERE IMPOSTO.CODI_EMP = ? AND IMPOSTO.DATA_SIM >= ? AND IMPOSTO.DATA_SIM <= ? AND IMPOSTO.CODI_IMP = 2 AND TDVALOR_IMPOSTO.VALOR_IMPOSTO > 0 ORDER BY 1
    """
}
# ------------------------------------------------------------------------------


# --- Funções de Lógica e Dados (Inalteradas) ---
def format_currency(value):
    try:
        val = float(value)
        if val == 0: return "0,00"
        return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError): return "0,00"

def get_params_for_query(query_id, codi_emp, data_inicial, data_final):
    params_53 = (codi_emp, data_inicial, data_final) * 4
    params_58 = (codi_emp, data_inicial, codi_emp, data_inicial, data_inicial, data_inicial, codi_emp)
    params_65 = (codi_emp, data_inicial, data_final) * 2
    params_map = {"23": (codi_emp,), "56": (codi_emp, data_inicial), "57": (codi_emp, data_inicial),
                  "53": params_53, "58": params_58, "65": params_65}
    return params_map.get(query_id, ())

def fetch_data_from_db(conn_str, templates, codi_emp, data_inicial, data_final):
    query_map = {}
    conn = None
    try:
        print("Conectando ao banco de dados...")
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        print("Conexão bem-sucedida.")
        for query_id, sql_template in templates.items():
            if not sql_template.strip() or "Cole a consulta" in sql_template:
                print(f"AVISO: Consulta {query_id} está vazia. Pulando.")
                query_map[query_id] = []
                continue
            params = get_params_for_query(query_id, codi_emp, data_inicial, data_final)
            print(f"Executando consulta: {query_id}...")
            cursor.execute(sql_template, params)
            columns = [column[0] for column in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            query_map[query_id] = results
    except pyodbc.Error as ex:
        print(f"ERRO DE BANCO DE DADOS: {ex}")
        return {key: [] for key in templates.keys()}
    except Exception as e:
        print(f"ERRO INESPERADO: {e}")
        return {key: [] for key in templates.keys()}
    finally:
        if conn: conn.close(); print("Conexão com o banco de dados fechada.")
    return query_map

def extract_report_data(query_map, data_inicial, data_final):
    report_data = {"header": {}, "entradas": [], "saidas": [], "resumo": {}}
    res56 = query_map.get("56", [{}])[0]; report_data["header"]["empresa"] = list(res56.values())[0] if res56 else ''
    res23 = query_map.get("23", [{}])[0]; report_data["header"]["cnpj"] = res23.get("CGCE_EMP")
    res57 = query_map.get("57", [{}])[0]; report_data["header"]["insc_est"] = list(res57.values())[0] if res57 else ''
    data_ini_obj = datetime.strptime(data_inicial, '%Y-%m-%d'); data_fim_obj = datetime.strptime(data_final, '%Y-%m-%d')
    report_data["header"]["periodo"] = f"{data_ini_obj.strftime('%d/%m/%Y')} a {data_fim_obj.strftime('%d/%m/%Y')}"
    for mov in query_map.get("53", []):
        if mov.get("TIPO") == 1: report_data["entradas"].append(mov)
        elif mov.get("TIPO") == 2: report_data["saidas"].append(mov)
    resumo_raw = query_map.get("58", [{}])[0] if query_map.get("58") else {}
    total_outros_creditos, detalhes_outros_creditos = 0.0, []
    for item in query_map.get("65", []):
        valor_item = float(item.get("VLOR", 0) or 0)
        if valor_item > 0:
            total_outros_creditos += valor_item; detalhes_outros_creditos.append({"desc": item.get("NOME"), "valor": valor_item})
    if resumo_raw:
        credito_entradas = float(resumo_raw.get("ENTRADAS", 0) or 0); credito_outros = total_outros_creditos
        saldo_anterior = float(resumo_raw.get("SALDOANT", 0) or 0); credito_subtotal = credito_entradas + credito_outros
        report_data["resumo"] = {"debito_saidas": float(resumo_raw.get("SAIDAS", 0) or 0), "debito_outros": 0.0,
                                 "debito_estornos": float(resumo_raw.get("ESTORNOC", 0) or 0), "debito_subtotal": float(resumo_raw.get("SAIDAS", 0) or 0),
                                 "credito_entradas": credito_entradas, "credito_outros": credito_outros, "credito_outros_detalhes": detalhes_outros_creditos,
                                 "credito_estornos": float(resumo_raw.get("ESTORNOD", 0) or 0), "credito_subtotal": credito_subtotal, "saldo_anterior": saldo_anterior,
                                 "credito_total": credito_subtotal + saldo_anterior, "saldo_devedor": float(resumo_raw.get("DEVEDOR", 0) or 0),
                                 "deducoes": float(resumo_raw.get("DEDUCOES", 0) or 0), "imposto_recolher": float(resumo_raw.get("RECOLHER", 0) or 0),
                                 "saldo_credor_final": float(resumo_raw.get("CREDOR", 0) or 0)}
    return report_data


# --- Classe Geradora de PDF (Inalterada) ---
class PDF(FPDF):
    def __init__(self, data, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.report_data = data
        self.set_auto_page_break(auto=True, margin=15)
        self.set_margins(10, 10, 10)
        self.header_bg_color = (220, 220, 220)
        self.line_height = 6
    def draw_page_header(self, title):
        self.set_font("Courier", "B", 12); self.cell(0, 10, title, 0, 1, "C")
        self.set_font("Courier", "", 10); info = self.report_data["header"]
        self.cell(180, 5, f"EMPRESA : {info.get('empresa', '')}", 0, 0, "L"); self.cell(0, 5, f"CNPJ: {info.get('cnpj', '')}", 0, 1, "L")
        self.cell(180, 5, f"INSC.EST.: {info.get('insc_est', '')}", 0, 0, "L"); self.cell(0, 5, f"MES OU PERÍODO/ANO: {info.get('periodo', '')}", 0, 1, "L")
        self.cell(0, 5, f"FOLHA : {self.page_no():03d}", 0, 1, "L"); self.ln(5)
    def draw_movimento_table(self, title, data):
        col_widths = {'contabil': 25, 'fiscal': 25, 'v_contabeis': 40, 'base_calculo': 40, 'imposto': 35, 'isentas': 55, 'outras': 35}
        codif_width, ipi_width, total_width = col_widths['contabil'] + col_widths['fiscal'], sum(col_widths.values()) - (col_widths['contabil'] + col_widths['fiscal']), sum(col_widths.values())
        h, start_x = self.line_height, self.get_x()
        self.set_font('Courier', 'B', 12); self.cell(total_width, 8, title, 1, 1, 'C')
        y_level2 = self.get_y(); self.set_font('Courier', 'B', 9)
        self.cell(codif_width, h, 'CODIFICAÇÃO', 1, 0, 'C'); self.cell(ipi_width, h, 'I P I - VALORES FISCAIS', 1, 1, 'C')
        y_level3 = self.get_y(); op_title = f"OPERAÇÃO COM {'CRÉDITO' if title == 'ENTRADAS' else 'DÉBITO'} DO IMPOSTO"; op_sem_title = f"OPERAÇÃO SEM {'CRÉDITO' if title == 'ENTRADAS' else 'DÉBITO'} DO IMPOSTO"
        self.set_x(start_x + codif_width); self.cell(col_widths['v_contabeis'], h * 2, '', 1, 0, 'C'); self.cell(col_widths['base_calculo'] + col_widths['imposto'], h, op_title, 1, 0, 'C'); self.cell(col_widths['isentas'] + col_widths['outras'], h, op_sem_title, 1, 1, 'C')
        self.set_font('Courier', 'B', 8); self.set_xy(start_x, y_level3); self.cell(col_widths['contabil'], h, 'CONTÁBIL', 1, 0, 'C'); self.cell(col_widths['fiscal'], h, 'FISCAL', 1, 0, 'C')
        self.set_xy(start_x + codif_width, y_level3); self.multi_cell(col_widths['v_contabeis'], h, 'VALORES\nCONTÁBEIS', 0, 'C')
        self.set_xy(start_x + codif_width + col_widths['v_contabeis'], y_level3 + h); self.cell(col_widths['base_calculo'], h, 'BASE DE CÁLCULO', 1, 0, 'C'); self.cell(col_widths['imposto'], h, 'IMPOSTO', 1, 0, 'C'); self.cell(col_widths['isentas'], h, 'ISENTAS OU NÃO TRIBUTADAS', 1, 0, 'C'); self.cell(col_widths['outras'], h, 'OUTRAS', 1, 1, 'C')
        self.set_font("Courier", "", 9)
        subtotals = {f'cat{i}': {k: 0.0 for k in ["vcon", "base", "vimp", "vise", "vout"]} for i in range(1, 4)}
        for row in data:
            self.cell(col_widths['contabil'], h, "", 1, 0, "C")
            self.cell(col_widths['fiscal'], h, str(row.get('CFOP', '')), 1, 0, "C")
            self.cell(col_widths['v_contabeis'], h, format_currency(row.get('VCON')), 1, 0, "R"); self.cell(col_widths['base_calculo'], h, format_currency(row.get('BASE')), 1, 0, "R"); self.cell(col_widths['imposto'], h, format_currency(row.get('VIMP')), 1, 0, "R"); self.cell(col_widths['isentas'], h, format_currency(row.get('VISE')), 1, 0, "R"); self.cell(col_widths['outras'], h, format_currency(row.get('VOUT')), 1, 1, "R")
            cfop_str = str(row.get('CFOP', '0')); cat = ''
            if cfop_str.startswith(('1', '5')): cat = 'cat1'
            elif cfop_str.startswith(('2', '6')): cat = 'cat2'
            elif cfop_str.startswith(('3', '7')): cat = 'cat3'
            if cat:
                for k in ["VCON", "BASE", "VIMP", "VISE", "VOUT"]: subtotals[cat][k.lower()] += float(row.get(k, 0) or 0)
        self.ln(2); self.set_font("Courier", "B", 9); self.cell(codif_width, h, f"SUBTOTAIS {title}", 0, 1, "L")
        subtotal_labels = {'ENTRADAS': ["1-000 DO ESTADO", "2-000 DE OUTROS ESTADOS", "3-000 DO EXTERIOR"], 'SAIDAS': ["5-000 PARA O ESTADO", "6-000 PARA OUTROS ESTADOS", "7-000 PARA O EXTERIOR"]}
        key_map = {'vcon': 'v_contabeis', 'base': 'base_calculo', 'vimp': 'imposto', 'vise': 'isentas', 'vout': 'outras'}
        for i, cat in enumerate(['cat1', 'cat2', 'cat3']):
            self.cell(codif_width, h, subtotal_labels[title][i], 1, 0, "L")
            for key in ["vcon", "base", "vimp", "vise", "vout"]: self.cell(col_widths[key_map[key]], h, format_currency(subtotals[cat][key]), 1, 1 if key == "vout" else 0, "R")
        totals = {k: sum(subtotals[cat][k] for cat in subtotals) for k in subtotals['cat1']}
        self.set_fill_color(*self.header_bg_color); self.cell(codif_width, h, "TOTAIS", 1, 0, "C", True)
        for key in ["vcon", "base", "vimp", "vise", "vout"]: self.cell(col_widths[key_map[key]], h, format_currency(totals[key]), 1, 1 if key == "vout" else 0, "R", True)
        self.ln(10)
    def draw_resumo_table(self):
        self.set_font("Courier", "B", 10); self.cell(0, 8, "R E S U M O   D A   A P U R A Ç Ã O   D O   I M P O S T O", 0, 1, "C"); self.ln(2)
        resumo, h = self.report_data["resumo"], 5; w_v, w_d, w_a, w_s = 6, 150, 40, 40; total_w, start_x = w_v + w_d + w_a + w_s, self.get_x()
        def draw_v_text(x, y, text, height):
            self.set_font("Courier", "B", 9); char_h = 3.5; start_y = y + (height / 2) - (len(text) * char_h / 2)
            for char in text: self.set_xy(x, start_y); self.cell(w_v, char_h, char, 0, 1, "C"); start_y += char_h
        def draw_section(title, rows, v_text):
            start_y, height = self.get_y(), (h * 2) + (len(rows) * h)
            self.set_font("Courier", "B", 9); self.set_xy(start_x + w_v, start_y); self.cell(w_d, h, title, 'B', 0, 'C'); self.cell(w_a + w_s, h, 'V A L O R E S', 'B', 1, 'C')
            self.set_xy(start_x + w_v, start_y + h); self.cell(w_d, h, '', 'B', 0); self.cell(w_a, h, 'COLUNA AUXILIAR', 'B', 0, 'C'); self.cell(w_s, h, 'SOMAS', 'B', 1, 'C')
            self.set_font("Courier", "", 9); self.set_y(start_y + h * 2)
            for desc, aux, somas in rows:
                self.set_x(start_x + w_v); self.cell(w_d, h, desc, 0, 0, 'L'); self.cell(w_a, h, format_currency(aux) if aux not in [None, ""] else '', 0, 0, 'R'); self.cell(w_s, h, format_currency(somas) if somas not in [None, ""] else '', 0, 1, 'R')
            self.rect(start_x, start_y, total_w, height); self.line(start_x + w_v, start_y, start_x + w_v, start_y + height); self.line(start_x + w_v + w_d, start_y, start_x + w_v + w_d, start_y + height); self.line(start_x + w_v + w_d + w_a, start_y + h, start_x + w_v + w_d + w_a, start_y + height); self.line(start_x + w_v + w_d, start_y + h, start_x + total_w, start_y + h)
            draw_v_text(start_x, start_y, v_text, height); self.set_y(start_y + height)
        debito_rows = [("001 - POR SAÍDAS/PRESTAÇÕES COM DÉBITO DO IMPOSTO", "", resumo.get("debito_saidas")), ("002 - OUTROS DÉBITOS(DISCRIMINAR ABAIXO)", "", resumo.get("debito_outros")), ("003 - ESTORNO DE CRÉDITOS(DISCRIMINAR ABAIXO)", "", resumo.get("debito_estornos")), ("004 - SUBTOTAL", "", resumo.get("debito_subtotal"))]; draw_section("D É B I T O   D O   I M P O S T O", debito_rows, "DÉBITO")
        credito_rows = [("005 - POR ENTRADAS/AQUISIÇÕES COM CRÉDITO DO IMPOSTO", "", resumo.get("credito_entradas")), ("006 - OUTROS CRÉDITOS(DISCRIMINAR ABAIXO)", "", resumo.get("credito_outros"))]
        for detalhe in resumo.get("credito_outros_detalhes", []): credito_rows.append((f"      ({detalhe['desc']})", detalhe['valor'], "")); 
        credito_rows.extend([("007 - ESTORNO DE DÉBITOS(DISCRIMINAR ABAIXO)", "", resumo.get("credito_estornos")), ("008 - SUBTOTAL", "", resumo.get("credito_subtotal")), ("009 - SALDO CREDOR DO PERÍODO ANTERIOR", "", resumo.get("saldo_anterior")), ("010 - TOTAL", "", resumo.get("credito_total"))]); draw_section("C R É D I T O   D O   I M P O S T O", credito_rows, "CRÉDITO")
        apuracao_rows = [("011 - SALDO DEVEDOR(DÉBITO - CRÉDITO)", "", resumo.get("saldo_devedor")), ("012 - DEDUÇÕES(DISCRIMINAR ABAIXO)", "", resumo.get("deducoes")), ("013 - IMPOSTO A RECOLHER", "", resumo.get("imposto_recolher")), ("014 - SALDO CREDOR(CRÉDITO - DÉBITO) A TRANSPORTAR PARA O PERÍODO SEGUINTE", "", resumo.get("saldo_credor_final"))]; draw_section("A P U R A Ç Ã O   D O   S A L D O", apuracao_rows, "SALDO")
        self.ln(5); self.set_font("Courier", "B", 9); self.cell(0, h, 'DISTRIBUIÇÃO DO SALDO DEVEDOR PÊLOS PRAZOS DE RECOLHIMENTO', 'T', 1, 'C'); self.cell(0, h, 'GUIA DE RECOLHIMENTO', 'B', 1, 'C')
        self.cell(40, h, 'PERÍODO', 'B', 0, 'L'); self.cell(40, h, 'DATA', 'B', 0, 'L'); self.cell(40, h, 'VALOR', 'B', 0, 'L'); self.cell(60, h, 'ORGÃO ARRECADADOR', 'B', 0, 'L'); self.cell(0, h, 'DATA DE ENTREGA LOCAL DE ENTREGA(BANCO/REPARTIÇÃO)', 'B', 1, 'L')


# --- CLASSE GERADORA DE XLSX (Inalterada) ---
class ExcelReport:
    def __init__(self, data, filename):
        self.report_data = data; self.filename = filename; self.workbook = openpyxl.Workbook(); self.workbook.remove(self.workbook.active)
        self.styles = self._create_styles()
    def _create_styles(self):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        return {"default_font": Font(name='Courier New', size=9), "header_main": Font(bold=True, size=12, name='Courier New'),
                "header_info": Font(bold=True, size=9, name='Courier New'), "header_table": Font(bold=True, size=9, name='Courier New'),
                "currency": 'R$ #,##0.00_-', "align_center": Alignment(horizontal='center', vertical='center', wrap_text=True),
                "align_right": Alignment(horizontal='right', vertical='center'), "align_left": Alignment(horizontal='left', vertical='center'),
                "gray_fill": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"), "thin_border": thin_border,
                "vertical_text": Alignment(text_rotation=90, horizontal='center', vertical='center')}
    def _set_cell(self, sheet, row, col, value, style_keys=[]):
        cell = sheet.cell(row=row, column=col, value=value); cell.font = self.styles["default_font"]
        for key in style_keys:
            style_obj = self.styles.get(key)
            if isinstance(style_obj, Font): cell.font = style_obj
            elif isinstance(style_obj, Alignment): cell.alignment = style_obj
            elif isinstance(style_obj, PatternFill): cell.fill = style_obj
            elif isinstance(style_obj, Border): cell.border = style_obj
            elif "currency" in key: cell.number_format = style_obj
    def _apply_border_to_range(self, sheet, cell_range):
        rows = sheet[cell_range]; border = self.styles["thin_border"]
        for row in rows:
            for cell in row: cell.border = border
    def draw_page_header(self, sheet, title, start_row):
        r = start_row; sheet.merge_cells(f'A{r}:H{r}'); self._set_cell(sheet, r, 1, title, ["header_main", "align_center"]); r += 1
        info = self.report_data["header"]
        sheet.merge_cells(f'A{r}:D{r}'); self._set_cell(sheet, r, 1, f"EMPRESA: {info.get('empresa', '')}", ["header_info"]); sheet.merge_cells(f'E{r}:H{r}'); self._set_cell(sheet, r, 5, f"CNPJ: {info.get('cnpj', '')}", ["header_info"]); r += 1
        sheet.merge_cells(f'A{r}:D{r}'); self._set_cell(sheet, r, 1, f"INSC.EST.: {info.get('insc_est', '')}", ["header_info"]); sheet.merge_cells(f'E{r}:H{r}'); self._set_cell(sheet, r, 5, f"MÊS OU PERÍODO/ANO: {info.get('periodo', '')}", ["header_info"]); return r + 2
    def draw_movimento_table(self, sheet, title, data, start_row):
        widths = {'A': 8, 'B': 5, 'C': 18, 'D': 18, 'E': 18, 'F': 22, 'G': 18}; [sheet.column_dimensions[col].__setattr__('width', width) for col, width in widths.items()]
        r = start_row; sheet.merge_cells(f'A{r}:G{r}'); self._set_cell(sheet, r, 1, title, ["header_table", "align_center"])
        r += 1; sheet.merge_cells(f'A{r}:B{r}'); self._set_cell(sheet, r, 1, "CODIFICAÇÃO", ["header_table", "align_center"]); sheet.merge_cells(f'C{r}:G{r}'); self._set_cell(sheet, r, 3, "I P I - VALORES FISCAIS", ["header_table", "align_center"])
        r += 1; op_t = f"OPERAÇÃO COM {'CRÉDITO' if title == 'ENTRADAS' else 'DÉBITO'} DO IMPOSTO"; op_sem_t = f"OPERAÇÃO SEM {'CRÉDITO' if title == 'ENTRADAS' else 'DÉBITO'} DO IMPOSTO"
        sheet.merge_cells(f'C{r}:C{r+1}'); self._set_cell(sheet, r, 3, "VALORES\nCONTÁBEIS", ["header_table", "align_center"]); sheet.merge_cells(f'D{r}:E{r}'); self._set_cell(sheet, r, 4, op_t, ["header_table", "align_center"]); sheet.merge_cells(f'F{r}:G{r}'); self._set_cell(sheet, r, 6, op_sem_t, ["header_table", "align_center"])
        self._set_cell(sheet, r, 1, "CONTÁBIL", ["header_table", "align_center"]); self._set_cell(sheet, r, 2, "FISCAL", ["header_table", "align_center"]); r += 1
        self._set_cell(sheet, r, 4, "BASE DE CÁLCULO", ["header_table", "align_center"]); self._set_cell(sheet, r, 5, "IMPOSTO", ["header_table", "align_center"]); self._set_cell(sheet, r, 6, "ISENTAS OU NÃO TRIBUTADAS", ["header_table", "align_center"]); self._set_cell(sheet, r, 7, "OUTRAS", ["header_table", "align_center"]); r += 1
        subtotals = {f'cat{i}': {k: 0.0 for k in ["VCON", "BASE", "VIMP", "VISE", "VOUT"]} for i in range(1, 4)}
        for row_data in data:
            self._set_cell(sheet, r, 1, ""); self._set_cell(sheet, r, 2, row_data.get('CFOP', ''), ["align_center"])
            for i, k in enumerate(["VCON", "BASE", "VIMP", "VISE", "VOUT"]): self._set_cell(sheet, r, 3 + i, float(row_data.get(k, 0) or 0), ["align_right", "currency"])
            r += 1; cfop_str = str(row_data.get('CFOP', '0')); cat = ''
            if cfop_str.startswith(('1', '5')): cat = 'cat1'
            elif cfop_str.startswith(('2', '6')): cat = 'cat2'
            elif cfop_str.startswith(('3', '7')): cat = 'cat3'
            if cat:
                for k in ["VCON", "BASE", "VIMP", "VISE", "VOUT"]: subtotals[cat][k] += float(row_data.get(k, 0) or 0)
        self._apply_border_to_range(sheet, f'A{start_row}:G{r-1}'); r += 1
        sheet.merge_cells(f'A{r}:B{r}'); self._set_cell(sheet, r, 1, f"SUBTOTAIS {title}", ["header_info"]); r += 1
        subtotal_labels = {'ENTRADAS': ["1-000 DO ESTADO", "2-000 DE OUTROS ESTADOS", "3-000 DO EXTERIOR"], 'SAIDAS': ["5-000 PARA O ESTADO", "6-000 PARA OUTROS ESTADOS", "7-000 PARA O EXTERIOR"]}
        for i, cat in enumerate(['cat1', 'cat2', 'cat3']):
            sheet.merge_cells(f'A{r}:B{r}'); self._set_cell(sheet, r, 1, subtotal_labels[title][i])
            for j, k in enumerate(["VCON", "BASE", "VIMP", "VISE", "VOUT"]): self._set_cell(sheet, r, 3 + j, subtotals[cat][k], ["align_right", "currency"])
            r += 1
        totals = {k: sum(subtotals[cat][k] for cat in subtotals) for k in subtotals['cat1']}
        sheet.merge_cells(f'A{r}:B{r}'); self._set_cell(sheet, r, 1, "TOTAIS", ["header_table", "align_center", "gray_fill"])
        for j, k in enumerate(["VCON", "BASE", "VIMP", "VISE", "VOUT"]): self._set_cell(sheet, r, 3 + j, totals[k], ["header_table", "align_right", "currency", "gray_fill"])
        self._apply_border_to_range(sheet, f'A{r-3}:G{r}'); return r + 2
    def draw_resumo_table(self, sheet, start_row):
        resumo = self.report_data["resumo"]
        widths = {'A': 5, 'B': 75, 'C': 20, 'D': 20}; [sheet.column_dimensions[col].__setattr__('width', width) for col, width in widths.items()]
        r = start_row; sheet.merge_cells(f'A{r}:D{r}'); self._set_cell(sheet, r, 1, "R E S U M O   D A   A P U R A Ç Ã O   D O   I M P O S T O", ["header_main", "align_center"]); r += 2
        def draw_section(title, rows_data, v_text):
            nonlocal r; start_r = r
            sheet.merge_cells(f'C{r}:D{r}'); self._set_cell(sheet, r, 3, "V A L O R E S", ["header_table", "align_center"]); self._set_cell(sheet, r, 2, title, ["header_table", "align_center"]); r += 1
            self._set_cell(sheet, r, 2, ""); self._set_cell(sheet, r, 3, "COLUNA AUXILIAR", ["header_table", "align_center"]); self._set_cell(sheet, r, 4, "SOMAS", ["header_table", "align_center"]); r += 1
            for desc, aux, somas in rows_data:
                self._set_cell(sheet, r, 2, desc, ["align_left"]); self._set_cell(sheet, r, 3, aux if aux not in [None, ""] else None, ["align_right", "currency"]); self._set_cell(sheet, r, 4, somas if somas not in [None, ""] else None, ["align_right", "currency"]); r += 1
            sheet.merge_cells(f'A{start_r}:A{r-1}'); self._set_cell(sheet, start_r, 1, v_text, ["header_table", "vertical_text"]); self._apply_border_to_range(sheet, f'A{start_r}:D{r-1}')
        debito_rows = [("001 - POR SAÍDAS/PRESTAÇÕES COM DÉBITO DO IMPOSTO", "", resumo.get("debito_saidas")), ("002 - OUTROS DÉBITOS(DISCRIMINAR ABAIXO)", "", resumo.get("debito_outros")), ("003 - ESTORNO DE CRÉDITOS(DISCRIMINAR ABAIXO)", "", resumo.get("debito_estornos")), ("004 - SUBTOTAL", "", resumo.get("debito_subtotal"))]; draw_section("D É B I T O   D O   I M P O S T O", debito_rows, "DÉBITO")
        credito_rows = [("005 - POR ENTRADAS/AQUISIÇÕES COM CRÉDITO DO IMPOSTO", "", resumo.get("credito_entradas")), ("006 - OUTROS CRÉDITOS(DISCRIMINAR ABAIXO)", "", resumo.get("credito_outros"))]
        for detalhe in resumo.get("credito_outros_detalhes", []): credito_rows.append((f"      ({detalhe['desc']})", detalhe['valor'], ""));
        credito_rows.extend([("007 - ESTORNO DE DÉBITOS(DISCRIMINAR ABAIXO)", "", resumo.get("credito_estornos")), ("008 - SUBTOTAL", "", resumo.get("credito_subtotal")), ("009 - SALDO CREDOR DO PERÍODO ANTERIOR", "", resumo.get("saldo_anterior")), ("010 - TOTAL", "", resumo.get("credito_total"))]); draw_section("C R É D I T O   D O   I M P O S T O", credito_rows, "CRÉDITO")
        apuracao_rows = [("011 - SALDO DEVEDOR(DÉBITO - CRÉDITO)", "", resumo.get("saldo_devedor")), ("012 - DEDUÇÕES(DISCRIMINAR ABAIXO)", "", resumo.get("deducoes")), ("013 - IMPOSTO A RECOLHER", "", resumo.get("imposto_recolher")), ("014 - SALDO CREDOR(CRÉDITO - DÉBITO) A TRANSPORTAR PARA O PERÍODO SEGUINTE", "", resumo.get("saldo_credor_final"))]; draw_section("A P U R A Ç Ã O   D O   S A L D O", apuracao_rows, "SALDO")
        return r
    def generate(self):
        sheet = self.workbook.create_sheet("Apuração IPI"); next_row = 1
        next_row = self.draw_page_header(sheet, "REGISTRO DE APURAÇÃO DO IPI", next_row)
        if self.report_data.get("entradas"): next_row = self.draw_movimento_table(sheet, "ENTRADAS", self.report_data["entradas"], next_row)
        if self.report_data.get("saidas"): next_row = self.draw_movimento_table(sheet, "SAIDAS", self.report_data["saidas"], next_row)
        next_row += 2
        if self.report_data.get("resumo"): self.draw_resumo_table(sheet, next_row)
    def save(self):
        self.workbook.save(self.filename)


# --- NOVA FUNÇÃO DE GERAÇÃO DE RELATÓRIO ---
def gerarLivroDeIpi(codi_emp, data_inicio, data_fim, gerar_pdf=False, gerar_xlsx=True):
    """
    Gera o relatório de Apuração de IPI em formato PDF e/ou XLSX.

    Esta função orquestra todo o processo: busca os dados no banco,
    processa-os e, em seguida, chama as classes apropriadas para
    gerar os arquivos de relatório solicitados.

    Args:
        codi_emp (int): O código da empresa para a qual o relatório será gerado.
        data_inicio (str): A data inicial do período no formato 'YYYY-MM-DD'.
        data_fim (str): A data final do período no formato 'YYYY-MM-DD'.
        gerar_pdf (bool, optional): Se True, gera o relatório em PDF. Defaults to False.
        gerar_xlsx (bool, optional): Se True, gera o relatório em XLSX. Defaults to True.

    Returns:
        list: Uma lista contendo os nomes dos arquivos que foram gerados com sucesso.
              Retorna uma lista vazia se nenhum arquivo for gerado ou se ocorrer um erro.
    """
    if not (gerar_pdf or gerar_xlsx):
        print("Nenhum formato de saída selecionado (PDF ou XLSX). Abortando.")
        return []

    print(f"Iniciando geração do relatório para a empresa {codi_emp} no período de {data_inicio} a {data_fim}.")
    
    query_map = fetch_data_from_db(CONN_STR, SQL_TEMPLATES, codi_emp, data_inicio, data_fim)
    if not any(query_map.values()):
        print("Nenhum dado foi retornado do banco. Abortando a geração do relatório.")
        return []

    print("\nExtraindo e estruturando os dados para o relatório...")
    report_data = extract_report_data(query_map, data_inicio, data_fim)

    try:
        data_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
        periodo_str = data_obj.strftime('%Y-%m')
    except ValueError:
        print(f"Formato de data inválido: {data_inicio}. Usando 'periodo_desconhecido'.")
        periodo_str = 'periodo_desconhecido'

    base_filename = f"Livro_IPI{codi_emp}"
    generated_files = []

    if gerar_pdf:
        pdf_filename = f"{base_filename}.pdf"
        print(f"\n--- Gerando PDF: {pdf_filename} ---")
        try:
            pdf = PDF(report_data, orientation="L", unit="mm", format="A4")
            pdf.add_page()
            pdf.draw_page_header("REGISTRO DE APURAÇÃO DO IPI")
            if report_data.get("entradas"):
                pdf.draw_movimento_table("ENTRADAS", report_data["entradas"])
            if report_data.get("saidas"):
                pdf.draw_movimento_table("SAIDAS", report_data["saidas"])
            
            # Adiciona nova página para o resumo para evitar quebras estranhas
            pdf.add_page()
            pdf.draw_page_header("REGISTRO DE APURAÇÃO DO IPI")
            if report_data.get("resumo"):
                pdf.draw_resumo_table()
            
            pdf.output(pdf_filename)
            print(f"PDF salvo com sucesso em '{pdf_filename}'")
            generated_files.append(pdf_filename)
        except Exception as e:
            print(f"ERRO AO GERAR/SALVAR PDF: {e}")
    
    if gerar_xlsx:
        xlsx_filename = f"{base_filename}.xlsx"
        print(f"\n--- Gerando XLSX: {xlsx_filename} ---")
        try:
            excel = ExcelReport(report_data, xlsx_filename)
            excel.generate()
            excel.save()
            print(f"XLSX salvo com sucesso em '{xlsx_filename}'")
            generated_files.append(xlsx_filename)
        except Exception as e:
            print(f"ERRO AO GERAR/SALVAR XLSX: {e}")

    return generated_files


