import pyodbc
import pandas as pd
import json
from datetime import datetime, date

from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Para geração do XLSX - instalar com: pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Definir pagesize customizado
CUSTOM_PAGESIZE = (11*inch, 20*inch)

# ------------------------------------------------------------
# ESTILOS E CONFIGURAÇÕES GLOBAIS
# ------------------------------------------------------------
styles = getSampleStyleSheet()
header_style = ParagraphStyle(
    'header',
    parent=styles['Normal'],
    fontName='Helvetica-Bold',
    fontSize=6,
    alignment=1,
    leading=7
)

# String de conexão com o banco de dados
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

# ------------------------------------------------------------
# CONSULTAS SQL
# ------------------------------------------------------------
SQL_LIVRO_ANTERIOR = """
SELECT LIVRO.NUME_LIV, TDDATA_ANTERIOR.DATA_LIVRO
FROM BETHADBA.EFLIVROS AS LIVRO,
LATERAL ( SELECT MAX ( LIVRO_AUX.DATA_LIV ) AS DATA_LIVRO
          FROM BETHADBA.EFLIVROS AS LIVRO_AUX
          WHERE LIVRO_AUX.CODI_EMP = LIVRO.CODI_EMP
            AND LIVRO_AUX.CODI_LIV = LIVRO.CODI_LIV
            AND LIVRO_AUX.DATA_LIV < ? ) AS TDDATA_ANTERIOR
WHERE LIVRO.CODI_EMP = ?
  AND LIVRO.CODI_LIV = 5
  AND LIVRO.DATA_LIV = TDDATA_ANTERIOR.DATA_LIVRO
"""

SQL_DOCUMENTOS_FISCAIS = """
SELECT S.DSER_SER AS DATA, S.NUME_SER AS NUMERO, S.CODI_SER AS CODIGO, S.CODI_ESP AS CODESP, E.NOME_ESP AS ESPECIE, TRIM(S.SERI_SER) AS SERIE, TRIM(RIGHT(TRIM(CAST(S.SUB_SERIE_SER AS CHAR(10))), 3)) AS SUB_SERIE, DAY(S.DSER_SER) AS DIA, S.VCON_SER AS VCON, S.ATEX_SER AS NUMEROF, TDPOSSUI_OBS.EXISTE AS POSSUI_OBS, 'V' AS TIPO, 0 AS ALIQUOTA, 0 AS ALIQTAB, 0 AS ALIQRET, 0 AS I_BILHETE FROM BETHADBA.EFSERVICOS AS S INNER JOIN BETHADBA.EFESPECIES AS E ON E.CODI_ESP = S.CODI_ESP, LATERAL(SELECT COALESCE(MAX('S'), 'N') AS EXISTE FROM BETHADBA.EFSERVICOS_OBSERVACAO_FISCO AS OBSERVACAO_FISCO WHERE OBSERVACAO_FISCO.CODI_EMP = S.CODI_EMP AND OBSERVACAO_FISCO.CODI_SER = S.CODI_SER) AS TDPOSSUI_OBS WHERE S.DSER_SER >= ? AND S.DSER_SER <= ? AND S.CODI_EMP = ? AND ('N' = 'S' AND (EXISTS(SELECT 1 FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS P WHERE P.CODI_EMP = S.CODI_EMP AND P.CODI_ACU = S.CODI_ACU AND P.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(S.CODI_EMP, S.CODI_ACU, S.DSER_SER) AND P.CODI_IMP = 44 AND -1 IN (- 1, 0) AND NOT EXISTS(SELECT 1 FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS P2 WHERE P2.CODI_EMP = P.CODI_EMP AND P2.CODI_ACU = P.CODI_ACU AND P2.VIGENCIA_ACU = P.VIGENCIA_ACU AND P2.CODI_IMP = 3) AND ((P.SIMPLESN_ANEXO_IAC = 5 AND P.SIMPLESN_TABELA_IAC = 4) OR (P.SIMPLESN_ANEXO_IAC = 3 AND P.SIMPLESN_SECAO_IAC = 6 AND P.SIMPLESN_TABELA_IAC = 1))) OR EXISTS(SELECT 1 FROM BETHADBA.EFIMPSER AS I, BETHADBA.EFACUMULADOR_VIGENCIA AS A, BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS P WHERE I.CODI_SER = S.CODI_SER AND I.CODI_EMP = S.CODI_EMP AND A.CODI_EMP = I.CODI_EMP AND A.CODI_ACU = S.CODI_ACU AND A.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(S.CODI_EMP, S.CODI_ACU, S.DSER_SER) AND P.VIGENCIA_ACU = A.VIGENCIA_ACU AND P.CODI_EMP = A.CODI_EMP AND P.CODI_ACU = A.CODI_ACU AND (I.CODI_IMP = 3 OR (I.CODI_IMP = 18 AND 0 = 1)) AND (-1 = - 1 OR I.ALIQ_ISE = -1) AND ((P.SIMPLESN_ANEXO_IAC = 5 AND P.SIMPLESN_TABELA_IAC = 4) OR (P.SIMPLESN_ANEXO_IAC = 3 AND P.SIMPLESN_SECAO_IAC = 6 AND P.SIMPLESN_TABELA_IAC = 1)))) OR ('N' = 'N' AND EXISTS(SELECT 1 FROM BETHADBA.EFIMPSER AS I WHERE I.CODI_SER = S.CODI_SER AND I.CODI_EMP = S.CODI_EMP AND (I.CODI_IMP = 3 OR (I.CODI_IMP = 18 AND 0 = 1)) AND (-1 = - 1 OR I.ALIQ_ISE = -1)))) ORDER BY 1, 2, 3
"""

SQL_TEMPLATE_OBSERVACOES = """
SELECT OBSERVACAO_FISCO.I_SEQUENCIAL_FISCO AS I_SEQUENCIAL_OBS, COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') AS OBSERVACAO_FISCO FROM BETHADBA.EFSERVICOS_OBSERVACAO_FISCO AS OBSERVACAO_FISCO WHERE OBSERVACAO_FISCO.CODI_EMP = ? AND OBSERVACAO_FISCO.CODI_SER = ? AND ? = 'V' AND COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') <> '' UNION ALL SELECT OBSERVACAO_FISCO.I_SEQUENCIAL_FISCO AS I_SEQUENCIAL_OBS, COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') AS OBSERVACAO_FISCO FROM BETHADBA.EFSAIDAS_OBSERVACAO_FISCO AS OBSERVACAO_FISCO WHERE OBSERVACAO_FISCO.CODI_EMP = ? AND OBSERVACAO_FISCO.CODI_SAI = ? AND ? = 'S' AND COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') <> '' UNION ALL SELECT OBSERVACAO_FISCO.I_SEQUENCIAL_FISCO AS I_SEQUENCIAL_OBS, COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') AS OBSERVACAO_FISCO FROM BETHADBA.EFECF_REDUCAO_Z_OBSERVACAO_FISCO AS OBSERVACAO_FISCO WHERE OBSERVACAO_FISCO.CODI_EMP = ? AND OBSERVACAO_FISCO.I_REDUCAO = ? AND ? = 'Z' AND COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') <> '' UNION ALL SELECT OBSERVACAO_FISCO.I_SEQUENCIAL_FISCO AS I_SEQUENCIAL_OBS, COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') AS OBSERVACAO_FISCO FROM BETHADBA.EFRESUMO_MOVIMENTO_DIARIO_BILHETE_OBSERVACAO_FISCO AS OBSERVACAO_FISCO WHERE OBSERVACAO_FISCO.CODI_EMP = ? AND OBSERVACAO_FISCO.I_RESUMO = ? AND OBSERVACAO_FISCO.I_BILHETE = 0 AND ? = 'RM' AND COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') <> '' UNION ALL SELECT OBSERVACAO_FISCO.I_SEQUENCIAL_FISCO AS I_SEQUENCIAL_OBS, COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') AS OBSERVACAO_FISCO FROM BETHADBA.EFBILHETE_PASSAGEM_OBSERVACAO_FISCO AS OBSERVACAO_FISCO WHERE OBSERVACAO_FISCO.CODI_EMP = ? AND OBSERVACAO_FISCO.I_BILHETE = ? AND ? = 'BP' AND COALESCE(OBSERVACAO_FISCO.DESCRICAO_FISCO, '') <> '' ORDER BY 1
"""

SQL_TEMPLATE_IMPOSTOS = """
SELECT I.CODI_EMP AS EMPRESA, I.CODI_SER AS NOTA, I.CODI_IMP AS IMPOSTO, I.SEQU_ISE AS SEQUENCIAL, I.BCAL_ISE AS BASE, I.ALIQ_ISE AS ALIQUOTA, I.VLOR_ISE AS VALOR, I.VISE_ISE AS ISENTAS, I.VOUT_ISE AS OUTRAS, TDAUX.MOSTRA AS MOSTRA, 0 AS ALIQTAB, 0 AS ALIQRET, 0 AS REDTAB FROM BETHADBA.EFIMPSER AS I, LATERAL(SELECT (CASE WHEN I.CODI_IMP = 18 THEN (I.BCAL_ISE + I.VISE_ISE + I.VOUT_ISE) ELSE 1 END) AS MOSTRA FROM DSDBA.DUMMY) AS TDAUX WHERE I.CODI_EMP = ? AND I.CODI_SER = ? AND 'N' = 'N' AND TDAUX.MOSTRA > 0 AND ? = 'V' AND ('N' = 'S' AND EXISTS(SELECT 1 FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS P INNER JOIN BETHADBA.EFSERVICOS AS S ON P.CODI_EMP = S.CODI_EMP AND P.CODI_ACU = S.CODI_ACU WHERE P.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(S.CODI_EMP, S.CODI_ACU, S.DSER_SER) AND I.CODI_EMP = S.CODI_EMP AND I.CODI_SER = S.CODI_SER AND P.CODI_IMP = 44 AND ((P.SIMPLESN_ANEXO_IAC = 5 AND P.SIMPLESN_TABELA_IAC = 4) OR (P.SIMPLESN_ANEXO_IAC = 3 AND P.SIMPLESN_SECAO_IAC = 6 AND P.SIMPLESN_TABELA_IAC = 1))) AND ((I.CODI_IMP = 3 OR (I.CODI_IMP = 18 AND 0 = 1)) AND (-1 = - 1 OR I.ALIQ_ISE = -1)) OR ('N' = 'N' AND (I.CODI_IMP = 3 OR (I.CODI_IMP = 18 AND 0 = 1)) AND (-1 = - 1 OR I.ALIQ_ISE = -1))) ORDER BY 1, 2, 3, 4
"""

SQL_TOTAIS_GERAIS = """
SELECT P.CODI_IMP AS IMPOSTO, SUM(COALESCE(P.BCAL_ISE, 0)) AS BASE, SUM(COALESCE(P.VLOR_ISE, 0)) AS VALOR, SUM(COALESCE(P.VISE_ISE, 0)) AS ISENTAS, SUM(COALESCE(P.VOUT_ISE, 0)) AS OUTRAS FROM BETHADBA.EFSERVICOS AS S INNER JOIN BETHADBA.EFIMPSER AS P ON P.CODI_EMP = S.CODI_EMP AND P.CODI_SER = S.CODI_SER WHERE S.CODI_EMP = ? AND S.DSER_SER >= ? AND S.DSER_SER <= ? AND ('N' = 'N' AND (P.CODI_IMP = 3 OR (P.CODI_IMP = 18 AND 0 = 1))) GROUP BY P.CODI_IMP
"""

SQL_DADOS_EMPRESA_BASICOS = """
SELECT NOME_EMP, CGCE_EMP, IEST_EMP, IMUN_EMP
FROM BETHADBA.GEEMPRE
WHERE CODI_EMP = ?
"""

# ------------------------------------------------------------
# FUNÇÕES AUXILIARES
# ------------------------------------------------------------
def formatar_cnpj(cnpj):
    if not cnpj:
        return ""
    cnpj_numeros = ''.join(filter(str.isdigit, str(cnpj)))
    if len(cnpj_numeros) == 14:
        return f"{cnpj_numeros[:2]}.{cnpj_numeros[2:5]}.{cnpj_numeros[5:8]}/{cnpj_numeros[8:12]}-{cnpj_numeros[12:14]}"
    return str(cnpj)

def formatar_data_brasileira(data_str):
    if not data_str:
        return ""
    try:
        if '/' in data_str:
            return data_str
        if '-' in data_str and len(data_str) == 10:
            ano, mes, dia = data_str.split('-')
            return f"{dia}/{mes}/{ano}"
    except:
        pass
    return data_str

def executar_consulta_json(conn, sql, params=()):
    df = pd.read_sql(sql, conn, params=params)
    def conv(o):
        if isinstance(o, (datetime, date)):
            return o.isoformat()
        return str(o)
    return json.loads(df.to_json(orient="records", date_format="iso"))

# ------------------------------------------------------------
# FUNÇÕES DE GERAÇÃO DE ARQUIVOS
# ------------------------------------------------------------
def gerar_pdf(parametros, livro_ant, documentos, totais, dados_empresa):
    doc = SimpleDocTemplate(
        "LivroISS.pdf",
        pagesize=CUSTOM_PAGESIZE,
        leftMargin=15,
        rightMargin=15,
        topMargin=15,
        bottomMargin=15
    )
    normal = styles["Normal"]
    normal.fontSize = 7
    normal.leading  = 9
    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading1"],
        alignment=1,
        fontSize=11
    )
    
    story = []
    
    story.append(Paragraph("REGISTRO DE NOTAS FISCAIS E SERVIÇOS PRESTADOS", title_style))
    story.append(Spacer(1, 6))
    
    nome_empresa = dados_empresa.get('nome', 'N/A')
    insc_municipal = dados_empresa.get('insc_municipal', 'N/A')
    insc_estadual = dados_empresa.get('insc_estadual', 'N/A')
    cnpj = dados_empresa.get('cnpj', 'N/A')
    
    story.append(Paragraph(f"EMPRESA: {nome_empresa}", normal))
    
    linha_inscricoes = f"INSC.MUN.: {insc_municipal}            INSC.EST.: {insc_estadual}"
    story.append(Paragraph(linha_inscricoes, normal))
    
    data_inicial_br = formatar_data_brasileira(parametros['data_inicial'])
    data_final_br = formatar_data_brasileira(parametros['data_final'])
    linha_cnpj_periodo = f"CNPJ: {cnpj}     MÊS OU PERÍODO/ANO: {data_inicial_br} a {data_final_br}"
    story.append(Paragraph(linha_cnpj_periodo, normal))
    
    story.append(Spacer(1, 6))

    # CORREÇÃO: trocado '\' por '\n' para corrigir SyntaxWarning e melhorar layout
    header_data = [
        ['DOCUMENTOS FISCAIS', '', '', '', 'VALOR\nCONTÁBIL', 'CLASSIFICAÇÃO\nCONTÁBIL', 'VALORES FISCAIS', '', '', '', '', '', 'OBSERVAÇÕES'],
        ['ESPÉCIE', 'SÉRIE', 'NÚMERO/ATÉ', 'DIA', '', '', 'ISS', 'OPERAÇÕES COM DÉBITO DO IMPOSTO', '', '', 'OPERAÇÕES SEM DÉBITO DO IMPOSTO', '', ''],
        ['', '', '', '', '', '', '', 'BASE DE\nCÁLCULO', 'ALÍQUOTA', 'IMPOSTO\nDEBITADO', 'ISENTAS', 'OUTRAS', '']
    ]

    data = []
    data.extend(header_data)
    
    for doc_item in documentos:
        especie_curta = doc_item.get("ESPECIE", "")[:11]
        imp_list = doc_item.get("impostos_detalhados", [])
        if imp_list:
            imp0    = imp_list[0]
            base    = imp0.get("BASE", 0)
            aliq    = imp0.get("ALIQUOTA", 0)
            imp_deb = imp0.get("VALOR", 0)
            isentas = imp0.get("ISENTAS", 0)
            outras  = imp0.get("OUTRAS", 0)
        else:
            base = aliq = imp_deb = isentas = 0
            outras = doc_item.get("VCON", 0)

        obs_text = ""
        if doc_item.get("observacoes_detalhadas"):
            obs_list = doc_item.get("observacoes_detalhadas", [])
            if obs_list:
                obs_text = obs_list[0].get("OBSERVACAO_FISCO","")
        
        obs_paragraph = Paragraph(obs_text, ParagraphStyle(
            'obs_style', parent=styles['Normal'], fontSize=6, leading=8, alignment=0, wordWrap='LTR'
        )) if obs_text else ""

        valor_cont = doc_item.get("VCON", 0)
        
        numero = doc_item.get("NUMERO", "")
        if isinstance(numero, (int, float)):
            numero = str(int(numero))

        row = [
            especie_curta, doc_item.get("SERIE",""), numero, doc_item.get("DIA",""),
            f"{float(valor_cont):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            "", "ISS",
            f"{float(base):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"{float(aliq):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"{float(imp_deb):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"{float(isentas):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"{float(outras):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            obs_paragraph
        ]
        data.append(row)

    col_widths = [0.6*inch, 0.5*inch, 0.7*inch, 0.4*inch, 0.6*inch, 0.9*inch, 
                  0.4*inch, 0.7*inch, 0.6*inch, 0.7*inch, 0.9*inch, 0.9*inch, 0.7*inch]
    
    tbl = Table(data, colWidths=col_widths, repeatRows=3)
    
    ts = TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black), ('ALIGN', (0,0), (-1,2), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (12,3), (12,-1), 'LEFT'),
        ('VALIGN', (12,3), (12,-1), 'TOP'), ('FONTNAME', (0,0), (-1,2), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 6), ('ROWBACKGROUNDS', (0,0), (-1,-1), [None]),
        ('SPAN', (0, 0), (3, 0)), ('SPAN', (4, 0), (4, 2)), ('SPAN', (5, 0), (5, 2)),
        ('SPAN', (6, 0), (11, 0)), ('SPAN', (12, 0), (12, 2)), ('SPAN', (6, 1), (6, 2)),
        ('SPAN', (7, 1), (9, 1)), ('SPAN', (10, 1), (11, 1)), ('SPAN', (0, 1), (0, 2)),
        ('SPAN', (1, 1), (1, 2)), ('SPAN', (2, 1), (2, 2)), ('SPAN', (3, 1), (3, 2)),
    ])
    
    tbl.setStyle(ts)
    story.append(tbl)
    story.append(Spacer(1, 6))

    if totais:
        t = totais[0]
        total_valor_contabil = sum(float(doc.get('VCON', 0)) for doc in documentos)
        valor_contabil_formatado = f"{total_valor_contabil:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        base_total = f"{t.get('BASE',0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        iss_total = f"{t.get('VALOR',0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        isentas_total = f"{t.get('ISENTAS',0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        outras_total = f"{t.get('OUTRAS',0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        totals_data = [
            ["TOTAL GERAL", "", "", "", valor_contabil_formatado, "", "", base_total, "", iss_total, isentas_total, outras_total, ""]
        ]
        
        totals_table = Table(totals_data, colWidths=col_widths)
        totals_style = TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black), ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 5.0),
        ])
        totals_table.setStyle(totals_style)
        story.append(totals_table)
    
    doc.build(story)
    print("PDF gerado: livro_fiscal_ISS.pdf")


def gerar_xlsx_simples(parametros, livro_ant, documentos, totais, dados_empresa):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Livro Fiscal ISS"
        
        # Estilos básicos
        font_title = Font(name='Arial', size=11, bold=True)
        font_header = Font(name='Arial', size=8, bold=True)
        font_data = Font(name='Arial', size=7)
        font_totals = Font(name='Arial', size=7, bold=True)
        
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # CABEÇALHO TEXTUAL
        ws.merge_cells('A1:M1')
        ws['A1'] = "REGISTRO DE NOTAS FISCAIS E SERVIÇOS PRESTADOS"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_center
        row += 2
        
        # Dados da empresa
        nome_empresa = str(dados_empresa.get('nome', 'N/A'))
        insc_municipal = str(dados_empresa.get('insc_municipal', 'N/A'))
        insc_estadual = str(dados_empresa.get('insc_estadual', 'N/A'))
        cnpj = dados_empresa.get('cnpj', 'N/A')
        
        ws[f'A{row}'] = f"EMPRESA: {nome_empresa}"; row += 1
        ws[f'A{row}'] = f"INSC.MUN.: {insc_municipal}            INSC.EST.: {insc_estadual}"; row += 1
        data_inicial_br = formatar_data_brasileira(parametros['data_inicial'])
        data_final_br = formatar_data_brasileira(parametros['data_final'])
        ws[f'A{row}'] = f"CNPJ: {cnpj}     MÊS OU PERÍODO/ANO: {data_inicial_br} a {data_final_br}"; row += 2
        
        header_start_row = row
        
        # LINHA 0
        ws.cell(row=row, column=1, value='DOCUMENTOS FISCAIS'); ws.cell(row=row, column=5, value='VALOR\nCONTÁBIL'); ws.cell(row=row, column=6, value='CLASSIFICAÇÃO\nCONTÁBIL'); ws.cell(row=row, column=7, value='VALORES FISCAIS'); ws.cell(row=row, column=13, value='OBSERVAÇÕES'); row += 1
        # LINHA 1
        ws.cell(row=row, column=1, value='ESPÉCIE'); ws.cell(row=row, column=2, value='SÉRIE'); ws.cell(row=row, column=3, value='NÚMERO/ATÉ'); ws.cell(row=row, column=4, value='DIA'); ws.cell(row=row, column=7, value='ISS'); ws.cell(row=row, column=8, value='OPERAÇÕES COM DÉBITO DO IMPOSTO'); ws.cell(row=row, column=11, value='OPERAÇÕES SEM DÉBITO DO IMPOSTO'); row += 1
        # LINHA 2
        ws.cell(row=row, column=8, value='BASE DE\nCÁLCULO'); ws.cell(row=row, column=9, value='ALÍQUOTA'); ws.cell(row=row, column=10, value='IMPOSTO\nDEBITADO'); ws.cell(row=row, column=11, value='ISENTAS'); ws.cell(row=row, column=12, value='OUTRAS'); row += 1
        
        header_end_row = row - 1

        for r in range(header_start_row, header_end_row + 1):
            for c in range(1, 14):
                cell = ws.cell(row=r, column=c)
                cell.font = font_header
                cell.alignment = align_center
                cell.border = border_thin

        # Mesclagens
        ws.merge_cells(start_row=header_start_row, start_column=1, end_row=header_start_row, end_column=4)      # DOCUMENTOS FISCAIS
        ws.merge_cells(start_row=header_start_row, start_column=5, end_row=header_end_row, end_column=5)        # VALOR CONTÁBIL
        ws.merge_cells(start_row=header_start_row, start_column=6, end_row=header_end_row, end_column=6)        # CLASSIFICAÇÃO CONTÁBIL
        ws.merge_cells(start_row=header_start_row, start_column=7, end_row=header_start_row, end_column=12)     # VALORES FISCAIS
        ws.merge_cells(start_row=header_start_row, start_column=13, end_row=header_end_row, end_column=13)      # OBSERVAÇÕES
        ws.merge_cells(start_row=header_start_row + 1, start_column=1, end_row=header_end_row, end_column=1)    # ESPÉCIE
        ws.merge_cells(start_row=header_start_row + 1, start_column=2, end_row=header_end_row, end_column=2)    # SÉRIE
        ws.merge_cells(start_row=header_start_row + 1, start_column=3, end_row=header_end_row, end_column=3)    # NÚMERO/ATÉ
        ws.merge_cells(start_row=header_start_row + 1, start_column=4, end_row=header_end_row, end_column=4)    # DIA
        ws.merge_cells(start_row=header_start_row + 1, start_column=7, end_row=header_end_row, end_column=7)    # ISS
        ws.merge_cells(start_row=header_start_row + 1, start_column=8, end_row=header_start_row + 1, end_column=10) # OP COM DÉBITO
        ws.merge_cells(start_row=header_start_row + 1, start_column=11, end_row=header_start_row + 1, end_column=12)# OP SEM DÉBITO
        
        # DADOS
        for doc_item in documentos:
            especie = str(doc_item.get("ESPECIE", ""))[:11]; serie = str(doc_item.get("SERIE", "")); numero = doc_item.get("NUMERO", "");
            if isinstance(numero, (int, float)) and numero: numero = int(numero)
            dia = doc_item.get("DIA", ""); valor_cont = float(doc_item.get("VCON", 0) or 0)
            imp_list = doc_item.get("impostos_detalhados", [])
            if imp_list: 
                imp0 = imp_list[0]; base = float(imp0.get("BASE", 0) or 0); aliq = float(imp0.get("ALIQUOTA", 0) or 0); imp_deb = float(imp0.get("VALOR", 0) or 0); isentas = float(imp0.get("ISENTAS", 0) or 0); outras = float(imp0.get("OUTRAS", 0) or 0)
            else: 
                base = aliq = imp_deb = isentas = 0.0; outras = valor_cont
            obs_text = ""
            obs_list = doc_item.get("observacoes_detalhadas", [])
            if obs_list: obs_text = str(obs_list[0].get("OBSERVACAO_FISCO", ""))

            row_data = [especie, serie, numero, dia, valor_cont, "", "ISS", base, aliq, imp_deb, isentas, outras, obs_text]
            ws.append(row_data)
            
            # Formatação da linha de dados
            for col_idx in range(1, 14):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = font_data
                cell.border = border_thin
                if col_idx in [5, 8, 9, 10, 11, 12]: # Colunas numéricas
                    cell.number_format = '#,##0.00'
                if col_idx == 13:
                    cell.alignment = align_left_wrap
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

        # TOTAIS
        if totais:
            t = totais[0]; total_valor_contabil = sum(float(doc.get('VCON', 0) or 0) for doc in documentos)
            base_total = float(t.get('BASE', 0) or 0); iss_total = float(t.get('VALOR', 0) or 0); isentas_total = float(t.get('ISENTAS', 0) or 0); outras_total = float(t.get('OUTRAS', 0) or 0)
            totals_data = ["TOTAL GERAL", "", "", "", total_valor_contabil, "", "", base_total, "", iss_total, isentas_total, outras_total, ""]
            
            ws.append(totals_data)
            for col_idx in range(1, 14):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = font_totals
                cell.alignment = align_center
                cell.border = border_thin
                if col_idx in [5, 8, 10, 11, 12]:
                     cell.number_format = '#,##0.00'

        # Ajustar larguras
        column_widths = [12, 10, 15, 8, 15, 20, 8, 15, 12, 15, 15, 15, 40]
        for i, width in enumerate(column_widths, 1): 
            ws.column_dimensions[get_column_letter(i)].width = width
        
        wb.save("LivroISS.xlsx")
        print("XLSX gerado: livro_fiscal_ISS.xlsx")
        
    except Exception as e:
        print(f"Erro geral ao gerar XLSX: {e}")
        import traceback
        traceback.print_exc()

# ------------------------------------------------------------
# FUNÇÃO PRINCIPAL REFATORADA
# ------------------------------------------------------------
# CORREÇÃO: Parâmetros renomeados de 'gerar_pdf' para 'exportar_pdf' e 'gerar_xlsx' para 'exportar_xlsx'
def gerar_livro_iss(codi_emp, data_inicio, data_fim, exportar_pdf=True, exportar_xlsx=False):
    """
    Gera o relatório do Livro Fiscal de ISS em formato PDF e/ou XLSX.
    
    Args:
        codi_emp (int): O código da empresa.
        data_inicio (str): A data de início no formato 'YYYY-MM-DD'.
        data_fim (str): A data de fim no formato 'YYYY-MM-DD'.
        exportar_pdf (bool): Se True, gera o arquivo PDF.
        exportar_xlsx (bool): Se True, gera o arquivo XLSX.

    Returns:
        list: Uma lista contendo os nomes dos arquivos gerados com sucesso.
              Retorna uma lista vazia em caso de falha.
    """
    if not (exportar_pdf or exportar_xlsx):
        print("Nenhum formato de saída foi selecionado (PDF ou XLSX).")
        return []

    # Lista para armazenar os nomes dos arquivos criados
    arquivos_gerados = []
    conn = None
    
    try:
        print("Conectando ao banco de dados...")
        conn = pyodbc.connect(CONN_STR)
        print("Conexão estabelecida.")

        print(f"Buscando dados para a empresa {codi_emp} no período de {data_inicio} a {data_fim}...")
        livro_ant = executar_consulta_json(conn, SQL_LIVRO_ANTERIOR, (data_fim, codi_emp))
        documentos = executar_consulta_json(conn, SQL_DOCUMENTOS_FISCAIS, (data_inicio, data_fim, codi_emp))
        
        for i, d in enumerate(documentos):
            cod = int(d['CODIGO'])
            tp = d['TIPO']
            d['observacoes_detalhadas'] = executar_consulta_json(conn, SQL_TEMPLATE_OBSERVACOES, (
                codi_emp, cod, tp, codi_emp, cod, tp, codi_emp, cod, tp,
                codi_emp, cod, tp, codi_emp, cod, tp
            ))
            d['impostos_detalhados'] = executar_consulta_json(conn, SQL_TEMPLATE_IMPOSTOS, (codi_emp, cod, tp))
        
        totais = executar_consulta_json(conn, SQL_TOTAIS_GERAIS, (codi_emp, data_inicio, data_fim))
        
        print(f"Buscando dados cadastrais da empresa {codi_emp}...")
        dados_empresa_result = executar_consulta_json(conn, SQL_DADOS_EMPRESA_BASICOS, (codi_emp,))
        
        if dados_empresa_result:
            empresa_data = dados_empresa_result[0]
            dados_empresa = {
                'nome': empresa_data.get('NOME_EMP'),
                'cnpj': formatar_cnpj(empresa_data.get('CGCE_EMP')),
                'insc_estadual': empresa_data.get('IEST_EMP'),
                'insc_municipal': empresa_data.get('IMUN_EMP'),
            }
        else:
            print(f"AVISO: Dados da empresa {codi_emp} não encontrados. Usando valores padrão.")
            dados_empresa = {'nome': 'N/A', 'cnpj': 'N/A', 'insc_estadual': 'N/A', 'insc_municipal': 'N/A'}
        
        print("=== DADOS DA EMPRESA PARA O RELATÓRIO ===")
        print(json.dumps(dados_empresa, indent=2))
        
        parametros_relatorio = {
            "empresa": codi_emp,
            "data_inicial": data_inicio,
            "data_final": data_fim
        }

        # Definindo nomes de arquivo de forma padronizada
        nome_pdf = "LivroISS.pdf"
        nome_xlsx = "LivroISS.xlsx"

        if exportar_pdf:
            print(f"\nIniciando a geração do PDF: {nome_pdf}...")
            gerar_pdf(parametros_relatorio, livro_ant, documentos, totais, dados_empresa)
            arquivos_gerados.append(nome_pdf)

        if exportar_xlsx:
            print(f"\nIniciando a geração do XLSX: {nome_xlsx}...")
            gerar_xlsx_simples(parametros_relatorio, livro_ant, documentos, totais, dados_empresa)
            arquivos_gerados.append(nome_xlsx)
            
        return arquivos_gerados

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"ERRO DE BANCO DE DADOS: {sqlstate}")
        print(ex)
        return []
    except Exception as e:
        print(f"UM ERRO INESPERADO OCORREU: {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        if conn:
            conn.close()
            print("\nConexão com o banco de dados fechada.")








gerar_livro_iss(962, data_inicio='2022-01-01', data_fim='2022-12-31', exportar_pdf=True, exportar_xlsx=True)