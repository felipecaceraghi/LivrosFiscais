import pyodbc
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from decimal import Decimal, getcontext
from datetime import datetime, date
from collections import defaultdict
import concurrent.futures
import json
import traceback

# --- IMPORTAÇÃO PARA O EXCEL ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIGURAÇÃO INICIAL ---
getcontext().prec = 10
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

# --- QUERIES (Permanecem inalteradas) ---
SQL_DADOS_CABECALHO = "SELECT G.IEST_EMP, G.RAZAO_EMP, G.CGCE_EMP FROM BETHADBA.GEEMPRE G WHERE G.CODI_EMP = ?"
SQL_DADOS_PRINCIPAIS = """
-- 1. NOTAS FISCAIS TRADICIONAIS
SELECT
    CAST(S.NUME_SAI AS DECIMAL(15,0)) AS NUMERO,
    S.CODI_SAI AS CODIGO,
    S.DDOC_SAI AS DDOC,
    DAY(S.DDOC_SAI) AS DIA,
    S.VCON_SAI AS VCON,
    S.CODI_NAT AS CFOP,
    COALESCE(S.SIGL_EST, '') AS UF,
    E.NOME_ESP AS ESPECIE,
    S.SERI_SAI AS SERIE,
    S.SEGI_SAI,
    E.MODE_ESP AS MODELO,
    'SAI' AS TIPO_NOTA,
    S.DATA_SAIDA AS DATAORD
FROM BETHADBA.EFSAIDAS AS S
INNER JOIN BETHADBA.EFESPECIES AS E ON S.CODI_ESP = E.CODI_ESP
WHERE S.CODI_EMP = ? AND S.DSAI_SAI BETWEEN ? AND ?
  AND S.CODI_NAT BETWEEN 5000 AND 7999
  AND E.DOCUMENTO_NAO_FISCAL = 'N'

UNION ALL

-- 2. REDUÇÕES Z DE ECF
SELECT
    (CASE WHEN PAR.EMITE_MAPA_RESUMO_ECF_PAR = 'N' THEN RED.CUPOM_INICIAL ELSE RED.MAPA_RESUMO END) AS NUMERO,
    RED.I_REDUCAO AS CODIGO,
    RED.DATA_REDUCAO AS DDOC,
    DAY(RED.DATA_REDUCAO) AS DIA,
    COALESCE(TDVALOR.VALOR_CONTABIL, 0) AS VCON,
    0 AS CFOP,
    TDCLIENTE.SIGL_EST AS UF,
    TDESPECIE.NOME_ESP AS ESPECIE,
    (CASE WHEN PAR.EMITE_MAPA_RESUMO_ECF_PAR = 'N' THEN
        CAST(RIGHT(RED.CODIGO_MAQUINA, 7) AS CHAR(7))
     ELSE CAST('ECF' AS CHAR(7)) END) AS SERIE,
    0 AS SEGI_SAI,
    90 AS MODELO,
    'Z2D' AS TIPO_NOTA,
    RED.DATA_REDUCAO AS DATAORD
FROM BETHADBA.EFECF_REDUCAO_Z AS RED
INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS PAR ON PAR.CODI_EMP = RED.CODI_EMP
    AND PAR.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(RED.CODI_EMP, RED.DATA_REDUCAO),
LATERAL(SELECT MIN(ESP.CODI_ESP) AS CODI_ESP FROM BETHADBA.EFESPECIES AS ESP WHERE ESP.CODIGO_MODELO = '2D') AS TDESPECIE_CODI,
LATERAL(SELECT MIN(ESP.NOME_ESP) AS NOME_ESP FROM BETHADBA.EFESPECIES AS ESP WHERE ESP.CODI_ESP = TDESPECIE_CODI.CODI_ESP) AS TDESPECIE,
LATERAL(SELECT CLI.SIGL_EST FROM BETHADBA.EFCLIENTES AS CLI WHERE CLI.CODI_EMP = RED.CODI_EMP AND CLI.CODI_CLI = PAR.CLIENTE_REDUCAO_Z_PAR) AS TDCLIENTE,
LATERAL(SELECT SUM(REDUCAO_Z_IMPOSTOS.VALOR_CONTABIL) AS VALOR_CONTABIL
        FROM BETHADBA.EFECF_REDUCAO_Z_IMPOSTOS AS REDUCAO_Z_IMPOSTOS
        WHERE REDUCAO_Z_IMPOSTOS.CODI_EMP = RED.CODI_EMP
          AND REDUCAO_Z_IMPOSTOS.I_REDUCAO = RED.I_REDUCAO
          AND REDUCAO_Z_IMPOSTOS.CODI_IMP IN (1, 3)) AS TDVALOR
WHERE RED.CODI_EMP = ? AND RED.DATA_REDUCAO BETWEEN ? AND ?

UNION ALL

-- 3. RESUMO MOVIMENTO DIÁRIO
SELECT
    MOVIMENTO_DIARIO.NUMERO AS NUMERO,
    MOVIMENTO_DIARIO.I_RESUMO AS CODIGO,
    MOVIMENTO_DIARIO.DATA_EMISSAO AS DDOC,
    DAY(MOVIMENTO_DIARIO.DATA_EMISSAO) AS DIA,
    SUM(MOVIMENTO_DIARIO_BILHETE.VALOR_TOTAL) AS VCON,
    MOVIMENTO_DIARIO_BILHETE.CODI_NAT AS CFOP,
    MOVIMENTO_DIARIO_BILHETE.UF_DESTINO AS UF,
    'RMD' AS ESPECIE,
    MOVIMENTO_DIARIO.SERIE AS SERIE,
    0 AS SEGI_SAI,
    0 AS MODELO,
    'MOV' AS TIPO_NOTA,
    MOVIMENTO_DIARIO.DATA_EMISSAO AS DATAORD
FROM BETHADBA.EFRESUMO_MOVIMENTO_DIARIO AS MOVIMENTO_DIARIO
INNER JOIN BETHADBA.EFRESUMO_MOVIMENTO_DIARIO_BILHETE AS MOVIMENTO_DIARIO_BILHETE
    ON MOVIMENTO_DIARIO_BILHETE.CODI_EMP = MOVIMENTO_DIARIO.CODI_EMP
    AND MOVIMENTO_DIARIO_BILHETE.I_RESUMO = MOVIMENTO_DIARIO.I_RESUMO
INNER JOIN BETHADBA.EFESPECIES AS EFESPECIES ON EFESPECIES.CODI_ESP = MOVIMENTO_DIARIO_BILHETE.CODI_ESP
INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS EFPARAMETRO_VIGENCIA ON EFPARAMETRO_VIGENCIA.CODI_EMP = MOVIMENTO_DIARIO.CODI_EMP
    AND EFPARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(MOVIMENTO_DIARIO.CODI_EMP, MOVIMENTO_DIARIO.DATA_EMISSAO)
WHERE MOVIMENTO_DIARIO.CODI_EMP = ? AND MOVIMENTO_DIARIO.DATA_EMISSAO BETWEEN ? AND ?
  AND EFESPECIES.DOCUMENTO_NAO_FISCAL = 'N'
GROUP BY MOVIMENTO_DIARIO.DATA_EMISSAO, MOVIMENTO_DIARIO.NUMERO, MOVIMENTO_DIARIO.I_RESUMO,
         MOVIMENTO_DIARIO.SERIE, MOVIMENTO_DIARIO_BILHETE.CODI_NAT, MOVIMENTO_DIARIO_BILHETE.UF_DESTINO

UNION ALL

-- 4. BILHETES DE PASSAGEM
SELECT
    EFBILHETE_PASSAGEM.DOCUMENTO_INICIAL AS NUMERO,
    EFBILHETE_PASSAGEM.I_BILHETE AS CODIGO,
    EFBILHETE_PASSAGEM.DATA_EMISSAO AS DDOC,
    DAY(EFBILHETE_PASSAGEM.DATA_EMISSAO) AS DIA,
    EFBILHETE_PASSAGEM.VALOR_TOTAL AS VCON,
    EFBILHETE_PASSAGEM.CODI_NAT AS CFOP,
    EFBILHETE_PASSAGEM.UF_DESTINO AS UF,
    EFESPECIES.NOME_ESP AS ESPECIE,
    EFBILHETE_PASSAGEM.SERIE AS SERIE,
    0 AS SEGI_SAI,
    EFESPECIES.MODE_ESP AS MODELO,
    'BPA' AS TIPO_NOTA,
    EFBILHETE_PASSAGEM.DATA_EMISSAO AS DATAORD
FROM BETHADBA.EFBILHETE_PASSAGEM AS EFBILHETE_PASSAGEM
INNER JOIN BETHADBA.EFESPECIES AS EFESPECIES ON EFESPECIES.CODI_ESP = EFBILHETE_PASSAGEM.CODI_ESP
INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS EFPARAMETRO_VIGENCIA ON EFPARAMETRO_VIGENCIA.CODI_EMP = EFBILHETE_PASSAGEM.CODI_EMP
    AND EFPARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(EFBILHETE_PASSAGEM.CODI_EMP, EFBILHETE_PASSAGEM.DATA_EMISSAO)
WHERE EFBILHETE_PASSAGEM.CODI_EMP = ? AND EFBILHETE_PASSAGEM.DATA_EMISSAO BETWEEN ? AND ?
  AND EFESPECIES.DOCUMENTO_NAO_FISCAL = 'N'

UNION ALL

-- 5. CUPONS FISCAIS ELETRÔNICOS (CF-e SAT)
SELECT
    CUPOM.EXTRATO AS NUMERO,
    CUPOM.I_CFE AS CODIGO,
    CUPOM.DATA_CFE AS DDOC,
    DAY(CUPOM.DATA_CFE) AS DIA,
    SUM(COALESCE(TDICMS.VALOR_CONTABIL, 0) + COALESCE(TDISS.VALOR_CONTABIL, 0)) AS VCON,
    0 AS CFOP,
    MAX(EMPRESA.ESTA_EMP) AS UF,
    'CF-e-SAT' AS ESPECIE,
    SAT_CFE.NUMERO_SERIE AS SERIE,
    0 AS SEGI_SAI,
    59 AS MODELO,
    'CFE' AS TIPO_NOTA,
    CUPOM.DATA_CFE AS DATAORD
FROM BETHADBA.EFCUPOM_FISCAL_ELETRONICO AS CUPOM
INNER JOIN BETHADBA.EFSAT_CF_E AS SAT_CFE ON SAT_CFE.CODI_EMP = CUPOM.CODI_EMP AND SAT_CFE.I_SAT_CF_E = CUPOM.SAT
INNER JOIN BETHADBA.EFPARAMETRO AS PARAMETRO ON PARAMETRO.CODI_EMP = CUPOM.CODI_EMP
INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS PARAMETRO_VIGENCIA ON PARAMETRO_VIGENCIA.CODI_EMP = CUPOM.CODI_EMP
    AND PARAMETRO_VIGENCIA.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(CUPOM.CODI_EMP, CUPOM.DATA_CFE)
INNER JOIN BETHADBA.GEEMPRE AS EMPRESA ON CUPOM.CODI_EMP = EMPRESA.CODI_EMP,
LATERAL(SELECT COALESCE(SUM(IMPOSTOS_DETALHAMENTO.VALOR_CONTABIL), 0) AS VALOR_CONTABIL
        FROM BETHADBA.EFCUPOM_FISCAL_ELETRONICO_IMPOSTOS AS IMPOSTOS
        LEFT JOIN BETHADBA.EFCUPOM_FISCAL_ELETRONICO_IMPOSTOS_DETALHAMENTO AS IMPOSTOS_DETALHAMENTO
            ON IMPOSTOS_DETALHAMENTO.CODI_EMP = IMPOSTOS.CODI_EMP
            AND IMPOSTOS_DETALHAMENTO.I_CFE = IMPOSTOS.I_CFE
            AND IMPOSTOS_DETALHAMENTO.I_IMPOSTO = IMPOSTOS.I_IMPOSTO
        WHERE IMPOSTOS.CODI_EMP = CUPOM.CODI_EMP
          AND IMPOSTOS.I_CFE = CUPOM.I_CFE
          AND IMPOSTOS_DETALHAMENTO.CODI_IMP = 1) AS TDICMS,
LATERAL(SELECT COALESCE(SUM(IMPOSTOS_DETALHAMENTO.VALOR_CONTABIL), 0) AS VALOR_CONTABIL
        FROM BETHADBA.EFCUPOM_FISCAL_ELETRONICO_IMPOSTOS AS IMPOSTOS
        LEFT JOIN BETHADBA.EFCUPOM_FISCAL_ELETRONICO_IMPOSTOS_DETALHAMENTO AS IMPOSTOS_DETALHAMENTO
            ON IMPOSTOS_DETALHAMENTO.CODI_EMP = IMPOSTOS.CODI_EMP
            AND IMPOSTOS_DETALHAMENTO.I_CFE = IMPOSTOS.I_CFE
            AND IMPOSTOS_DETALHAMENTO.I_IMPOSTO = IMPOSTOS.I_IMPOSTO
        WHERE IMPOSTOS.CODI_EMP = CUPOM.CODI_EMP
          AND IMPOSTOS.I_CFE = CUPOM.I_CFE
          AND IMPOSTOS_DETALHAMENTO.CODI_IMP = 3) AS TDISS
WHERE CUPOM.CODI_EMP = ? AND CUPOM.DATA_CFE BETWEEN ? AND ?
  AND PARAMETRO.EMITE_CF_E_PAR = 'S'
GROUP BY CUPOM.DATA_CFE, CUPOM.EXTRATO, CUPOM.I_CFE, SAT_CFE.NUMERO_SERIE

ORDER BY DATAORD, MODELO, SERIE, NUMERO, CODIGO, SEGI_SAI;
"""
SQL_DETALHES_IMPOSTOS = """
-- Esta consulta unifica os detalhes de impostos e CFOP de diferentes tipos de documentos

-- 1. IMPOSTOS DE NOTAS FISCAIS TRADICIONAIS (EFSAIDAS)
SELECT
    I.CODI_SAI AS DOCUMENTO_ID,
    'SAI' AS TIPO_DOCUMENTO,
    I.CODI_IMP AS IMPOSTO_ID,
    S.CODI_NAT AS CFOP,
    I.VCON_ISA AS VALOR_CONTABIL_IMPOSTO,
    I.BCAL_ISA AS BASE_CALCULO,
    I.ALIQ_ISA AS ALIQUOTA,
    I.VLOR_ISA AS VALOR_IMPOSTO,
    (COALESCE(I.VISE_ISA, 0) + COALESCE(I.NTRI_ISA, 0)) AS ISENTAS_NT,
    I.VOUT_ISA AS OUTRAS
FROM BETHADBA.EFIMPSAI AS I
INNER JOIN BETHADBA.EFSAIDAS AS S ON I.CODI_EMP = S.CODI_EMP AND I.CODI_SAI = S.CODI_SAI
WHERE I.CODI_EMP = ? AND S.DSAI_SAI BETWEEN ? AND ?

UNION ALL

-- 2. IMPOSTOS DE REDUÇÕES Z (ECF)
SELECT
    RED_IMP.I_REDUCAO AS DOCUMENTO_ID,
    'Z2D' AS TIPO_DOCUMENTO,
    RED_IMP.CODI_IMP AS IMPOSTO_ID,
    RED_IMP.CODI_NAT AS CFOP,
    RED_IMP.VALOR_CONTABIL AS VALOR_CONTABIL_IMPOSTO,
    RED_IMP.BASE_CALCULO AS BASE_CALCULO,
    RED_IMP.ALIQUOTA AS ALIQUOTA,
    RED_IMP.IMPOSTO AS VALOR_IMPOSTO,
    (COALESCE(RED_IMP.ISENTAS, 0) + COALESCE(RED_IMP.NAO_TRIBUTADAS, 0)) AS ISENTAS_NT,
    RED_IMP.OUTROS AS OUTRAS
FROM BETHADBA.EFECF_REDUCAO_Z_IMPOSTOS AS RED_IMP
INNER JOIN BETHADBA.EFECF_REDUCAO_Z AS RED ON RED_IMP.CODI_EMP = RED.CODI_EMP AND RED_IMP.I_REDUCAO = RED.I_REDUCAO
WHERE RED_IMP.CODI_EMP = ? AND RED.DATA_REDUCAO BETWEEN ? AND ?
  AND RED_IMP.CODI_IMP IN (1, 2, 3) -- ICMS, IPI, ISS

UNION ALL

-- 3. IMPOSTOS DE RESUMO MOVIMENTO DIÁRIO
SELECT
    MDB.I_RESUMO AS DOCUMENTO_ID,
    'MOV' AS TIPO_DOCUMENTO,
    MDBI.CODI_IMP AS IMPOSTO_ID,
    MDB.CODI_NAT AS CFOP,
    SUM(MDBI.VALOR_CONTABIL) AS VALOR_CONTABIL_IMPOSTO,
    SUM(MDBI.BASE_CALCULO) AS BASE_CALCULO,
    AVG(MDBI.ALIQUOTA) AS ALIQUOTA,
    SUM(MDBI.VALOR_IMPOSTO) AS VALOR_IMPOSTO,
    SUM(MDBI.ISENTAS) AS ISENTAS_NT,
    SUM(MDBI.OUTROS) AS OUTRAS
FROM BETHADBA.EFRESUMO_MOVIMENTO_DIARIO_BILHETE AS MDB
INNER JOIN BETHADBA.EFRESUMO_MOVIMENTO_DIARIO_BILHETE_IMPOSTOS AS MDBI ON MDB.CODI_EMP = MDBI.CODI_EMP AND MDB.I_RESUMO = MDBI.I_RESUMO AND MDB.I_BILHETE = MDBI.I_BILHETE
INNER JOIN BETHADBA.EFRESUMO_MOVIMENTO_DIARIO MD ON MDB.CODI_EMP = MD.CODI_EMP AND MDB.I_RESUMO = MD.I_RESUMO
WHERE MDB.CODI_EMP = ? AND MD.DATA_EMISSAO BETWEEN ? AND ?
  AND MDBI.CODI_IMP IN (1, 2, 3)
GROUP BY MDB.I_RESUMO, MDBI.CODI_IMP, MDB.CODI_NAT

UNION ALL

-- 4. IMPOSTOS DE BILHETES DE PASSAGEM
SELECT
    IMPOSTOS_BP.I_BILHETE AS DOCUMENTO_ID,
    'BPA' AS TIPO_DOCUMENTO,
    IMPOSTOS_BP.CODI_IMP AS IMPOSTO_ID,
    BP.CODI_NAT AS CFOP,
    IMPOSTOS_BP.VALOR_CONTABIL AS VALOR_CONTABIL_IMPOSTO,
    IMPOSTOS_BP.BASE_CALCULO AS BASE_CALCULO,
    IMPOSTOS_BP.ALIQUOTA AS ALIQUOTA,
    IMPOSTOS_BP.VALOR_IMPOSTO AS VALOR_IMPOSTO,
    IMPOSTOS_BP.ISENTAS AS ISENTAS_NT,
    IMPOSTOS_BP.OUTROS AS OUTRAS
FROM BETHADBA.EFBILHETE_PASSAGEM_IMPOSTOS AS IMPOSTOS_BP
INNER JOIN BETHADBA.EFBILHETE_PASSAGEM AS BP ON IMPOSTOS_BP.CODI_EMP = BP.CODI_EMP AND IMPOSTOS_BP.I_BILHETE = BP.I_BILHETE
WHERE IMPOSTOS_BP.CODI_EMP = ? AND BP.DATA_EMISSAO BETWEEN ? AND ?
  AND IMPOSTOS_BP.CODI_IMP IN (1, 2, 3)

UNION ALL

-- 5. IMPOSTOS DE CUPONS FISCAIS ELETRÔNICOS (CF-e SAT)
SELECT
    IMPOSTOS.I_CFE AS DOCUMENTO_ID,
    'CFE' AS TIPO_DOCUMENTO,
    IMPOSTOS_DETALHAMENTO.CODI_IMP AS IMPOSTO_ID,
    IMPOSTOS_DETALHAMENTO.CODI_NAT AS CFOP,
    SUM(IMPOSTOS_DETALHAMENTO.VALOR_CONTABIL) AS VALOR_CONTABIL_IMPOSTO,
    SUM(IMPOSTOS_DETALHAMENTO.BASE_CALCULO) AS BASE_CALCULO,
    AVG(IMPOSTOS_DETALHAMENTO.ALIQUOTA) AS ALIQUOTA,
    SUM(IMPOSTOS_DETALHAMENTO.VALOR_IMPOSTO) AS VALOR_IMPOSTO,
    SUM(COALESCE(IMPOSTOS_DETALHAMENTO.VALOR_ISENTAS, 0)) AS ISENTAS_NT,
    SUM(COALESCE(IMPOSTOS_DETALHAMENTO.VALOR_OUTROS, 0)) AS OUTRAS
FROM BETHADBA.EFCUPOM_FISCAL_ELETRONICO AS CUPOM
INNER JOIN BETHADBA.EFCUPOM_FISCAL_ELETRONICO_IMPOSTOS AS IMPOSTOS ON IMPOSTOS.CODI_EMP = CUPOM.CODI_EMP AND IMPOSTOS.I_CFE = CUPOM.I_CFE
INNER JOIN BETHADBA.EFCUPOM_FISCAL_ELETRONICO_IMPOSTOS_DETALHAMENTO AS IMPOSTOS_DETALHAMENTO ON IMPOSTOS_DETALHAMENTO.CODI_EMP = IMPOSTOS.CODI_EMP AND IMPOSTOS_DETALHAMENTO.I_CFE = IMPOSTOS.I_CFE AND IMPOSTOS_DETALHAMENTO.I_IMPOSTO = IMPOSTOS.I_IMPOSTO
WHERE CUPOM.CODI_EMP = ? AND CUPOM.DATA_CFE BETWEEN ? AND ?
  AND IMPOSTOS_DETALHAMENTO.CODI_IMP IN (1, 2, 3)
GROUP BY IMPOSTOS.I_CFE, IMPOSTOS_DETALHAMENTO.CODI_IMP, IMPOSTOS_DETALHAMENTO.CODI_NAT;
"""

# --- FUNÇÕES AUXILIARES E DE GERAÇÃO (Permanecem inalteradas) ---
def executar_consulta(sql, params=()):
    conn = None
    try:
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        cursor.execute(sql, params)
        try:
            colunas = [coluna[0] for coluna in cursor.description]
            return [dict(zip(colunas, linha)) for linha in cursor.fetchall()]
        except TypeError:
            return []
    except Exception as e:
        print(f"Erro ao executar a consulta: {e}")
        return None
    finally:
        if conn:
            conn.close()

def formatar_valor(valor):
    if valor is None or valor == 0: return '0,00'
    valor_decimal = Decimal(valor)
    return f"{valor_decimal:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_cnpj(cnpj):
    if not cnpj or not cnpj.isdigit() or len(cnpj) != 14: return cnpj
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, (date, datetime)):
            return obj.isoformat()
        elif isinstance(obj, defaultdict):
            return dict(obj)
        return super(DecimalEncoder, self).default(obj)

def converter_para_json_serializavel(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, (date, datetime)):
        return obj.isoformat()
    elif isinstance(obj, defaultdict):
        return converter_para_json_serializavel(dict(obj))
    elif isinstance(obj, dict):
        return {k: converter_para_json_serializavel(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [converter_para_json_serializavel(item) for item in obj]
    else:
        return obj

def salvar_dados_json(nome_arquivo, dados_completos):
    try:
        dados_serializaveis = converter_para_json_serializavel(dados_completos)
        with open(nome_arquivo, 'w', encoding='utf-8') as arquivo_json:
            json.dump(dados_serializaveis, arquivo_json,
                     ensure_ascii=False,
                     indent=2,
                     sort_keys=True)
        print(f"✅ JSON '{nome_arquivo}' gerado com sucesso.")
    except Exception as e:
        print(f"❌ Erro ao salvar JSON: {e}")

class CabecalhoHelper:
    def __init__(self, dados_cabecalho_empresa): self.dados_cabecalho = dados_cabecalho_empresa
    def cabecalho_pagina(self, canvas, doc):
        canvas.saveState()
        largura, altura = landscape(A4); margem_esquerda = 10*mm; margem_topo = 15*mm
        titulo_style = ParagraphStyle('TituloPrincipal', fontSize=11, alignment=TA_CENTER, fontName='Helvetica-Bold')
        subtitulo_style = ParagraphStyle('Subtitulo', fontSize=10, alignment=TA_CENTER, fontName='Helvetica-Bold')
        p_titulo = Paragraph("LIVRO REGISTRO DE SAÍDAS - RS - MODELO P2", titulo_style)
        p_subtitulo = Paragraph("REGISTRO DE SAÍDAS", subtitulo_style)
        largura_titulo, altura_titulo = p_titulo.wrapOn(canvas, largura - 2*margem_esquerda, altura)
        p_titulo.drawOn(canvas, margem_esquerda, altura - margem_topo - altura_titulo)
        largura_subtitulo, altura_subtitulo = p_subtitulo.wrapOn(canvas, largura - 2*margem_esquerda, altura)
        p_subtitulo.drawOn(canvas, margem_esquerda, altura - margem_topo - altura_titulo - altura_subtitulo - 1*mm)
        cnpj_formatado = formatar_cnpj(self.dados_cabecalho['cnpj']); folha_formatada = f"{doc.page:03d}"
        dados_empresa_tabela = [['EMPRESA:', self.dados_cabecalho['nome'], 'CNPJ:', cnpj_formatado], ['INSC.EST.:', self.dados_cabecalho['insc_est'], '', ''], ['FOLHA:', folha_formatada, 'MÊS OU PERÍODO/ANO:', f"{self.dados_cabecalho['periodo_inicio']} a {self.dados_cabecalho['periodo_fim']}"]]
        larguras_empresa = [25*mm, 115*mm, 40*mm, 97*mm]
        tabela_empresa = Table(dados_empresa_tabela, colWidths=larguras_empresa, rowHeights=5*mm)
        tabela_empresa.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, colors.black), ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('FONTSIZE', (0, 0), (-1, -1), 8), ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('LEFTPADDING', (0, 0), (-1, -1), 2), ('SPAN', (1, 1), (3, 1)),]))
        largura_tabela_empresa, altura_tabela_empresa = tabela_empresa.wrapOn(canvas, largura - 2*margem_esquerda, altura)
        tabela_empresa.drawOn(canvas, margem_esquerda, altura - margem_topo - altura_titulo - altura_subtitulo - altura_tabela_empresa - 3*mm)
        canvas.restoreState()

def criar_pdf_livro_saidas(nome_arquivo, dados_cabecalho, documentos, totais_cfop, totais_gerais, totais_uf):
    try:
        helper_cabecalho = CabecalhoHelper(dados_cabecalho)
        doc_template = SimpleDocTemplate(nome_arquivo, pagesize=landscape(A4), topMargin=45*mm, bottomMargin=15*mm, leftMargin=10*mm, rightMargin=10*mm)
        elementos = []
        
        cabecalho_principal_dados = [['DOCUMENTOS FISCAIS', None, None, None, None, 'VALOR\nCONTÁBIL', 'CODIFICAÇÃO', None, 'VALORES FISCAIS', None, None, None, None, None, 'OBSERVAÇÕES'], [None, None, None, None, None, None, None, None, 'OPERAÇÕES COM DÉBITO DO IMPOSTO', None, None, 'OPERAÇÕES SEM DÉBITO DO IMPOSTO', None, None, None], ['ESPÉCIE', 'SÉRIE/\nSUB\nSÉRIE', 'NÚMERO/ATÉ', 'DIA', 'UF\nDEST', None, 'CONTA\nBIL', 'FISCAL', 'ICMS', 'BASE DE CÁLCULO', 'ALIQ.', 'IMP.DEBITADO', 'ISENTAS OU\nNÃO TRIBUTADAS', 'OUTRAS', None], [None, None, None, None, None, None, None, None, 'IPI', None, None, None, None, None, None]]
        dados_tabela_principal = []; estilo_dinamico_principal = []; current_row = len(cabecalho_principal_dados)
        for documento in documentos:
            icms, ipi = documento['impostos'].get(1), documento['impostos'].get(2)
            if not documento['impostos']: continue 
            
            dia, especie, serie = documento.get('DIA', ''), documento.get('ESPECIE', 'Nota Fis')[:9], documento.get('SERIE', '1')
            cfop_formatado = f"{documento['CFOP'] // 1000}.{documento['CFOP'] % 1000}" if documento.get('CFOP', 0) > 0 else ''
            dados_doc = [especie, serie, documento['NUMERO'], dia, documento['UF'], formatar_valor(documento['VCON']), '', cfop_formatado]
            
            linhas_impostos = []
            if icms:
                linhas_impostos.append(['ICMS', formatar_valor(icms.get('BASE_CALCULO', 0)), formatar_valor(icms.get('ALIQUOTA', 0)), formatar_valor(icms.get('VALOR_IMPOSTO', 0)), formatar_valor(icms.get('ISENTAS_NT', 0)), formatar_valor(icms.get('OUTRAS', 0)), ''])
            if ipi:
                linhas_impostos.append(['IPI', formatar_valor(ipi.get('BASE_CALCULO', 0)), formatar_valor(ipi.get('ALIQUOTA', 0)), formatar_valor(ipi.get('VALOR_IMPOSTO', 0)), formatar_valor(ipi.get('ISENTAS_NT', 0)), formatar_valor(ipi.get('OUTRAS', 0)), ''])
            
            if not linhas_impostos and documento['impostos']:
                imposto_qualquer = next(iter(documento['impostos'].values()))
                linhas_impostos.append(['OUTROS', formatar_valor(imposto_qualquer.get('BASE_CALCULO', 0)), formatar_valor(imposto_qualquer.get('ALIQUOTA', 0)), formatar_valor(imposto_qualquer.get('VALOR_IMPOSTO', 0)), formatar_valor(imposto_qualquer.get('ISENTAS_NT', 0)), formatar_valor(imposto_qualquer.get('OUTRAS', 0)), ''])

            for i, linha_imposto in enumerate(linhas_impostos):
                prefixo = dados_doc if i == 0 else [''] * 8
                dados_tabela_principal.append(prefixo + linha_imposto)

            num_linhas_imposto = len(linhas_impostos)
            if num_linhas_imposto > 1:
                end_row = current_row + num_linhas_imposto - 1
                for col in [0, 1, 2, 3, 4, 5, 6, 7, 14]: estilo_dinamico_principal.append(('SPAN', (col, current_row), (col, end_row)))
            current_row += num_linhas_imposto

        larguras_tabela = [18*mm, 15*mm, 20*mm, 10*mm, 8*mm, 18*mm, 16*mm, 16*mm, 10*mm, 24*mm, 12*mm, 20*mm, 24*mm, 19*mm, 35*mm]
        tabela_principal = Table(cabecalho_principal_dados + dados_tabela_principal, colWidths=larguras_tabela, repeatRows=4)
        estilo_base = [('BOX', (0, 0), (-1, -1), 1, colors.black), ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('LEFTPADDING', (0,0), (-1,-1), 1), ('RIGHTPADDING', (0,0), (-1,-1), 1), ('FONTSIZE', (0, 0), (-1, 3), 6), ('FONTSIZE', (0, 4), (-1, -1), 7), ('SPAN', (0, 0), (4, 1)), ('SPAN', (6, 0), (7, 1)), ('SPAN', (8, 0), (13, 0)), ('SPAN', (8, 1), (10, 1)), ('SPAN', (11, 1), (13, 1)), ('SPAN', (5, 0), (5, 3)), ('SPAN', (14, 0), (14, 3)), ('SPAN', (0, 2), (0, 3)), ('SPAN', (1, 2), (1, 3)), ('SPAN', (2, 2), (2, 3)), ('SPAN', (3, 2), (3, 3)), ('SPAN', (4, 2), (4, 3)), ('SPAN', (6, 2), (6, 3)), ('SPAN', (7, 2), (7, 3)), ('SPAN', (9, 2), (9, 3)), ('SPAN', (10, 2), (10, 3)), ('SPAN', (11, 2), (11, 3)), ('SPAN', (12, 2), (12, 3)), ('SPAN', (13, 2), (13, 3)), ('ALIGN', (0, 0), (-1, 3), 'CENTER'), ('FONTNAME', (0, 0), (-1, 3), 'Helvetica-Bold'), ('FONTNAME', (0, 4), (-1, -1), 'Helvetica'), ('ALIGN', (1, 4), (4, -1), 'CENTER'), ('ALIGN', (5, 4), (5, -1), 'RIGHT'), ('ALIGN', (9, 4), (-2, -1), 'RIGHT'),]
        tabela_principal.setStyle(TableStyle(estilo_base + estilo_dinamico_principal))
        elementos.append(tabela_principal)

        dados_totais_mensais = []; estilo_totais_mensais = []; total_labels = ['TOTAL ICMS MENSAL', 'TOTAL IPI MENSAL', 'TOTAL S.T. MENSAL']; total_data = [totais_gerais['icms'], totais_gerais['ipi'], {'VCON':0, 'BASE_CALCULO':0, 'VALOR_IMPOSTO':0, 'ISENTAS_NT':0, 'OUTRAS':0}]
        for i, label in enumerate(total_labels):
            dados = total_data[i]
            linha_total = [label, None, None, None, None, formatar_valor(dados['VCON']), None, None, None, formatar_valor(dados['BASE_CALCULO']), None, formatar_valor(dados['VALOR_IMPOSTO']), formatar_valor(dados['ISENTAS_NT']), formatar_valor(dados['OUTRAS']), None]
            dados_totais_mensais.append(linha_total); estilo_totais_mensais.extend([('SPAN', (0, i), (4, i)), ('SPAN', (6, i), (8, i)), ('SPAN', (10, i), (10, i)), ('SPAN', (14, i), (14, i)), ('ALIGN', (0, i), (0, i), 'LEFT'), ('FONTNAME', (0, i), (0, i), 'Helvetica-Bold')])
        tabela_totais_mensais = Table(dados_totais_mensais, colWidths=larguras_tabela)
        tabela_totais_mensais.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, colors.black), ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black), ('FONTSIZE', (0, 0), (-1, -1), 7), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('ALIGN', (5, 0), (5, -1), 'RIGHT'), ('ALIGN', (9, 0), (-2, -1), 'RIGHT'),] + estilo_totais_mensais))
        elementos.append(tabela_totais_mensais)
        elementos.append(Spacer(1, 4*mm))

        dados_demonstrativo_uf = [['DEMONSTRATIVO POR ESTADO', None, None, None, None, None, None, None, None, None, None, None, None, None, None]]
        estilo_demonstrativo_uf = [('SPAN', (0, 0), (-1, 0)), ('ALIGN', (0, 0), (0, 0), 'LEFT'), ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold')]
        
        current_row_uf = 1
        for uf, valores in sorted(totais_uf.items()):
            linha_uf = ['', '', '', '', uf, formatar_valor(valores['VCON']), '', '', '', formatar_valor(valores['BASE_CALCULO']), formatar_valor(valores['ALIQUOTA_MEDIA']), formatar_valor(valores['VALOR_IMPOSTO']), formatar_valor(valores['ISENTAS_NT']), formatar_valor(valores['OUTRAS']), '']
            dados_demonstrativo_uf.append(linha_uf)
            estilo_demonstrativo_uf.extend([('SPAN', (0, current_row_uf), (3, current_row_uf)), ('SPAN', (6, current_row_uf), (8, current_row_uf)), ('SPAN', (10, current_row_uf), (10, current_row_uf)), ('ALIGN', (4, current_row_uf), (4, current_row_uf), 'CENTER'),])
            current_row_uf += 1
        
        base_style_demo = [('BOX', (0, 0), (-1, -1), 1, colors.black), ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black), ('FONTSIZE', (0, 0), (-1, -1), 7), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('ALIGN', (5, 1), (5, -1), 'RIGHT'), ('ALIGN', (9, 1), (-2, -1), 'RIGHT'),]
        tabela_demonstrativo_uf = Table(dados_demonstrativo_uf, colWidths=larguras_tabela)
        tabela_demonstrativo_uf.setStyle(TableStyle(base_style_demo + estilo_demonstrativo_uf))
        elementos.append(tabela_demonstrativo_uf)
        elementos.append(Spacer(1, 4*mm))

        dados_demonstrativo_cfop = [['DEMONSTRATIVO POR CFOP', None, None, None, None, None, None, None, None, None, None, None, None, None, None]]
        estilo_demonstrativo_cfop = [('SPAN', (0, 0), (-1, 0)), ('ALIGN', (0, 0), (0, 0), 'LEFT'), ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold')]
        current_row_demo = 1
        for cfop, valores in sorted(totais_cfop.items()):
            linha_cfop = ['', '', '', '', f"{cfop // 1000}.{cfop % 1000}", formatar_valor(valores['VCON']), '', '', '', formatar_valor(valores['BASE_CALCULO']), formatar_valor(valores['ALIQUOTA_MEDIA']), formatar_valor(valores['VALOR_IMPOSTO']), formatar_valor(valores['ISENTAS_NT']), formatar_valor(valores['OUTRAS']), '']
            dados_demonstrativo_cfop.append(linha_cfop)
            estilo_demonstrativo_cfop.extend([('SPAN', (0, current_row_demo), (3, current_row_demo)), ('SPAN', (6, current_row_demo), (8, current_row_demo)), ('SPAN', (10, current_row_demo), (10, current_row_demo)), ('ALIGN', (4, current_row_demo), (4, current_row_demo), 'CENTER'),])
            current_row_demo += 1
        tabela_demonstrativo_cfop = Table(dados_demonstrativo_cfop, colWidths=larguras_tabela)
        tabela_demonstrativo_cfop.setStyle(TableStyle(base_style_demo + estilo_demonstrativo_cfop))
        elementos.append(tabela_demonstrativo_cfop)
        
        doc_template.build(elementos, onFirstPage=helper_cabecalho.cabecalho_pagina, onLaterPages=helper_cabecalho.cabecalho_pagina)
        print(f"✅ PDF '{nome_arquivo}' gerado com sucesso.")
    except Exception as e:
        print(f"❌ Erro ao gerar PDF '{nome_arquivo}': {e}")
        traceback.print_exc()

def criar_xlsx_livro_saidas(nome_arquivo, dados_cabecalho, documentos, totais_cfop, totais_gerais, totais_uf):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Livro Saídas"
        ws.sheet_view.showGridLines = False

        font_bold = Font(name='Calibri', size=11, bold=True)
        font_bold_small = Font(name='Calibri', size=8, bold=True)
        font_normal_small = Font(name='Calibri', size=9)
        align_center_middle = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left_middle = Alignment(horizontal='left', vertical='center')
        align_right_middle = Alignment(horizontal='right', vertical='center')
        num_format_brl = '#,##0.00'
        num_format_decimal = '0.00'
        gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        
        larguras_base = [3, 20, 15, 22, 11, 8, 20, 18, 18, 12, 26, 14, 22, 26, 21, 38]
        for i, width in enumerate(larguras_base):
            ws.column_dimensions[get_column_letter(i + 1)].width = width

        current_row = 1
        col_offset = 1 # Começa a escrever na coluna B

        ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=15 + col_offset)
        cell = ws.cell(row=current_row, column=1 + col_offset, value="LIVRO REGISTRO DE SAÍDAS - RS - MODELO P2")
        cell.font = font_bold; cell.alignment = align_center_middle
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=15 + col_offset)
        cell = ws.cell(row=current_row, column=1 + col_offset, value="REGISTRO DE SAÍDAS")
        cell.font = font_bold; cell.alignment = align_center_middle
        current_row += 2

        cnpj_formatado = formatar_cnpj(dados_cabecalho['cnpj'])
        ws.cell(row=current_row, column=1 + col_offset, value="EMPRESA:").font = font_normal_small
        ws.cell(row=current_row, column=2 + col_offset, value=dados_cabecalho['nome']).font = font_normal_small
        ws.cell(row=current_row, column=10 + col_offset, value="CNPJ:").font = font_normal_small
        ws.cell(row=current_row, column=11 + col_offset, value=cnpj_formatado).font = font_normal_small
        ws.merge_cells(start_row=current_row, start_column=2 + col_offset, end_row=current_row, end_column=9 + col_offset)
        ws.merge_cells(start_row=current_row, start_column=11 + col_offset, end_row=current_row, end_column=15 + col_offset)
        current_row += 1
        ws.cell(row=current_row, column=1 + col_offset, value="INSC.EST.:").font = font_normal_small
        ws.cell(row=current_row, column=2 + col_offset, value=dados_cabecalho['insc_est']).font = font_normal_small
        ws.merge_cells(start_row=current_row, start_column=2 + col_offset, end_row=current_row, end_column=15 + col_offset)
        current_row += 1
        ws.cell(row=current_row, column=1 + col_offset, value="FOLHA:").font = font_normal_small
        ws.cell(row=current_row, column=2 + col_offset, value="001").font = font_normal_small
        ws.cell(row=current_row, column=10 + col_offset, value="MÊS OU PERÍODO/ANO:").font = font_normal_small
        ws.cell(row=current_row, column=11 + col_offset, value=f"{dados_cabecalho['periodo_inicio']} a {dados_cabecalho['periodo_fim']}").font = font_normal_small
        ws.merge_cells(start_row=current_row, start_column=2 + col_offset, end_row=current_row, end_column=9 + col_offset)
        ws.merge_cells(start_row=current_row, start_column=11 + col_offset, end_row=current_row, end_column=15 + col_offset)
        current_row += 2
        
        header_start_row = current_row
        cabecalho_dados = [['DOCUMENTOS FISCAIS', None, None, None, None, '\n\n\nVALOR\nCONTÁBIL', 'CODIFICAÇÃO', None, 'VALORES FISCAIS', None, None, None, None, None, 'OBSERVAÇÕES'], [None, None, None, None, None, None, None, None, 'OPERAÇÕES COM DÉBITO DO IMPOSTO', None, None, 'OPERAÇÕES SEM DÉBITO DO IMPOSTO', None, None, None], ['ESPÉCIE', 'SÉRIE/\nSUB\nSÉRIE', 'NÚMERO/ATÉ', 'DIA', 'UF\nDEST', None, 'CONTA\nBIL', 'FISCAL', 'ICMS', 'BASE DE CÁLCULO', 'ALIQ.', 'IMP.DEBITADO', 'ISENTAS OU\nNÃO TRIBUTADAS', 'OUTRAS', None], [None, None, None, None, None, None, None, None, 'IPI', None, None, None, None, None, None]]
        for r, row_data in enumerate(cabecalho_dados):
            for c, value in enumerate(row_data):
                cell = ws.cell(row=current_row + r, column=c + 1 + col_offset, value=value)
                cell.font = font_bold_small
                cell.alignment = align_center_middle
                cell.fill = gray_fill
        ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row+1, end_column=5 + col_offset)
        ws.merge_cells(start_row=current_row, start_column=7 + col_offset, end_row=current_row+1, end_column=8 + col_offset)
        ws.merge_cells(start_row=current_row, start_column=9 + col_offset, end_row=current_row, end_column=14 + col_offset)
        ws.merge_cells(start_row=current_row+1, start_column=9 + col_offset, end_row=current_row+1, end_column=11 + col_offset)
        ws.merge_cells(start_row=current_row+1, start_column=12 + col_offset, end_row=current_row+1, end_column=14 + col_offset)
        ws.merge_cells(start_row=current_row, start_column=6 + col_offset, end_row=current_row+3, end_column=6 + col_offset)
        ws.merge_cells(start_row=current_row, start_column=15 + col_offset, end_row=current_row+3, end_column=15 + col_offset)
        merges_v = [(3, 1), (3, 2), (3, 3), (3, 4), (3, 5), (3, 7), (3, 8), (3, 10), (3, 11), (3, 12), (3, 13), (3, 14)]
        for r, c in merges_v:
            ws.merge_cells(start_row=current_row+r-1, start_column=c + col_offset, end_row=current_row+r, end_column=c + col_offset)
        current_row += len(cabecalho_dados)

        def write_row(data_list):
            for c, value in enumerate(data_list):
                ws.cell(row=current_row, column=c + 1 + col_offset, value=value)

        for documento in documentos:
            doc_row_start = current_row
            icms, ipi = documento['impostos'].get(1), documento['impostos'].get(2)
            if not documento['impostos']: continue
            
            cfop_formatado = f"{documento['CFOP'] // 1000}.{documento['CFOP'] % 1000}" if documento.get('CFOP', 0) > 0 else ''
            dados_doc_comuns = [documento.get('ESPECIE', 'Nota Fis')[:9], documento.get('SERIE', '1'), documento['NUMERO'], documento.get('DIA', ''), documento['UF'], documento['VCON'], '', cfop_formatado]
            
            row_count_doc = 0
            linhas_impostos_xls = []
            if icms:
                linhas_impostos_xls.append(['ICMS', icms.get('BASE_CALCULO', 0), icms.get('ALIQUOTA', 0), icms.get('VALOR_IMPOSTO', 0), icms.get('ISENTAS_NT', 0), icms.get('OUTRAS', 0), ''])
            if ipi:
                linhas_impostos_xls.append(['IPI', ipi.get('BASE_CALCULO', 0), ipi.get('ALIQUOTA', 0), ipi.get('VALOR_IMPOSTO', 0), ipi.get('ISENTAS_NT', 0), ipi.get('OUTRAS', 0), ''])

            if not linhas_impostos_xls and documento['impostos']:
                imposto_qualquer = next(iter(documento['impostos'].values()))
                linhas_impostos_xls.append(['OUTROS', imposto_qualquer.get('BASE_CALCULO', 0), imposto_qualquer.get('ALIQUOTA', 0), imposto_qualquer.get('VALOR_IMPOSTO', 0), imposto_qualquer.get('ISENTAS_NT', 0), imposto_qualquer.get('OUTRAS', 0), ''])

            for i, linha_imposto in enumerate(linhas_impostos_xls):
                prefixo = dados_doc_comuns if i == 0 else [''] * 8
                write_row(prefixo + linha_imposto)
                row_count_doc += 1; current_row += 1

            if row_count_doc > 1:
                for col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 15]:
                    ws.merge_cells(start_row=doc_row_start, start_column=col_idx + col_offset, end_row=doc_row_start + row_count_doc - 1, end_column=col_idx + col_offset)
        
        current_row += 1
        total_labels = ['TOTAL ICMS MENSAL', 'TOTAL IPI MENSAL', 'TOTAL S.T. MENSAL']
        total_data = [totais_gerais['icms'], totais_gerais['ipi'], {'VCON':Decimal(0), 'BASE_CALCULO':Decimal(0), 'VALOR_IMPOSTO':Decimal(0), 'ISENTAS_NT':Decimal(0), 'OUTRAS':Decimal(0)}]
        for i, label in enumerate(total_labels):
            dados = total_data[i]
            write_row([label, None, None, None, None, dados['VCON'], None, None, None, dados['BASE_CALCULO'], None, dados['VALOR_IMPOSTO'], dados['ISENTAS_NT'], dados['OUTRAS'], None])
            ws.cell(row=current_row, column=1 + col_offset).font = font_bold_small
            ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=5 + col_offset); ws.merge_cells(start_row=current_row, start_column=7 + col_offset, end_row=current_row, end_column=9 + col_offset); ws.merge_cells(start_row=current_row, start_column=11 + col_offset, end_row=current_row, end_column=11 + col_offset); ws.merge_cells(start_row=current_row, start_column=15 + col_offset, end_row=current_row, end_column=15 + col_offset)
            current_row += 1
        current_row += 1

        write_row(['DEMONSTRATIVO POR ESTADO'])
        ws.cell(row=current_row, column=1 + col_offset).font = font_bold_small
        ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=15 + col_offset)
        current_row += 1
        
        for uf, valores in sorted(totais_uf.items()):
            write_row(['', '', '', '', uf, valores['VCON'], '', '', '', valores['BASE_CALCULO'], valores['ALIQUOTA_MEDIA'], valores['VALOR_IMPOSTO'], valores['ISENTAS_NT'], valores['OUTRAS'], ''])
            ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=4 + col_offset)
            ws.merge_cells(start_row=current_row, start_column=7 + col_offset, end_row=current_row, end_column=9 + col_offset)
            ws.merge_cells(start_row=current_row, start_column=11 + col_offset, end_row=current_row, end_column=11 + col_offset)
            ws.cell(row=current_row, column=5 + col_offset).alignment = align_center_middle
            current_row += 1
        current_row += 1

        write_row(['DEMONSTRATIVO POR CFOP'])
        ws.cell(row=current_row, column=1 + col_offset).font = font_bold_small
        ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=15 + col_offset)
        current_row += 1
        for cfop, valores in sorted(totais_cfop.items()):
            write_row(['', '', '', '', f"{cfop // 1000}.{cfop % 1000}", valores['VCON'], '', '', '', valores['BASE_CALCULO'], valores['ALIQUOTA_MEDIA'], valores['VALOR_IMPOSTO'], valores['ISENTAS_NT'], valores['OUTRAS'], ''])
            ws.merge_cells(start_row=current_row, start_column=1 + col_offset, end_row=current_row, end_column=4 + col_offset); ws.merge_cells(start_row=current_row, start_column=7 + col_offset, end_row=current_row, end_column=9 + col_offset); ws.merge_cells(start_row=current_row, start_column=11 + col_offset, end_row=current_row, end_column=11 + col_offset)
            ws.cell(row=current_row, column=5 + col_offset).alignment = align_center_middle
            current_row += 1
        
        for row in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1 + col_offset, max_col=15 + col_offset):
            for cell in row:
                if cell.row > header_start_row + 3:
                    cell.font = font_normal_small
                    if cell.column in [2+c for c in [1, 2, 3, 4, 8]]: cell.alignment = align_center_middle
                    elif cell.column in [2+c for c in [5, 6, 9, 10, 11, 12, 13]]: cell.alignment = align_right_middle
                    else: cell.alignment = align_left_middle
                    if isinstance(cell.value, Decimal):
                      if cell.column in [2+c for c in [5, 9, 11, 12, 13]]: cell.number_format = num_format_brl
                      if cell.column == 2+10: cell.number_format = num_format_decimal


        wb.save(nome_arquivo)
        print(f"✅ XLSX '{nome_arquivo}' gerado com sucesso.")
    except Exception as e:
        print(f"❌ Erro ao gerar XLSX '{nome_arquivo}': {e}")
        traceback.print_exc()

# --- NOVA FUNÇÃO PRINCIPAL ---
def gerar_livro_saidas(codi_emp, data_inicio, data_fim, gerar_pdf=True, gerar_xlsx=False):
    """
    Gera o relatório Livro de Saídas para uma empresa e período específicos.

    Args:
        codi_emp (int): O código da empresa.
        data_inicio (str): A data de início no formato 'YYYY-MM-DD'.
        data_fim (str): A data de fim no formato 'YYYY-MM-DD'.
        gerar_pdf (bool): Se True, gera o arquivo PDF.
        gerar_xlsx (bool): Se True, gera o arquivo XLSX.

    Returns:
        list: Uma lista contendo os nomes dos arquivos gerados com sucesso.
              Retorna uma lista vazia em caso de falha.
    """
    # Lista para armazenar os nomes dos arquivos criados
    arquivos_gerados = []
    
    try:
        print("\n--- INICIANDO GERAÇÃO DO LIVRO DE SAÍDAS ---")
        print(f"Empresa: {codi_emp}, Período: {data_inicio} a {data_fim}")
        print("Iniciando a busca de dados concorrente...")

        params_principais = (
            codi_emp, data_inicio, data_fim, codi_emp, data_inicio, data_fim,
            codi_emp, data_inicio, data_fim, codi_emp, data_inicio, data_fim,
            codi_emp, data_inicio, data_fim
        )
        params_impostos = (
            codi_emp, data_inicio, data_fim, codi_emp, data_inicio, data_fim,
            codi_emp, data_inicio, data_fim, codi_emp, data_inicio, data_fim,
            codi_emp, data_inicio, data_fim
        )

        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            future_cabecalho = executor.submit(executar_consulta, SQL_DADOS_CABECALHO, (codi_emp,))
            future_documentos = executor.submit(executar_consulta, SQL_DADOS_PRINCIPAIS, params_principais)
            future_impostos = executor.submit(executar_consulta, SQL_DETALHES_IMPOSTOS, params_impostos)
            
            cabecalho_info = future_cabecalho.result()[0]
            lista_documentos = future_documentos.result()
            detalhes_impostos_raw = future_impostos.result()
        
        if lista_documentos is None or detalhes_impostos_raw is None or cabecalho_info is None:
            print("❌ Falha na busca de dados. Abortando.")
            return [] # Retorna lista vazia em caso de falha inicial

        print(f"Busca de dados concluída. Documentos: {len(lista_documentos)}, Detalhes de Impostos: {len(detalhes_impostos_raw)}")
        
        dados_cabecalho = {
            'nome': cabecalho_info['RAZAO_EMP'], 'cnpj': cabecalho_info['CGCE_EMP'], 'insc_est': cabecalho_info['IEST_EMP'], 
            'periodo_inicio': datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y'), 
            'periodo_fim': datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
        }
        
        print("Processando e vinculando dados...")
        impostos_por_nota = defaultdict(dict)
        for imposto in detalhes_impostos_raw:
            chave_unica = (imposto['TIPO_DOCUMENTO'], imposto['DOCUMENTO_ID'])
            impostos_por_nota[chave_unica][imposto['IMPOSTO_ID']] = imposto
        
        for documento in lista_documentos:
            chave_doc = (documento['TIPO_NOTA'], documento['CODIGO'])
            documento['impostos'] = impostos_por_nota[chave_doc]
            documento['VCON'] = Decimal(documento.get('VCON', 0)) if documento.get('VCON') is not None else Decimal(0)

            if documento['impostos']:
                primeiro_imposto = next(iter(documento['impostos'].values()))
                if primeiro_imposto.get('CFOP'):
                    documento['CFOP'] = primeiro_imposto['CFOP']
                for imp_id, imp_data in documento['impostos'].items():
                    for k, v in imp_data.items():
                        if k in ['VALOR_CONTABIL_IMPOSTO', 'BASE_CALCULO', 'ALIQUOTA', 'VALOR_IMPOSTO', 'ISENTAS_NT', 'OUTRAS']:
                            imp_data[k] = Decimal(v or 0)
        
        print("Calculando totais...")
        totais_cfop = defaultdict(lambda: defaultdict(Decimal))
        totais_gerais = {'icms': defaultdict(Decimal), 'ipi': defaultdict(Decimal)}
        totais_uf = defaultdict(lambda: defaultdict(Decimal))
        
        documentos_processados = set()

        for documento in lista_documentos:
            if not documento['impostos']: continue
            chave_doc = (documento['TIPO_NOTA'], documento['CODIGO'])
            if chave_doc in documentos_processados: continue
            
            documentos_processados.add(chave_doc)
            cfop_doc, uf_doc = documento.get('CFOP'), documento.get('UF') or 'S/UF'
            primeiro_imposto = next(iter(documento['impostos'].values()))
            base_doc = primeiro_imposto.get('BASE_CALCULO', Decimal(0))
            isentas_doc = primeiro_imposto.get('ISENTAS_NT', Decimal(0))
            outras_doc = primeiro_imposto.get('OUTRAS', Decimal(0))

            totais_uf[uf_doc]['VCON'] += documento['VCON']; totais_uf[uf_doc]['BASE_CALCULO'] += base_doc
            totais_uf[uf_doc]['ISENTAS_NT'] += isentas_doc; totais_uf[uf_doc]['OUTRAS'] += outras_doc
            
            if cfop_doc:
                totais_cfop[cfop_doc]['VCON'] += documento['VCON']; totais_cfop[cfop_doc]['BASE_CALCULO'] += base_doc
                totais_cfop[cfop_doc]['ISENTAS_NT'] += isentas_doc; totais_cfop[cfop_doc]['OUTRAS'] += outras_doc
            
            for imp_id, imp_data in documento['impostos'].items():
                valor_imposto, vcon_imp = imp_data.get('VALOR_IMPOSTO', Decimal(0)), imp_data.get('VALOR_CONTABIL_IMPOSTO', Decimal(0))
                totais_uf[uf_doc]['VALOR_IMPOSTO'] += valor_imposto
                if cfop_doc: totais_cfop[cfop_doc]['VALOR_IMPOSTO'] += valor_imposto
                
                if imp_id == 1: # ICMS
                    for key in ['BASE_CALCULO', 'VALOR_IMPOSTO', 'ISENTAS_NT', 'OUTRAS']: totais_gerais['icms'][key] += imp_data.get(key, Decimal(0))
                    totais_gerais['icms']['VCON'] += vcon_imp
                elif imp_id == 2: # IPI
                    for key in ['BASE_CALCULO', 'VALOR_IMPOSTO', 'ISENTAS_NT', 'OUTRAS']: totais_gerais['ipi'][key] += imp_data.get(key, Decimal(0))
                    totais_gerais['ipi']['VCON'] += vcon_imp
        
        for totais in (totais_cfop, totais_uf):
            for key in totais:
                totais[key]['ALIQUOTA_MEDIA'] = (totais[key]['VALOR_IMPOSTO'] / totais[key]['BASE_CALCULO'] * 100) if totais[key]['BASE_CALCULO'] > 0 else Decimal(0)

        # A variável 'dados_completos' ainda pode ser útil internamente, se necessário.
        # dados_completos = {
        #     'metadata': { 'empresa_codigo': codi_emp, 'periodo_inicio': data_inicio, 'periodo_fim': data_fim, 'data_geracao': datetime.now().isoformat() },
        #     'cabecalho_empresa': dados_cabecalho, 'documentos': lista_documentos, 'totais_por_cfop': totais_cfop, 'totais_por_uf': totais_uf, 'totais_gerais': totais_gerais
        # }
        
        nome_base = f"Livro_Saidas{codi_emp}"
        nome_pdf = f"{nome_base}.pdf"
        nome_xlsx = f"{nome_base}.xlsx"
        
        documentos_para_relatorio = [doc for doc in lista_documentos if doc['impostos']]
        print(f"Total de documentos com impostos a serem exibidos: {len(documentos_para_relatorio)}")

        if gerar_pdf:
            print(f"Gerando o arquivo PDF: {nome_pdf}...")
            criar_pdf_livro_saidas(nome_pdf, dados_cabecalho, documentos_para_relatorio, totais_cfop, totais_gerais, totais_uf)
            arquivos_gerados.append(nome_pdf)
        
        if gerar_xlsx:
            print(f"Gerando o arquivo XLSX: {nome_xlsx}...")
            criar_xlsx_livro_saidas(nome_xlsx, dados_cabecalho, documentos_para_relatorio, totais_cfop, totais_gerais, totais_uf)
            arquivos_gerados.append(nome_xlsx)
        
        # O retorno agora é a lista de nomes de arquivos
        return arquivos_gerados
        
    except Exception as e:
        print(f"❌ Ocorreu um erro inesperado na função principal: {e}")
        traceback.print_exc()
        # Retorna uma lista vazia em caso de qualquer exceção
        return []