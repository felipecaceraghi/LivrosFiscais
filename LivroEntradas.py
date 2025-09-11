import pyodbc
import json
import os
import locale
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak, Paragraph, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.pdfgen import canvas

# NOVA IMPORTAÇÃO PARA XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIGURAÇÕES GLOBAIS ---
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    print("Locale 'pt_BR.UTF-8' não encontrado. Usando locale padrão.")

# Configuração da conexão com o banco
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

HEADER_INFO = {}  # Dicionário global para informações do cabeçalho

# --- TEMPLATES DE CONSULTAS COMPLETOS ---
QUERIES_SETUP_TEMPLATE = {
    "1": "SELECT P.INICIOEFETIVO_PAR , V.SIMPLESN_OPTANTE_PAR , V.CENTRAL_PAR , V.SIMPLESN_ULTRAPASSOU_PAR , V.SIMPLESN_ICMS_NORMAL_PAR , V.SIMPLESN_ULTRAPASSOU_PAR FROM BETHADBA.EFPARAMETRO_VIGENCIA AS V INNER JOIN BETHADBA.EFPARAMETRO AS P ON P.CODI_EMP =V.CODI_EMP WHERE V.CODI_EMP = {codi_emp} AND V.VIGENCIA_PAR =DSDBA.D_BUSCA_VIGENCIA_PARAMETRO ( {codi_emp} , DATE ( '{data_inicio}' ) )",
    "2": "SELECT IMPOSTO_VIGENCIA.PDIC_IMP FROM BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA WHERE IMPOSTO_VIGENCIA.CODI_EMP = {codi_emp} AND IMPOSTO_VIGENCIA.CODI_IMP = 1 AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO ( IMPOSTO_VIGENCIA.CODI_EMP , IMPOSTO_VIGENCIA.CODI_IMP , '{data_inicio}' )",
    "3": "SELECT P.PINI_PAR FROM BETHADBA.EFPARAMETRO AS P WHERE P.CODI_EMP = {codi_emp}",
    "4": "SELECT COUNT ( *) FROM BETHADBA.EFSDOIMP WHERE CODI_EMP = {codi_emp} AND DATA_SIM >= '{data_inicio}' AND DATA_SIM <= '{data_fim}'",
    "5": "SELECT VALOR FROM BETHADBA.GEINICIAL AS GEINICIAL WHERE CHAVE ='' AND SECAO ='' AND USUARIO ='ESTATISTICA002'",
    "6": "SELECT VALOR FROM BETHADBA.GEINICIAL AS GEINICIAL WHERE CHAVE ='' AND SECAO ='' AND USUARIO ='TODOS'",
    "7": "SELECT TINS_EMP FROM BETHADBA.GEEMPRE AS GEEMPRE WHERE CODI_EMP = {codi_emp}",
    "8": "SELECT IMPOSTO_RECOLHIMENTO.I_RECOLHIMENTO AS I_RECOLHIMENTO, IMPOSTO_RECOLHIMENTO.CODIGO_RECOLHIMENTO AS CODIGO_RECOLHIMENTO, IMPOSTO_RECOLHIMENTO.DESCRICAO AS DESCRICAO_RECOLHIMENTO FROM BETHADBA.GEIMPOSTO_VIGENCIA_RECOLHIMENTO AS IMPOSTO_RECOLHIMENTO WHERE IMPOSTO_RECOLHIMENTO.CODI_EMP = {codi_emp} AND IMPOSTO_RECOLHIMENTO.CODI_IMP = 2 AND IMPOSTO_RECOLHIMENTO.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(IMPOSTO_RECOLHIMENTO.CODI_EMP, IMPOSTO_RECOLHIMENTO.CODI_IMP, DATE('{data_inicio}')) AND 2 = 2 UNION ALL SELECT 0 AS I_RECOLHIMENTO, '' AS CODIGO_RECOLHIMENTO, '' AS DESCRICAO_RECOLHIMENTO FROM DSDBA.DUMMY WHERE 2 = 30 ORDER BY 1, 2",
    "20": "SELECT MAX ( 1 ) FROM BETHADBA.EFTMP_APURACAO_PERIODO WHERE CODI_EMP = {codi_emp} AND COMPETENCIA ='{data_inicio}' AND TIPO =2",
    "21": "INSERT INTO BETHADBA.EFTMP_APURACAO_PERIODO ( CODI_EMP , COMPETENCIA , TIPO ) VALUES ( {codi_emp} , '{data_inicio}' , 2 )",
    "22": "INSERT INTO BETHADBA.EFTMP_ENTRADAS_PERIODO ( CODI_EMP , CODI_ENT , DENT_ENT , DATA_ESCRITURACAO , CODI_NAT , CODI_ESP , CODI_ACU , CODI_FOR , SITUACAO_ENT , VCON_ENT , VEXC_ENT , DATA_ENTRADA , CODIGO_FETHAB_ENT , I_RECOLHIMENTO_FETHAB , DDOC_ENT , MODALIDADE_IMPORTACAO_ENT , SERI_ENT , NUME_ENT , CIFOB_ENT , SUB_SERIE_ENT , COMPTE_ENT , EMITENTE_ENT , CHAVE_NFE_ENT , VPROD_ENT , I_SCP , VFRE_ENT , VSEG_ENT , VDESACE_ENT , VPIS_ENT , VCOF_ENT , TIPO_IMPORTACAO_ENT , DECLARACAO_IMPORTACAO_ENT , NUMERO_AC_DRAWBACK_ENT , COD_SITTRIB_ENT , TIPO_SERVICO , ATEX_ENT , VALOR_DESCONTO_ENT , VALOR_RESSARCIMENTO_1603_2603_SUBTRI , PEDAGIO_ENT , IPI_ENT , ICMS_ST_ENT , VALOR_IPI_DEVOLUCAO , RESPONSAVEL_FETHAB_ENT , MODALIDADE_FRETE_ENT , VALOR_ICMS_DESONERADO_ENT , VIGENCIA_PAR , VIGENCIA_ACU , SEGI_ENT , ACU_IDEV_ACU , CEXC_ENT , NRO_PARECER_FISCAL_ENT , DATA_PARECER_FISCAL_ENT , CODIGO_MUNICIPIO , CODIGO_MUNICIPIO_DESTINO , TIPO_FATURAMENTO_NFCOM , MODELO_REFERENCIADO , HASH_REFERENCIADO , SERIE_REFERENCIADA , NUMERO_REFERENCIADO , PERIODO_REFERENCIADO , TIPO_CTE_ENT , CTE_REFERENCIA_ENT ) SELECT N.CODI_EMP , N.CODI_ENT , N.DENT_ENT , N.DATA_ESCRITURACAO , N.CODI_NAT , N.CODI_ESP , N.CODI_ACU , N.CODI_FOR , N.SITUACAO_ENT , COALESCE ( N.VCON_ENT , 0 ) , COALESCE ( N.VEXC_ENT , 0 ) , N.DATA_ENTRADA , N.CODIGO_FETHAB_ENT , N.I_RECOLHIMENTO_FETHAB , N.DDOC_ENT , N.MODALIDADE_IMPORTACAO_ENT , N.SERI_ENT , N.NUME_ENT , N.CIFOB_ENT , N.SUB_SERIE_ENT , N.COMPTE_ENT , N.EMITENTE_ENT , N.CHAVE_NFE_ENT , COALESCE ( N.VPROD_ENT , 0 ) , N.I_SCP , COALESCE ( N.VFRE_ENT , 0 ) , COALESCE ( N.VSEG_ENT , 0 ) , COALESCE ( N.VDESACE_ENT , 0 ) , COALESCE ( N.VPIS_ENT , 0 ) , COALESCE ( N.VCOF_ENT , 0 ) , N.TIPO_IMPORTACAO_ENT , N.DECLARACAO_IMPORTACAO_ENT , N.NUMERO_AC_DRAWBACK_ENT , N.COD_SITTRIB_ENT , TIPO_SERVICO , ATEX_ENT , COALESCE ( N.VALOR_DESCONTO_ENT , 0 ) , COALESCE ( N.VALOR_RESSARCIMENTO_1603_2603_SUBTRI , 0 ) , N.PEDAGIO_ENT , N.IPI_ENT , N.ICMS_ST_ENT , N.VALOR_IPI_DEVOLUCAO , N.RESPONSAVEL_FETHAB_ENT , N.MODALIDADE_FRETE_ENT , N.VALOR_ICMS_DESONERADO_ENT , DSDBA.D_BUSCA_VIGENCIA_PARAMETRO ( N.CODI_EMP , N.DENT_ENT ) , ACUMULADOR.VIGENCIA_ACU , N.SEGI_ENT , ACUMULADOR.IDEV_ACU , N.CEXC_ENT , N.NRO_PARECER_FISCAL_ENT , N.DATA_PARECER_FISCAL_ENT , N.CODIGO_MUNICIPIO , N.CODIGO_MUNICIPIO_DESTINO , N.TIPO_FATURAMENTO_NFCOM , N.MODELO_REFERENCIADO , N.HASH_REFERENCIADO , N.SERIE_REFERENCIADA , N.NUMERO_REFERENCIADO , N.PERIODO_REFERENCIADO , N.TIPO_CTE_ENT , N.CTE_REFERENCIA_ENT FROM BETHADBA.EFENTRADAS AS N INNER JOIN BETHADBA.EFACUMULADOR_VIGENCIA AS ACUMULADOR ON ACUMULADOR.CODI_EMP =N.CODI_EMP AND ACUMULADOR.CODI_ACU =N.CODI_ACU AND ACUMULADOR.VIGENCIA_ACU =DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR ( N.CODI_EMP , N.CODI_ACU , N.DENT_ENT ) WHERE N.CODI_EMP = {codi_emp} AND N.DENT_ENT BETWEEN DATE ( '{data_inicio}' ) AND DATE ( '{data_fim}' )",
    "23": "CREATE STATISTICS BETHADBA.EFTMP_ENTRADAS_PERIODO;",
    "25": "SELECT E.CGCE_EMP FROM BETHADBA.GEEMPRE AS E WHERE E.CODI_EMP = {codi_emp}",
    "26": "SELECT G.IEST_EMP, G.RAZAO_EMP FROM BETHADBA.GEEMPRE G WHERE CODI_EMP = {codi_emp}",
    "34": "SELECT COALESCE ( DSDBA.S_BUSCA_ALTERACAO_EMPRESA ( {codi_emp} , '{data_inicio}' , 2 ) , 'EMPRESA PADRAO' ) AS NOME_EMP FROM DSDBA.DUMMY",
    "35": "SELECT COALESCE ( DSDBA.S_BUSCA_ALTERACAO_EMPRESA ( {codi_emp} , '{data_inicio}' , 12 ) , '' ) AS IE_EMP FROM DSDBA.DUMMY",
}

# Consulta SQL para buscar dados brutos, sem filtros de negócio ou ordenação.
LISTA_NOTAS_TEMPLATE = """
    SELECT E.DENT_ENT AS DMOV, E.DDOC_ENT AS DDOC, F.NOME_FOR AS NOME_FORNECEDOR,
           P.CODIGO_MODELO AS MODELO, E.SEGI_ENT AS SEGMENTO,
           E.CODI_ENT AS CODIGO, E.CODI_FOR AS FORNECEDOR, 
           E.SERI_ENT AS SERIE, E.SUB_SERIE_ENT AS SUB_SERIE, E.NUME_ENT AS NUMERO, 
           E.CODI_NAT AS CFOP, E.VCON_ENT AS VCON, 
           E.CODI_ESP AS CODESP, A.CDEB_ACU AS CONTA_ACU,
           P.NOME_ESP AS ESPECIE, F.SIGL_EST AS UF, E.EMITENTE_ENT AS EMITENTE
    FROM BETHADBA.EFTMP_ENTRADAS_PERIODO AS E 
    INNER JOIN BETHADBA.EFACUMULADOR_VIGENCIA AS A ON A.CODI_EMP = E.CODI_EMP AND A.CODI_ACU = E.CODI_ACU AND A.VIGENCIA_ACU = E.VIGENCIA_ACU 
    INNER JOIN BETHADBA.EFFORNECE AS F ON F.CODI_EMP = E.CODI_EMP AND F.CODI_FOR = E.CODI_FOR 
    INNER JOIN BETHADBA.EFESPECIES AS P ON E.CODI_ESP = P.CODI_ESP 
    WHERE E.CODI_EMP = {codi_emp} AND E.DENT_ENT BETWEEN '{data_inicio}' AND '{data_fim}'
"""

LOOP_QUERIES_TEMPLATE = {
    'impostos': """
        SELECT I.CODI_EMP AS EMPRESA, I.CODI_ENT AS NOTA, 
               CASE 
                   WHEN I.CODI_IMP = 30 THEN 2 
                   WHEN I.CODI_IMP = 1 AND TDEFACUM_TEM_FUNDAP.TEM_FUNDAP = 'S' AND G.ESTA_EMP = 'ES' AND I.VLOR_IEN > 0 THEN 130 
                   WHEN I.CODI_IMP = 125 THEN 1 
                   ELSE I.CODI_IMP 
               END AS IMPOSTO, 
               (CASE WHEN TDAUX2.GERA_ALIQUOTA = 1 THEN I.ALIQ_IEN ELSE 0.00 END) AS ALIQUOTA, 
               I.SEQU_IEN AS SEQUENCIAL, 
               CASE 
                   WHEN TDAUX.SOMA = 0 OR TDAUX.RJBARES = 1 OR (E.CODI_NAT IN (1111, 1113, 2111, 2113) AND R.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S') OR (R.SP_ESCRITURAR_REMESSA_PARA_VENDA_FORA_ESTAB = 'S' AND ESP.CODIGO_MODELO IN('01', '1B', '04', '55') AND TDEFACUMULADOR_VIGENCIA_IMPOSTOS_ICMS.TEM_ICMS = 'S' AND E.CODI_NAT IN (1904, 2904)) OR (G.ESTA_EMP = 'MG' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515 = 'S' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515_TIPO IN (1, 2, 4)) 
                   THEN 0.00 
                   ELSE I.BCAL_IEN 
               END AS BASE, 
               (CASE WHEN TDAUX2.GERA_VALOR_IMPOSTO = 1 THEN (CASE WHEN G.ESTA_EMP = 'RS' AND I.CODI_IMP IN (1, 9) THEN I.VALOR_FUNDO_AMPARA_IEN ELSE 0 END) + TD_VALORIMP.VALOR_IMP ELSE 0 END) + (CASE WHEN G.ESTA_EMP = 'ES' AND I.CODI_IMP = 1 THEN TDMOVIMENTO.VALOR_ICMS_MONOFASICO ELSE 0 END) AS VALOR, 
               CASE 
                   WHEN TDAUX.SOMA = 0 OR TDAUX.RJBARES = 1 OR (E.CODI_NAT IN (1111, 1113, 2111, 2113) AND R.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S') OR (R.SP_ESCRITURAR_REMESSA_PARA_VENDA_FORA_ESTAB = 'S' AND ESP.CODIGO_MODELO IN('01', '1B', '04', '55') AND TDEFACUMULADOR_VIGENCIA_IMPOSTOS_ICMS.TEM_ICMS = 'S' AND E.CODI_NAT IN (1904, 2904)) OR (G.ESTA_EMP = 'MG' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515 = 'S' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515_TIPO IN (1, 2, 4)) 
                   THEN 0.00 
                   ELSE CASE WHEN I.CODI_IMP = 1 THEN I.VISE_IEN + I.NTRI_IEN + I.REDU_IEN ELSE I.VISE_IEN END 
               END AS ISENTAS, 
               CASE 
                   WHEN TDAUX.SOMA = 0 THEN I.VCON_IEN 
                   ELSE CASE WHEN TDAUX.RJBARES = 1 OR (E.CODI_NAT IN (1111, 1113, 2111, 2113) AND R.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S') OR (R.SP_ESCRITURAR_REMESSA_PARA_VENDA_FORA_ESTAB = 'S' AND ESP.CODIGO_MODELO IN('01', '1B', '04', '55') AND TDEFACUMULADOR_VIGENCIA_IMPOSTOS_ICMS.TEM_ICMS = 'S' AND E.CODI_NAT IN (1904, 2904)) OR (G.ESTA_EMP = 'MG' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515 = 'S' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515_TIPO IN (1, 2, 4)) THEN 0 ELSE I.VOUT_IEN END 
               END AS OUTRAS, 
               I.VLORIPI_IEN AS VIPI, 
               TDAUX.VALOR_SUBTRI AS VSUB, 
               I.VCON_IEN AS VALOR_CONTABIL, 
               I.VALOR_TARE_IEN AS VALOR_TARE_IEN, 
               I.ALIQUOTA_TARE_IEN AS ALIQUOTA_TARE_IEN, 
               CASE WHEN TDEFACUM_TEM_FUNDAP.TEM_FUNDAP = 'S' THEN 'FUNDAP' ELSE P.SIGL_IMP END AS SIGLA 
        FROM BETHADBA.EFIMPENT AS I 
        INNER JOIN BETHADBA.EFIMPOSTO AS P ON I.CODI_EMP = P.CODI_EMP AND I.CODI_IMP = P.CODI_IMP 
        INNER JOIN BETHADBA.EFENTRADAS AS E ON E.CODI_EMP = I.CODI_EMP AND E.CODI_ENT = I.CODI_ENT 
        INNER JOIN BETHADBA.EFACUMULADOR_VIGENCIA AS A ON A.CODI_EMP = E.CODI_EMP AND A.CODI_ACU = E.CODI_ACU 
        INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS R ON R.CODI_EMP = E.CODI_EMP 
        INNER JOIN BETHADBA.GEEMPRE AS G ON G.CODI_EMP = E.CODI_EMP 
        INNER JOIN BETHADBA.EFESPECIES AS ESP ON ESP.CODI_ESP = E.CODI_ESP, 
        LATERAL(SELECT (COALESCE(MAX('S'), 'N')) AS TEM_ICMS FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS EFACUMULADOR_VIGENCIA_IMPOSTOS WHERE EFACUMULADOR_VIGENCIA_IMPOSTOS.CODI_EMP = E.CODI_EMP AND EFACUMULADOR_VIGENCIA_IMPOSTOS.CODI_ACU = E.CODI_ACU AND EFACUMULADOR_VIGENCIA_IMPOSTOS.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(E.CODI_EMP, E.CODI_ACU, E.DENT_ENT) AND EFACUMULADOR_VIGENCIA_IMPOSTOS.CODI_IMP = 1) AS TDEFACUMULADOR_VIGENCIA_IMPOSTOS_ICMS, 
        LATERAL(SELECT (COALESCE(MAX('S'), 'N')) AS TEM_FUNDAP FROM BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS EFACUMULADOR_VIGENCIA_IMPOSTOS WHERE EFACUMULADOR_VIGENCIA_IMPOSTOS.CODI_EMP = E.CODI_EMP AND EFACUMULADOR_VIGENCIA_IMPOSTOS.CODI_ACU = E.CODI_ACU AND EFACUMULADOR_VIGENCIA_IMPOSTOS.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(E.CODI_EMP, E.CODI_ACU, E.DENT_ENT) AND EFACUMULADOR_VIGENCIA_IMPOSTOS.CODI_IMP = 130 AND ((A.IDEV_ACU = 'S') AND ESP.CODIGO_MODELO IN ('01', '1B', '04', '07', '08', '8B', '09', '10', '11', '26', '27', '55', '57', '67') OR ESP.CODIGO_MODELO IN ('07', '08', '8B', '09', '10', '11', '26', '27', '57', '67'))) AS TDEFACUM_TEM_FUNDAP, 
        LATERAL(SELECT COUNT(1) AS CONTA FROM BETHADBA.EFIMPENT AS X WHERE X.CODI_EMP = E.CODI_EMP AND X.CODI_ENT = E.CODI_ENT AND X.CODI_IMP = 45) AS TD45, 
        LATERAL(SELECT COALESCE(MAX(TCAL_IMP), 'L') AS TIPO_IMPOSTO FROM BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA WHERE IMPOSTO_VIGENCIA.CODI_EMP = I.CODI_EMP AND IMPOSTO_VIGENCIA.CODI_IMP = 9 AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(I.CODI_EMP, 9, E.DENT_ENT)) AS TD_SUBTRI, 
        LATERAL(SELECT COALESCE(MAX(TCAL_IMP), 'L') AS TIPO_IMPOSTO FROM BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA WHERE IMPOSTO_VIGENCIA.CODI_EMP = I.CODI_EMP AND IMPOSTO_VIGENCIA.CODI_IMP = 31 AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(I.CODI_EMP, IMPOSTO_VIGENCIA.CODI_IMP, E.DENT_ENT)) AS TD_ICMS_STAT, 
        LATERAL(SELECT (CASE WHEN ((A.SIMPLESN_CREDITO_PRESUMIDO_TIPO_ACU = 'A' AND A.SIMPLESN_CREDITO_PRESUMIDO_ACU = 'S' AND G.ESTA_EMP = 'SC' AND I.CODI_IMP = 1 AND E.DENT_ENT >= '2009-01-01') OR (A.SIMPLESN_CREDITO_PRESUMIDO_ACU = 'S' AND G.ESTA_EMP IN ('GO') AND I.CODI_IMP = 1) OR (G.ESTA_EMP = 'RS' AND I.CODI_IMP = 1 AND (A.IMPORTACAO_ACU = 'S' OR A.GERA_CRED_PAG_IMPORT_ACU = 'S')) OR (G.ESTA_EMP = 'RS' AND I.CODI_IMP = 1 AND TD45.CONTA > 0 AND COALESCE(E.MODALIDADE_IMPORTACAO_ENT, 0) IN (1, 2, 3, 4, 5))) THEN 0 ELSE 1 END) AS SOMA, (CASE WHEN (G.ESTA_EMP = 'RJ' AND I.CODI_IMP = 1 AND R.RJ_BARES_RESTAURANTES_PAR = 'S') THEN 1 ELSE 0 END) AS RJBARES, I.VLORSUBTRI_IEN + CASE WHEN G.ESTA_EMP = 'MT' AND I.CODI_IMP = 129 THEN I.VLOR_IEN ELSE 0 END AS VALOR_SUBTRI FROM DSDBA.DUMMY) AS TDAUX, 
        LATERAL(SELECT SUM(1) AS FAVORECIDA FROM BETHADBA.GEMUNICIPIO AS MUNICIPIO WHERE MUNICIPIO.CODIGO_MUNICIPIO = E.CODIGO_MUNICIPIO AND MUNICIPIO.CODIGO_UF = 25) AS TD_UF_FAVORECIDA, 
        LATERAL(SELECT COUNT(1) AS QTD FROM BETHADBA.EFIMPENT AS IMPOSTO WHERE IMPOSTO.CODI_EMP = I.CODI_EMP AND IMPOSTO.CODI_ENT = I.CODI_ENT AND IMPOSTO.CODI_IMP = 9) AS TD_TEM_SUBTRI, 
        LATERAL(SELECT (CASE WHEN G.ESTA_EMP = 'SC' AND I.CODI_IMP = 1 AND A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO = 'S' AND A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO_PERCENTUAL_TIPO = 1 AND ESP.CODIGO_MODELO IN ('06', '66') AND R.REST_PAR = 1 AND E.DENT_ENT < DATE('2020-01-01') THEN ((I.VLOR_IEN + (CASE WHEN TD_TEM_SUBTRI.QTD = 0 THEN I.VLORSUBTRI_IEN ELSE 0 END)) * 80/100) WHEN G.ESTA_EMP = 'SC' AND I.CODI_IMP = 1 AND A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO = 'S' AND ESP.CODIGO_MODELO IN ('06', '66') AND R.REST_PAR = 1 AND E.DENT_ENT >= DATE('2020-01-01') THEN ((I.VLOR_IEN + (CASE WHEN TD_TEM_SUBTRI.QTD = 0 THEN I.VLORSUBTRI_IEN ELSE 0 END)) * A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO_PERCENTUAL/100) WHEN G.ESTA_EMP = 'SC' AND I.CODI_IMP = 9 AND A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO = 'S' AND A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO_PERCENTUAL_TIPO = 1 AND ESP.CODIGO_MODELO IN ('06', '66') AND R.REST_PAR = 1 AND E.DENT_ENT < DATE('2020-01-01') THEN (I.VLOR_IEN * 80/100) WHEN G.ESTA_EMP = 'SC' AND I.CODI_IMP = 9 AND A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO = 'S' AND ESP.CODIGO_MODELO IN ('06', '66') AND R.REST_PAR = 1 AND E.DENT_ENT >= DATE('2020-01-01') THEN (I.VLOR_IEN * A.SC_CREDITO_ENERGIA_ELETRICA_LAUDO_TECNICO_PERCENTUAL/100) ELSE I.VLOR_IEN END) AS VALOR_IMP FROM DSDBA.DUMMY) AS TD_VALORIMP, 
        LATERAL(SELECT COALESCE(SUM(MOVIMENTO.VALOR_ICMS_MONOFASICO), 0) AS VALOR_ICMS_MONOFASICO FROM BETHADBA.EFMVEPRO_ICMS_MONOFASICO AS MOVIMENTO WHERE MOVIMENTO.CODI_EMP = E.CODI_EMP AND MOVIMENTO.CODI_ENT = E.CODI_ENT) TDMOVIMENTO, 
        LATERAL(SELECT (CASE WHEN (TDAUX.SOMA = 0) OR (TDAUX.RJBARES = 1) OR (E.CODI_NAT IN (1111, 1113, 2111, 2113) AND R.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S') OR (R.SP_ESCRITURAR_REMESSA_PARA_VENDA_FORA_ESTAB = 'S' AND ESP.CODIGO_MODELO IN('01', '1B', '04', '55') AND TDEFACUMULADOR_VIGENCIA_IMPOSTOS_ICMS.TEM_ICMS = 'S' AND E.CODI_NAT IN (1904, 2904)) OR (G.ESTA_EMP = 'MG' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515 = 'S' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515_TIPO IN (1, 2, 4)) OR (G.ESTA_EMP = 'RJ' AND E.DENT_ENT >= DATE('2020-04-01') AND A.SIMPLESN_CREDITO_PRESUMIDO_ACU = 'S') THEN 0 ELSE 1 END) AS GERA_ALIQUOTA, (CASE WHEN (TDAUX.SOMA = 0) OR (TDAUX.RJBARES = 1) OR (E.CODI_NAT IN (1111, 1113, 2111, 2113) AND R.DESCONTAR_VALOR_ICMS_IPI_CFOP_MERCANTIL = 'S') OR (R.SP_ESCRITURAR_REMESSA_PARA_VENDA_FORA_ESTAB = 'S' AND ESP.CODIGO_MODELO IN('01', '1B', '04', '55') AND TDEFACUMULADOR_VIGENCIA_IMPOSTOS_ICMS.TEM_ICMS = 'S' AND E.CODI_NAT IN (1904, 2904)) OR (G.ESTA_EMP = 'MG' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515 = 'S' AND A.MG_APROVEITAMENTO_CREDITO_ICMS_485515_TIPO IN (1, 2, 4)) OR (G.ESTA_EMP = 'RJ' AND E.DENT_ENT >= DATE('2020-04-01') AND A.SIMPLESN_CREDITO_PRESUMIDO_ACU = 'S') THEN 0 ELSE 1 END) AS GERA_VALOR_IMPOSTO FROM DSDBA.DUMMY) AS TDAUX2 
        WHERE A.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(E.CODI_EMP, E.CODI_ACU, E.DENT_ENT) 
          AND R.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(E.CODI_EMP, E.DENT_ENT) 
          AND I.CODI_EMP = {codi_emp} AND I.CODI_ENT = {codi_ent} 
          AND ((I.CODI_IMP IN (2, 30)) OR (I.CODI_IMP = 1) OR (I.CODI_IMP = 8) OR (I.CODI_IMP = 9 AND 1 = 1 AND TD_SUBTRI.TIPO_IMPOSTO <> 'P') OR (I.CODI_IMP = 11 AND 0 = 1) OR (I.CODI_IMP = 31 AND 0 = 1 AND TD_ICMS_STAT.TIPO_IMPOSTO <> 'P') OR (I.CODI_IMP = 57 AND G.ESTA_EMP = 'RJ') OR (I.CODI_IMP = 69 AND TD_UF_FAVORECIDA.FAVORECIDA = 1 AND 1 = 1) OR (I.CODI_IMP = 125) OR (I.CODI_IMP = 129 AND G.ESTA_EMP = 'MT' AND A.IDEV_ACU = 'S')) 
          AND NOT(G.ESTA_EMP = 'RN' AND E.CODI_NAT IN (1602, 1605)) 
        UNION 
        SELECT E.CODI_EMP AS EMPRESA, E.CODI_ENT AS NOTA, 1 AS IMPOSTO, 0 AS ALIQUOTA, 1 AS SEQUENCIAL, 0 AS BASE, 0 AS VALOR, 0 AS ISENTAS, E.VCON_ENT AS OUTRAS, 0 AS VIPI, 0 AS VSUB, 0 AS VALOR_CONTABIL, 0 AS VALOR_TARE_IEN, 0 AS ALIQUOTA_TARE_IEN, COALESCE((SELECT I.SIGL_IMP FROM BETHADBA.EFIMPOSTO AS I WHERE I.CODI_EMP = E.CODI_EMP AND I.CODI_IMP = 1), '') AS SIGLA 
        FROM BETHADBA.EFENTRADAS AS E 
        INNER JOIN BETHADBA.EFESPECIES AS EFESPECIES ON EFESPECIES.CODI_ESP = E.CODI_ESP 
        INNER JOIN BETHADBA.EFPARAMETRO_VIGENCIA AS P ON P.CODI_EMP = E.CODI_EMP 
        WHERE P.VIGENCIA_PAR = DSDBA.D_BUSCA_VIGENCIA_PARAMETRO(E.CODI_EMP, E.DENT_ENT) 
          AND E.CODI_EMP = {codi_emp} AND E.CODI_ENT = {codi_ent} 
          AND EFESPECIES.DOCUMENTO_NAO_FISCAL = 'N' AND 'N' = 'S' AND P.SIMPLESN_ULTRAPASSOU_PAR = 'N' 
          AND NOT EXISTS (SELECT 1 FROM BETHADBA.EFIMPENT AS I WHERE I.CODI_EMP = E.CODI_EMP AND I.CODI_ENT = E.CODI_ENT AND I.CODI_IMP = 1) 
        UNION ALL 
        SELECT ENTRADA.CODI_EMP AS EMPRESA, ENTRADA.CODI_ENT AS NOTA, 9 AS IMPOSTO, 0 AS ALIQUOTA, 1 AS SEQUENCIAL, SUM(TDMOVIMENTO.BICMSST_MEP) AS BASE, SUM(TDAUX_02.VALOR_SUBTRI) AS VALOR, 0 AS ISENTAS, 0 AS OUTRAS, 0 AS VIPI, SUM(TDMOVIMENTO.VALOR_SUBTRI_MEP) + SUM(TD_AMPARA.VALOR_ICMS_ST) AS VSUB, 0 AS VALOR_CONTABIL, 0 AS VALOR_TARE_IEN, 0 AS ALIQUOTA_TARE_IEN, TD_SUBTRI.SIGLA AS SIGLA 
        FROM BETHADBA.EFENTRADAS AS ENTRADA 
        INNER JOIN BETHADBA.EFACUMULADOR_VIGENCIA_IMPOSTOS AS ACUMULADOR_IMPOSTOS ON ACUMULADOR_IMPOSTOS.CODI_EMP = ENTRADA.CODI_EMP AND ACUMULADOR_IMPOSTOS.CODI_ACU = ENTRADA.CODI_ACU AND ACUMULADOR_IMPOSTOS.CODI_IMP = 9 AND ACUMULADOR_IMPOSTOS.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(ACUMULADOR_IMPOSTOS.CODI_EMP, ACUMULADOR_IMPOSTOS.CODI_ACU, ENTRADA.DENT_ENT) 
        INNER JOIN BETHADBA.GEEMPRE AS GEEMPRE ON GEEMPRE.CODI_EMP = ENTRADA.CODI_EMP, 
        LATERAL(SELECT COALESCE(MAX(IMPOSTO_VIGENCIA.TCAL_IMP), 'L') AS TIPO_IMPOSTO, COALESCE(MAX(IMPOSTO.SIGL_IMP), '') AS SIGLA FROM BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA INNER JOIN BETHADBA.GEIMPOSTO AS IMPOSTO ON IMPOSTO.CODI_EMP = IMPOSTO_VIGENCIA.CODI_EMP AND IMPOSTO.CODI_IMP = IMPOSTO_VIGENCIA.CODI_IMP WHERE IMPOSTO_VIGENCIA.CODI_EMP = ENTRADA.CODI_EMP AND IMPOSTO_VIGENCIA.CODI_IMP = ACUMULADOR_IMPOSTOS.CODI_IMP AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(ENTRADA.CODI_EMP, IMPOSTO_VIGENCIA.CODI_IMP, ENTRADA.DENT_ENT)) AS TD_SUBTRI, 
        LATERAL(SELECT COUNT(1) AS QTDE_REGISTRO, COALESCE(SUM(MOVIMENTO.BICMSST_MEP), 0) AS BICMSST_MEP, COALESCE(SUM(MOVIMENTO.VALOR_SUBTRI_MEP), 0) AS VALOR_SUBTRI_MEP FROM BETHADBA.EFMVEPRO AS MOVIMENTO WHERE MOVIMENTO.CODI_EMP = ENTRADA.CODI_EMP AND MOVIMENTO.CODI_ENT = ENTRADA.CODI_ENT) TDMOVIMENTO, 
        LATERAL(SELECT CASE WHEN ENTRADA.CODI_NAT IN (1603, 2603) AND ENTRADA.SITUACAO_ENT = 9 THEN 'S' ELSE 'N' END AS RESSARCIMENTO_ICMS_ST FROM DSDBA.DUMMY) AS TDAUX, 
        LATERAL(SELECT CASE WHEN TDAUX.RESSARCIMENTO_ICMS_ST = 'S' THEN COALESCE(ENTRADA.VALOR_RESSARCIMENTO_1603_2603_SUBTRI, 0) ELSE TDMOVIMENTO.VALOR_SUBTRI_MEP END AS VALOR_SUBTRI FROM DSDBA.DUMMY) AS TDAUX_02, 
        LATERAL(SELECT COALESCE(SUM(MOVIMENTO.VALOR_ICMS_ST), 0) AS VALOR_ICMS_ST FROM BETHADBA.EFMVEPRO_FUNDO_AMPARA AS MOVIMENTO WHERE MOVIMENTO.CODI_EMP = ENTRADA.CODI_EMP AND MOVIMENTO.CODI_ENT = ENTRADA.CODI_ENT AND GEEMPRE.ESTA_EMP = 'RS') AS TD_AMPARA 
        WHERE ENTRADA.CODI_EMP = {codi_emp} AND ENTRADA.CODI_ENT = {codi_ent} AND TD_SUBTRI.TIPO_IMPOSTO = 'P' AND (TDMOVIMENTO.QTDE_REGISTRO > 0 OR TDAUX.RESSARCIMENTO_ICMS_ST = 'S') AND 1 = 1 
        GROUP BY ENTRADA.CODI_EMP, ENTRADA.CODI_ENT, TD_SUBTRI.SIGLA 
        UNION ALL 
        SELECT ENTRADA.CODI_EMP AS EMPRESA, ENTRADA.CODI_ENT AS NOTA, 31 AS IMPOSTO, 0 AS ALIQUOTA, 1 AS SEQUENCIAL, SUM(ICMS_STAT.BASE_CALCULO) AS BASE, SUM(ICMS_STAT.VALOR_IMPOSTO) AS VALOR, 0 AS ISENTAS, 0 AS OUTRAS, 0 AS VIPI, 0 AS VSUB, 0 AS VALOR_CONTABIL, 0 AS VALOR_TARE_IEN, 0 AS ALIQUOTA_TARE_IEN, IMPOSTO.SIGL_IMP AS SIGLA 
        FROM BETHADBA.EFENTRADAS AS ENTRADA 
        INNER JOIN BETHADBA.EFMVEPRO AS MOVIMENTO ON MOVIMENTO.CODI_EMP = ENTRADA.CODI_EMP AND MOVIMENTO.CODI_ENT = ENTRADA.CODI_ENT 
        INNER JOIN BETHADBA.EFMVEPRO_ICMS_STAT AS ICMS_STAT ON ICMS_STAT.CODI_EMP = MOVIMENTO.CODI_EMP AND ICMS_STAT.CODI_ENT = MOVIMENTO.CODI_ENT AND ICMS_STAT.NUME_MEP = MOVIMENTO.NUME_MEP 
        INNER JOIN BETHADBA.GEIMPOSTO AS IMPOSTO ON IMPOSTO.CODI_EMP = ENTRADA.CODI_EMP AND IMPOSTO.CODI_IMP = 31 
        INNER JOIN BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA ON IMPOSTO_VIGENCIA.CODI_EMP = ENTRADA.CODI_EMP AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(ENTRADA.CODI_EMP, IMPOSTO_VIGENCIA.CODI_IMP, ENTRADA.DENT_ENT) 
        INNER JOIN BETHADBA.GEEMPRE AS EMPRESA ON EMPRESA.CODI_EMP = ENTRADA.CODI_EMP 
        INNER JOIN BETHADBA.EFESPECIES AS ESPECIES ON ENTRADA.CODI_ESP = ESPECIES.CODI_ESP 
        INNER JOIN BETHADBA.EFACUMULADOR_VIGENCIA AS ACUMULADOR ON ACUMULADOR.CODI_EMP = ENTRADA.CODI_EMP AND ACUMULADOR.CODI_ACU = ENTRADA.CODI_ACU AND ACUMULADOR.VIGENCIA_ACU = DSDBA.D_BUSCA_VIGENCIA_ACUMULADOR(ENTRADA.CODI_EMP, ENTRADA.CODI_ACU, ENTRADA.DENT_ENT) 
        , LATERAL(SELECT MAX(1) AS TEM_IMPOSTO, SUM(I.BCAL_IEN) AS BASE_CALCULO, SUM(I.VLOR_IEN) AS VALOR_IMPOSTO, SUM(I.VISE_IEN) AS VALOR_ISENTAS, SUM(I.VOUT_IEN) AS VALOR_OUTRAS, SUM(I.VCON_IEN) AS VALOR_CONTABIL FROM BETHADBA.EFIMPENT AS I WHERE I.CODI_EMP = ENTRADA.CODI_EMP AND I.CODI_ENT = ENTRADA.CODI_ENT AND I.CODI_IMP = 31) AS TDNOTA_IMPOSTO 
        WHERE ENTRADA.CODI_EMP = {codi_emp} AND ENTRada.CODI_ENT = {codi_ent} AND EMPRESA.ESTA_EMP = 'BA' 
          AND COALESCE(TDNOTA_IMPOSTO.TEM_IMPOSTO, 0) > 0 AND ACUMULADOR.IDEV_ACU = 'N' AND IMPOSTO_VIGENCIA.CODI_IMP = 31 AND IMPOSTO_VIGENCIA.TCAL_IMP = 'P' 
          AND ESPECIES.CODIGO_MODELO IN('07', '08', '8B', '09', '10', '11', '26', '27', '57', '67') AND 0 = 1 
        GROUP BY ENTRADA.CODI_EMP, ENTRADA.CODI_ENT, IMPOSTO.SIGL_IMP 
        UNION ALL 
        SELECT ENTRADA.CODI_EMP AS EMPRESA, ENTRADA.CODI_ENT AS NOTA, 8 AS IMPOSTO, 0 AS ALIQUOTA, 1 AS SEQUENCIAL, SUM(MOVIMENTO_RJ.BASE_CALCULO_DIFALI) AS BASE, SUM(MOVIMENTO_RJ.VALOR_DIFALI) AS VALOR, 0 AS ISENTAS, 0 AS OUTRAS, 0 AS VIPI, 0 AS VSUB, 0 AS VALOR_CONTABIL, 0 AS VALOR_TARE_IEN, 0 AS ALIQUOTA_TARE_IEN, IMPOSTO.SIGL_IMP AS SIGLA 
        FROM BETHADBA.EFENTRADAS AS ENTRADA 
        INNER JOIN BETHADBA.EFMVEPRO AS MOVIMENTO ON MOVIMENTO.CODI_EMP = ENTRADA.CODI_EMP AND MOVIMENTO.CODI_ENT = ENTRADA.CODI_ENT 
        INNER JOIN BETHADBA.EFMVEPRO_DIFALI AS MOVIMENTO_RJ ON MOVIMENTO_RJ.CODI_EMP = MOVIMENTO.CODI_EMP AND MOVIMENTO_RJ.CODI_ENT = MOVIMENTO.CODI_ENT AND MOVIMENTO_RJ.NUME_MEP = MOVIMENTO.NUME_MEP 
        INNER JOIN BETHADBA.GEIMPOSTO AS IMPOSTO ON IMPOSTO.CODI_EMP = ENTRADA.CODI_EMP AND IMPOSTO.CODI_IMP = 8 
        INNER JOIN BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA ON IMPOSTO_VIGENCIA.CODI_EMP = IMPOSTO.CODI_EMP AND IMPOSTO_VIGENCIA.CODI_IMP = IMPOSTO.CODI_IMP AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(IMPOSTO_VIGENCIA.CODI_EMP, IMPOSTO_VIGENCIA.CODI_IMP, ENTRADA.DENT_ENT) 
        INNER JOIN BETHADBA.GEEMPRE AS EMPRESA ON EMPRESA.CODI_EMP = ENTRADA.CODI_EMP 
        WHERE ENTRADA.CODI_EMP = {codi_emp} AND ENTRADA.CODI_ENT = {codi_ent} AND EMPRESA.ESTA_EMP = 'RJ' 
          AND IMPOSTO_VIGENCIA.TCAL_IMP = 'P' AND 0 = 1 
        GROUP BY ENTRADA.CODI_EMP, ENTRADA.CODI_ENT, IMPOSTO.SIGL_IMP 
        UNION ALL 
        SELECT ENTRADA.CODI_EMP AS EMPRESA, ENTRADA.CODI_ENT AS NOTA, 8 AS IMPOSTO, 0 AS ALIQUOTA, 1 AS SEQUENCIAL, SUM(MOVIMENTO_MT.BASE_CALCULO_DIFALI) AS BASE, SUM(MOVIMENTO_MT.VALOR_DIFALI) AS VALOR, 0 AS ISENTAS, 0 AS OUTRAS, 0 AS VIPI, 0 AS VSUB, 0 AS VALOR_CONTABIL, 0 AS VALOR_TARE_IEN, 0 AS ALIQUOTA_TARE_IEN, IMPOSTO.SIGL_IMP AS SIGLA 
        FROM BETHADBA.EFENTRADAS AS ENTRADA 
        INNER JOIN BETHADBA.EFMVEPRO AS MOVIMENTO ON MOVIMENTO.CODI_EMP = ENTRADA.CODI_EMP AND MOVIMENTO.CODI_ENT = ENTRADA.CODI_ENT 
        INNER JOIN BETHADBA.EFMVEPRO_DIFALI AS MOVIMENTO_MT ON MOVIMENTO_MT.CODI_EMP = MOVIMENTO.CODI_EMP AND MOVIMENTO_MT.CODI_ENT = MOVIMENTO.CODI_ENT AND MOVIMENTO_MT.NUME_MEP = MOVIMENTO.NUME_MEP 
        INNER JOIN BETHADBA.GEIMPOSTO AS IMPOSTO ON IMPOSTO.CODI_EMP = ENTRADA.CODI_EMP AND IMPOSTO.CODI_IMP = 8 
        INNER JOIN BETHADBA.GEIMPOSTO_VIGENCIA AS IMPOSTO_VIGENCIA ON IMPOSTO_VIGENCIA.CODI_EMP = IMPOSTO.CODI_EMP AND IMPOSTO_VIGENCIA.CODI_IMP = IMPOSTO.CODI_IMP AND IMPOSTO_VIGENCIA.VIGENCIA_IMP = DSDBA.D_BUSCA_VIGENCIA_IMPOSTO(IMPOSTO_VIGENCIA.CODI_EMP, IMPOSTO_VIGENCIA.CODI_IMP, ENTRADA.DENT_ENT) 
        INNER JOIN BETHADBA.GEEMPRE AS EMPRESA ON EMPRESA.CODI_EMP = ENTRADA.CODI_EMP 
        INNER JOIN BETHADBA.EFESPECIES AS ESP ON ESP.CODI_ESP = ENTRADA.CODI_ESP 
        WHERE ENTRADA.CODI_EMP = {codi_emp} AND ENTRADA.CODI_ENT = {codi_ent} AND EMPRESA.ESTA_EMP = 'MT' 
          AND IMPOSTO_VIGENCIA.TCAL_IMP = 'P' AND ESP.CODIGO_MODELO IN ('01', '1B', '04', '55', '65') AND 0 = 1 
        GROUP BY ENTRADA.CODI_EMP, ENTRADA.CODI_ENT, IMPOSTO.SIGL_IMP 
        ORDER BY 1, 2, 3, 4, 5
    """
}
def execute_query(cursor, sql, is_setup=False):
    try:
        cursor.execute(sql)
        if cursor.description:
            columns = [column[0] for column in cursor.description]
            rows = cursor.fetchall()
            return [dict(zip(columns, row)) for row in rows]
        if not is_setup:
            cursor.connection.commit()
        return {"status": "Comando executado com sucesso", "rows_affected": cursor.rowcount}
    except pyodbc.Error as ex:
        if ex.args[0] in ('42W08', '52011', '52016'):
            return {"status": "Comando ignorado (provavelmente já existe)", "error_code": ex.args[0]}
        else:
            raise

def get_end_of_month(date_str):
    """Calcula o último dia do mês a partir de uma data de início"""
    start_date = datetime.strptime(date_str, '%Y-%m-%d')
    next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
    end_of_month = next_month - timedelta(days=1)
    return end_of_month.strftime('%Y-%m-%d')

def validate_date_range(data_inicio, data_fim):
    """Valida se o período informado é válido"""
    try:
        inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
        fim = datetime.strptime(data_fim, '%Y-%m-%d')
        
        if inicio > fim:
            raise ValueError(f"Data início ({data_inicio}) não pode ser maior que data fim ({data_fim})")
            
        # Aviso se período for muito longo (mais de 1 ano)
        if (fim - inicio).days > 365:
            print(f"⚠️  AVISO: Período muito longo ({(fim - inicio).days} dias). Isso pode impactar a performance.")
            
        return True
    except ValueError as e:
        print(f"❌ ERRO na validação de datas: {e}")
        return False

def extract_data_from_db(codi_emp, data_inicio, data_fim=None):
    # Se data_fim não foi fornecida, calcula automaticamente
    if data_fim is None:
        data_fim = get_end_of_month(data_inicio)
        print(f"  ✅ Data fim calculada automaticamente: {data_fim}")
    else:
        print(f"  ✅ Data fim fornecida pelo usuário: {data_fim}")
    
    # Validar período
    if not validate_date_range(data_inicio, data_fim):
        return None, None
    
    data = {"header": {}, "transactions": [], "tax_details": {}}
    raw_data = {"lista_notas_raw": [], "impostos_raw": {}}
    
    try:
        with pyodbc.connect(CONN_STR, autocommit=False) as conn:
            with conn.cursor() as cursor:
                print("--- Executando queries de preparação (SETUP) ---")
                for key, template_sql in QUERIES_SETUP_TEMPLATE.items():
                    if not template_sql:  # Pula queries vazias
                        continue
                    sql = template_sql.format(codi_emp=codi_emp, data_inicio=data_inicio, data_fim=data_fim)
                    try:
                        print(f"  - Executando Setup Query #{key}...")
                        result = execute_query(cursor, sql, is_setup=True)
                        if key == "25" and result: data["header"]["cnpj"] = result[0].get("CGCE_EMP")
                        elif key == "26" and result: 
                            # Usar dados diretos da GEEMPRE (mais confiáveis)
                            razao_emp = (result[0].get("RAZAO_EMP") or "EMPRESA PADRAO").strip()
                            iest_emp = (result[0].get("IEST_EMP") or "").strip()
                            data["header"]["empresa_nome"] = razao_emp
                            data["header"]["ie"] = iest_emp
                            print(f"    ✅ Dados da empresa capturados: '{razao_emp}' | IE: '{iest_emp}'")
                        elif key == "34" and result and not data["header"].get("empresa_nome"): 
                            # Fallback se a query 26 não trouxer resultado
                            data["header"]["empresa_nome"] = result[0].get("NOME_EMP")
                        elif key == "35" and result and not data["header"].get("ie"): 
                            # Fallback se a query 26 não trouxer resultado
                            data["header"]["ie"] = result[0].get("IE_EMP")
                    except pyodbc.Error as ex:
                        print(f"    ERRO na query de setup #{key}: {ex}")
                        if key in ["21", "22", "23"]: continue
                        conn.rollback()
                        return None, None
                conn.commit()
                print("Setup concluído com sucesso.")

                print("\n--- Obtendo a lista de notas fiscais (dados brutos) ---")
                if LISTA_NOTAS_TEMPLATE.strip():
                    sql_notas = LISTA_NOTAS_TEMPLATE.format(codi_emp=codi_emp, data_inicio=data_inicio, data_fim=data_fim)
                    lista_de_notas = execute_query(cursor, sql_notas)
                    raw_data["lista_notas_raw"] = lista_de_notas
                    
                    if not lista_de_notas:
                        print("Nenhuma nota encontrada.")
                        return data, raw_data
                    
                    print(f"{len(lista_de_notas)} registros brutos encontrados para processar.")
                    
                    for nota in lista_de_notas:
                        data["transactions"].append({
                            "codigo": nota['CODIGO'],
                            "data_entrada": nota['DMOV'],
                            "data_doc": nota['DDOC'],
                            "nome_fornecedor": nota['NOME_FORNECEDOR'],
                            "modelo": nota['MODELO'], 
                            "segmento": nota['SEGMENTO'], 
                            "especie": nota['ESPECIE'].replace("Eletrônica", "Elet.").replace("Energia Elétrica", "Co").replace("- CTRC Elet.", " CTE")[:5],
                            "serie": str(nota.get('SERIE', '')),
                            "sub_serie": str(nota.get('SUB_SERIE', '')),
                            "numero": str(nota['NUMERO']),
                            "cod_emitente": str(nota['FORNECEDOR']),
                            "uf": nota.get('UF', ''),
                            "cfop": nota['CFOP'],
                            "valor_contabil": Decimal(str(nota.get('VCON', '0.00'))),
                            "emitente": nota.get('EMITENTE', '')
                        })

                    if data["transactions"]:
                        dates = [t['data_entrada'] for t in data['transactions'] if t.get('data_entrada')]
                        if dates:
                            valid_dates = [d for d in dates if d]
                            if valid_dates:
                               data["header"]["periodo_inicio"] = min(valid_dates)
                               data["header"]["periodo_fim"] = max(valid_dates)

                print("\n--- Buscando detalhes de impostos por nota ---")
                codigos_processados = set()
                for transaction in data["transactions"]:
                    cod_entrada = transaction['codigo']
                    if cod_entrada in codigos_processados:
                        continue
                    if LOOP_QUERIES_TEMPLATE['impostos'].strip():
                        sql_impostos = LOOP_QUERIES_TEMPLATE['impostos'].format(codi_emp=codi_emp, codi_ent=cod_entrada)
                        impostos_resultado = execute_query(cursor, sql_impostos)
                        raw_data["impostos_raw"][str(cod_entrada)] = impostos_resultado
                        
                        if impostos_resultado:
                            if cod_entrada not in data['tax_details']:
                                data['tax_details'][cod_entrada] = defaultdict(list)

                            for imposto in impostos_resultado:
                                sigla = imposto.get('SIGLA', '').strip()
                                if not sigla: continue
                                data['tax_details'][cod_entrada][sigla].append({
                                    'BASE': imposto.get('BASE'), 'VALOR': imposto.get('VALOR'),
                                    'ISENTAS': imposto.get('ISENTAS'), 'OUTRAS': imposto.get('OUTRAS'),
                                    'ALIQUOTA': imposto.get('ALIQUOTA'), 'SEQUENCIAL': imposto.get('SEQUENCIAL'),
                                    'IMPOSTO': imposto.get('IMPOSTO'), 'VALOR_CONTABIL': imposto.get('VALOR_CONTABIL'),
                                    'VSUB': imposto.get('VSUB')
                                })
                    codigos_processados.add(cod_entrada)
    except pyodbc.Error as ex:
        print(f"ERRO DE CONEXÃO: {ex}")
        return None, None
    
    global HEADER_INFO
    HEADER_INFO = data["header"]
    print(f"\n=== DADOS EXTRAÍDOS ===\nTransações: {len(data['transactions'])}\nDetalhes de impostos: {len(data['tax_details'])} notas")
    return data, raw_data

def calculate_summaries(transactions, tax_details):
    """Calcula todos os resumos a partir da lista de transações já filtrada e agrupada."""
    print("  - Iniciando cálculo de resumos em Python...")
    
    summaries = {
        'cfop_summary': defaultdict(lambda: defaultdict(Decimal)),
        'total_contabil': Decimal(0),
        'total_icms': Decimal(0),  # ← CORRIGIDO: Total específico de ICMS
        'total_ipi': Decimal(0),
        'total_subtri': Decimal(0),  # ← NOVO: Total de Substituição Tributária
        'icms_valores': defaultdict(Decimal),
        'ipi_valores': defaultdict(Decimal),
        'icms_difal': defaultdict(Decimal),
        'subtri_valores': defaultdict(Decimal)  # ← NOVO: Valores de ST
    }

    # DEBUG: Verificar siglas disponíveis nos dados
    siglas_encontradas = set()
    for doc in transactions:
        for item in doc['items']:
            item_tax_details = tax_details.get(item['codigo'], {})
            siglas_encontradas.update(item_tax_details.keys())
    
    print(f"  - Siglas de impostos encontradas: {sorted(siglas_encontradas)}")

    for doc in transactions:
        summaries['total_contabil'] += doc.get('valor_contabil_total', Decimal(0))
        
        for item in doc['items']:
            cfop = item.get('cfop')
            item_code = item['codigo']
            item_tax_details = tax_details.get(item_code, {})

            # Procurar IPI com diferentes possibilidades de sigla
            siglas_ipi = ['IPI', 'IPI     ', 'IPI    ']
            for sigla in siglas_ipi:
                if sigla in item_tax_details:
                    for ipi_info in item_tax_details[sigla]:
                        valor_ipi = Decimal(str(ipi_info.get('VALOR', 0)))
                        summaries['total_ipi'] += valor_ipi
                        summaries['ipi_valores']['BCAL'] += Decimal(str(ipi_info.get('BASE', 0)))
                        summaries['ipi_valores']['VLOR'] += valor_ipi
                        summaries['ipi_valores']['VISE'] += Decimal(str(ipi_info.get('ISENTAS', 0)))
                        summaries['ipi_valores']['VOUT'] += Decimal(str(ipi_info.get('OUTRAS', 0)))
                    break  # Para no primeiro encontrado

            # Procurar ICMS com diferentes possibilidades de sigla
            siglas_icms = ['ICMS', 'ICMS    ', 'ICMS   ']
            for sigla in siglas_icms:
                if sigla in item_tax_details:
                    for icms_info in item_tax_details[sigla]:
                        valor_icms = Decimal(str(icms_info.get('VALOR', 0)))
                        summaries['total_icms'] += valor_icms  # ← CORRIGIDO
                        summaries['icms_valores']['BCAL'] += Decimal(str(icms_info.get('BASE', 0)))
                        summaries['icms_valores']['VLOR'] += valor_icms
                        summaries['icms_valores']['VISE'] += Decimal(str(icms_info.get('ISENTAS', 0)))
                        summaries['icms_valores']['VOUT'] += Decimal(str(icms_info.get('OUTRAS', 0)))
                    break

            # Procurar SUBTRI (Substituição Tributária)
            siglas_subtri = ['SUBTRI', 'SUBTRI  ', 'SUBTRI ']
            for sigla in siglas_subtri:
                if sigla in item_tax_details:
                    for subtri_info in item_tax_details[sigla]:
                        valor_subtri = Decimal(str(subtri_info.get('VALOR', 0)))
                        summaries['total_subtri'] += valor_subtri
                        summaries['subtri_valores']['BCAL'] += Decimal(str(subtri_info.get('BASE', 0)))
                        summaries['subtri_valores']['VLOR'] += valor_subtri
                    break

            # Procurar DIFALI (Diferencial de Alíquota)
            siglas_difal = ['DIFALI', 'DIFALI  ', 'DIFAL']
            for sigla in siglas_difal:
                if sigla in item_tax_details:
                    for difal_info in item_tax_details[sigla]:
                        summaries['icms_difal']['VLOR'] += Decimal(str(difal_info.get('VALOR', 0)))
                    break
            
            # Acumula valores para o resumo por CFOP
            if cfop:
                summaries['cfop_summary'][cfop]['CODNAT'] = cfop
                summaries['cfop_summary'][cfop]['VC'] += item.get('valor_operacao', Decimal(0))
                
                # *** LÓGICA CORRIGIDA: Usar a mesma função de determinação ***
                # Buscar o primeiro ICMS para determinar o código fiscal correto
                codigo_fiscal_cfop = '3'  # Default
                for sigla in siglas_icms:
                    if sigla in item_tax_details and item_tax_details[sigla]:
                        primeiro_icms = item_tax_details[sigla][0]
                        codigo, _, _, _ = determinar_codigo_fiscal_e_valores(primeiro_icms)
                        if codigo:
                            codigo_fiscal_cfop = codigo
                        break
                
                summaries['cfop_summary'][cfop]['CODIGO_FISCAL'] = codigo_fiscal_cfop
                
                # Acumular valores por tipo de código fiscal
                for sigla in siglas_icms:
                    if sigla in item_tax_details:
                        for icms_info in item_tax_details[sigla]:
                            summaries['cfop_summary'][cfop]['BC'] += Decimal(str(icms_info.get('BASE', 0)))
                            summaries['cfop_summary'][cfop]['VI'] += Decimal(str(icms_info.get('VALOR', 0)))
                            summaries['cfop_summary'][cfop]['VS'] += Decimal(str(icms_info.get('ISENTAS', 0)))
                            summaries['cfop_summary'][cfop]['VO'] += Decimal(str(icms_info.get('OUTRAS', 0)))
                        break

    summaries['cfop_summary'] = list(summaries['cfop_summary'].values())
    print("  - Cálculo de resumos concluído.")
    return summaries

def calculate_detailed_summaries(transactions, tax_details):
    """Calcula resumos detalhados para os 5 demonstrativos necessários."""
    print("  - Calculando resumos detalhados para demonstrativos...")
    
    # Estruturas para os 5 demonstrativos
    summaries = {
        'total_icms': {'base': Decimal(0), 'valor': Decimal(0), 'isentas': Decimal(0), 'outras': Decimal(0)},
        'total_ipi': {'base': Decimal(0), 'valor': Decimal(0), 'isentas': Decimal(0), 'outras': Decimal(0)},
        'total_subtri': {'base': Decimal(0), 'valor': Decimal(0)},  # ← NOVO: Total ST
        'por_aliquota': defaultdict(lambda: {'base': Decimal(0), 'valor': Decimal(0), 'isentas': Decimal(0), 'outras': Decimal(0), 'valor_contabil': Decimal(0)}),
        'por_cfop': defaultdict(lambda: {
            'valor_contabil': Decimal(0), 
            'codigo_fiscal': '3',
            'tipo_1': {'base': Decimal(0), 'valor': Decimal(0)},
            'tipo_2': {'valor': Decimal(0)},
            'tipo_3': {'valor': Decimal(0)},
            'tipo_4': {'base': Decimal(0), 'valor': Decimal(0)},
            # Mantém campos antigos para compatibilidade
            'base': Decimal(0), 'valor': Decimal(0), 'isentas': Decimal(0), 'outras': Decimal(0)
        }),
        'por_estado': defaultdict(lambda: {
            'valor_contabil': Decimal(0),
            'tipo_1': {'base': Decimal(0), 'valor': Decimal(0)},
            'tipo_2': {'valor': Decimal(0)},
            'tipo_3': {'valor': Decimal(0)},
            'tipo_4': {'base': Decimal(0), 'valor': Decimal(0)}
        }),
        'total_contabil': Decimal(0),
        'total_difal': Decimal(0)
    }
    
    # Debug: encontrar siglas disponíveis
    siglas_encontradas = set()
    for doc in transactions:
        for item in doc['items']:
            item_tax_details = tax_details.get(item['codigo'], {})
            siglas_encontradas.update(item_tax_details.keys())
    print(f"  - Siglas encontradas: {sorted(siglas_encontradas)}")
    
    for doc in transactions:
        summaries['total_contabil'] += doc.get('valor_contabil_total', Decimal(0))
        uf_doc = doc.get('uf', 'XX')  # UF do documento
        
        for item in doc['items']:
            cfop = item.get('cfop')
            item_code = item['codigo']
            item_tax_details = tax_details.get(item_code, {})
            valor_operacao = item.get('valor_operacao', Decimal(0))
            
            # Acumular por UF - valor contábil total
            summaries['por_estado'][uf_doc]['valor_contabil'] += valor_operacao
            
            # Acumular por CFOP
            if cfop:
                summaries['por_cfop'][cfop]['valor_contabil'] += valor_operacao
            
            # Processar ICMS
            siglas_icms = ['ICMS', 'ICMS    ', 'ICMS   ']
            for sigla in siglas_icms:
                if sigla in item_tax_details:
                    for icms_info in item_tax_details[sigla]:
                        base = Decimal(str(icms_info.get('BASE', 0)))
                        valor = Decimal(str(icms_info.get('VALOR', 0)))
                        isentas = Decimal(str(icms_info.get('ISENTAS', 0)))
                        outras = Decimal(str(icms_info.get('OUTRAS', 0)))
                        aliquota = str(icms_info.get('ALIQUOTA', '0')).replace(',', '.')
                        
                        # Determinar tipo de operação para o estado
                        codigo_fiscal, _, _, _ = determinar_codigo_fiscal_e_valores(icms_info)
                        
                        if codigo_fiscal == "1" and base > 0:
                            # Tipo 1: Operações com crédito
                            summaries['por_estado'][uf_doc]['tipo_1']['base'] += base
                            summaries['por_estado'][uf_doc]['tipo_1']['valor'] += valor
                        elif codigo_fiscal == "2" and isentas > 0:
                            # Tipo 2: Isentas/Não tributadas
                            summaries['por_estado'][uf_doc]['tipo_2']['valor'] += isentas
                        elif codigo_fiscal == "3" and outras > 0:
                            # Tipo 3: Outras operações
                            summaries['por_estado'][uf_doc]['tipo_3']['valor'] += outras
                        
                        # Totais ICMS
                        summaries['total_icms']['base'] += base
                        summaries['total_icms']['valor'] += valor
                        summaries['total_icms']['isentas'] += isentas
                        summaries['total_icms']['outras'] += outras
                        
                        # Por alíquota
                        if aliquota and aliquota != '0':
                            summaries['por_aliquota'][aliquota]['base'] += base
                            summaries['por_aliquota'][aliquota]['valor'] += valor
                            summaries['por_aliquota'][aliquota]['isentas'] += isentas
                            summaries['por_aliquota'][aliquota]['outras'] += outras
                            summaries['por_aliquota'][aliquota]['valor_contabil'] += valor_operacao
                        
                        # Por CFOP
                        if cfop:
                            summaries['por_cfop'][cfop]['base'] += base
                            summaries['por_cfop'][cfop]['valor'] += valor
                            summaries['por_cfop'][cfop]['isentas'] += isentas
                            summaries['por_cfop'][cfop]['outras'] += outras
                            
                            # Determinar código fiscal e acumular por tipo
                            if codigo_fiscal:
                                summaries['por_cfop'][cfop]['codigo_fiscal'] = codigo_fiscal
                                
                                # Acumular por tipo específico
                                if codigo_fiscal == "1" and base > 0:
                                    summaries['por_cfop'][cfop]['tipo_1']['base'] += base
                                    summaries['por_cfop'][cfop]['tipo_1']['valor'] += valor
                                elif codigo_fiscal == "2" and isentas > 0:
                                    summaries['por_cfop'][cfop]['tipo_2']['valor'] += isentas
                                elif codigo_fiscal == "3" and outras > 0:
                                    summaries['por_cfop'][cfop]['tipo_3']['valor'] += outras
                    break
            
            # Processar SUBTRI (Substituição Tributária) - Tipo 4
            siglas_subtri = ['SUBTRI', 'SUBTRI  ', 'SUBTRI ']
            for sigla in siglas_subtri:
                if sigla in item_tax_details:
                    for subtri_info in item_tax_details[sigla]:
                        base_st = Decimal(str(subtri_info.get('BASE', 0)))
                        valor_st = Decimal(str(subtri_info.get('VALOR', 0)))
                        
                        # Totais ST globais
                        summaries['total_subtri']['base'] += base_st
                        summaries['total_subtri']['valor'] += valor_st
                        
                        if valor_st > 0:
                            # Tipo 4: Substituição Tributária por estado
                            summaries['por_estado'][uf_doc]['tipo_4']['base'] += base_st
                            summaries['por_estado'][uf_doc]['tipo_4']['valor'] += valor_st
                            
                            # Tipo 4: Substituição Tributária por CFOP
                            if cfop:
                                summaries['por_cfop'][cfop]['tipo_4']['base'] += base_st
                                summaries['por_cfop'][cfop]['tipo_4']['valor'] += valor_st
                    break
            
            # Processar IPI
            siglas_ipi = ['IPI', 'IPI     ', 'IPI    ']
            for sigla in siglas_ipi:
                if sigla in item_tax_details:
                    for ipi_info in item_tax_details[sigla]:
                        base = Decimal(str(ipi_info.get('BASE', 0)))
                        valor = Decimal(str(ipi_info.get('VALOR', 0)))
                        isentas = Decimal(str(ipi_info.get('ISENTAS', 0)))
                        outras = Decimal(str(ipi_info.get('OUTRAS', 0)))
                        
                        summaries['total_ipi']['base'] += base
                        summaries['total_ipi']['valor'] += valor
                        summaries['total_ipi']['isentas'] += isentas
                        summaries['total_ipi']['outras'] += outras
                    break
            
            # Processar DIFAL
            siglas_difal = ['DIFALI', 'DIFALI  ', 'DIFAL']
            for sigla in siglas_difal:
                if sigla in item_tax_details:
                    for difal_info in item_tax_details[sigla]:
                        summaries['total_difal'] += Decimal(str(difal_info.get('VALOR', 0)))
                    break
    
    # Converter defaultdicts para dicts normais
    summaries['por_aliquota'] = dict(summaries['por_aliquota'])
    summaries['por_cfop'] = dict(summaries['por_cfop'])
    summaries['por_estado'] = dict(summaries['por_estado'])
    
    print("  - Resumos detalhados calculados com sucesso.")
    return summaries

def create_summary_flowables(summaries_data):
    """Cria os 5 demonstrativos necessários: IPI, ICMS, Alíquota, CFOP e Estado."""
    if not summaries_data:
        return []
    
    flowables = []
    styles = getSampleStyleSheet()
    
    # Estilos
    title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], 
                                fontName='Helvetica-Bold', fontSize=11, spaceAfter=4, alignment=1)
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], 
                                 fontName='Helvetica-Bold', fontSize=9, alignment=1)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], 
                               fontName='Helvetica', fontSize=8, alignment=1)
    right_style = ParagraphStyle('RightStyle', parent=cell_style, alignment=2)
    
    # Verificar se há dados para mostrar antes de forçar nova página
    tem_dados = (summaries_data.get('total_ipi') or 
                 summaries_data.get('total_icms') or 
                 summaries_data.get('total_subtri') or  # ← ADICIONADO
                 summaries_data.get('por_aliquota') or 
                 summaries_data.get('por_cfop') or 
                 summaries_data.get('por_estado'))
    
    if not tem_dados:
        return flowables
    
    # ===== 1. TOTAL IPI MENSAL =====
    ipi_title = Paragraph("TOTAL IPI MENSAL", title_style)
    
    ipi_data = [
        [Paragraph("Código", header_style), Paragraph("Valor", header_style)],
        [Paragraph("1", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_ipi', {}).get('base', 0)), right_style)],
        [Paragraph("2", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_ipi', {}).get('valor', 0)), right_style)],
        [Paragraph("3", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_ipi', {}).get('isentas', 0)), right_style)],
        [Paragraph("4", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_ipi', {}).get('outras', 0)), right_style)]
    ]
    
    ipi_table = Table(ipi_data, colWidths=[3*cm, 7*cm])
    ipi_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    ipi_table.hAlign = 'CENTER'
    
    ipi_section = KeepTogether([ipi_title, ipi_table])
    flowables.append(ipi_section)
    flowables.append(Spacer(1, 0.8*cm))
    
    # ===== 2. TOTAL ICMS MENSAL =====
    icms_title = Paragraph("TOTAL ICMS MENSAL", title_style)
    
    icms_data = [
        [Paragraph("Código", header_style), Paragraph("Valor", header_style)],
        [Paragraph("1", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_icms', {}).get('base', 0)), right_style)],
        [Paragraph("2", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_icms', {}).get('valor', 0)), right_style)],
        [Paragraph("3", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_icms', {}).get('isentas', 0)), right_style)],
        [Paragraph("4", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_icms', {}).get('outras', 0)), right_style)],
        [Paragraph("ST", cell_style), 
         Paragraph(format_currency(summaries_data.get('total_subtri', {}).get('valor', 0)), right_style)]
    ]
    
    # Adicionar DIFAL se houver
    if summaries_data.get('total_difal', 0) > 0:
        icms_data.append([
            Paragraph("DIFAL", cell_style),
            Paragraph(format_currency(summaries_data.get('total_difal', 0)), right_style)
        ])
    
    icms_table = Table(icms_data, colWidths=[3*cm, 7*cm])
    icms_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    icms_table.hAlign = 'CENTER'
    
    icms_section = KeepTogether([icms_title, icms_table])
    flowables.append(icms_section)
    flowables.append(Spacer(1, 0.8*cm))
    
    # ===== 3. DEMONSTRATIVO POR ALÍQUOTA =====
    if summaries_data.get('por_aliquota'):
        aliq_title = Paragraph("DEMONSTRATIVO POR ALÍQUOTA", title_style)
        
        aliq_data = [
            [Paragraph("Alíquota %", header_style), Paragraph("Valor Contábil", header_style), 
             Paragraph("Base Cálculo", header_style), Paragraph("Imposto Creditado", header_style)]
        ]
        
        for aliquota in sorted(summaries_data['por_aliquota'].keys(), key=lambda x: float(x) if x.replace('.','').replace(',','').isdigit() else 0):
            dados = summaries_data['por_aliquota'][aliquota]
            aliq_data.append([
                Paragraph(f"{aliquota}%", cell_style),
                Paragraph(format_currency(dados.get('valor_contabil', 0)), right_style),
                Paragraph(format_currency(dados.get('base', 0)), right_style),
                Paragraph(format_currency(dados.get('valor', 0)), right_style)
            ])
        
        aliq_table = Table(aliq_data, colWidths=[2.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        aliq_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        aliq_table.hAlign = 'CENTER'
        
        aliq_section = KeepTogether([aliq_title, aliq_table])
        flowables.append(aliq_section)
        flowables.append(Spacer(1, 0.8*cm))
    
    # ===== 4. DEMONSTRATIVO POR CFOP =====
    if summaries_data.get('por_cfop'):
        cfop_title = Paragraph("DEMONSTRATIVO POR CFOP", title_style)
        
        cfop_data = [
            [Paragraph("CFOP", header_style), 
             Paragraph("Valor Contábil", header_style),
             Paragraph("1-Base", header_style), 
             Paragraph("1-Valor", header_style),
             Paragraph("2-Isentas", header_style),
             Paragraph("3-Outras", header_style),
             Paragraph("4-Base ST", header_style),
             Paragraph("4-Imposto Creditado", header_style)]
        ]
        
        for cfop in sorted(summaries_data['por_cfop'].keys()):
            dados = summaries_data['por_cfop'][cfop]
            
            cfop_data.append([
                Paragraph(format_codificacao_fiscal(cfop), cell_style),
                Paragraph(format_currency(dados.get('valor_contabil', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_1', {}).get('base', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_1', {}).get('valor', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_2', {}).get('valor', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_3', {}).get('valor', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_4', {}).get('base', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_4', {}).get('valor', 0)), right_style)
            ])
        
        cfop_table = Table(cfop_data, colWidths=[1.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm])
        cfop_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),  # Valores à direita
            ('ALIGN', (0,0), (0,-1), 'CENTER'),  # CFOP centralizado
        ]))
        cfop_table.hAlign = 'CENTER'
        
        cfop_section = KeepTogether([cfop_title, cfop_table])
        flowables.append(cfop_section)
        flowables.append(Spacer(1, 0.8*cm))
    
    # ===== 5. DEMONSTRATIVO POR ESTADO =====
    if summaries_data.get('por_estado'):
        estado_title = Paragraph("DEMONSTRATIVO POR ESTADO", title_style)
        
        estado_data = [
            [Paragraph("UF", header_style), 
             Paragraph("Valor Contábil", header_style),
             Paragraph("1-Base", header_style), 
             Paragraph("1-Valor", header_style),
             Paragraph("2-Isentas", header_style),
             Paragraph("3-Outras", header_style),
             Paragraph("4-Base ST", header_style),
             Paragraph("4-Imposto Creditado", header_style)]  # ← CORRIGIDO
        ]
        
        for uf in sorted(summaries_data['por_estado'].keys()):
            dados = summaries_data['por_estado'][uf]
            
            estado_data.append([
                Paragraph(uf, cell_style),
                Paragraph(format_currency(dados.get('valor_contabil', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_1', {}).get('base', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_1', {}).get('valor', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_2', {}).get('valor', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_3', {}).get('valor', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_4', {}).get('base', 0)), right_style),
                Paragraph(format_currency(dados.get('tipo_4', {}).get('valor', 0)), right_style)
            ])
        
        estado_table = Table(estado_data, colWidths=[1.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm])
        estado_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),  # Valores à direita
            ('ALIGN', (0,0), (0,-1), 'CENTER'),  # UF centralizada
        ]))
        estado_table.hAlign = 'CENTER'
        
        estado_section = KeepTogether([estado_title, estado_table])
        flowables.append(estado_section)
    
    return flowables

# ===== NOVAS FUNÇÕES PARA GERAR XLSX =====

def create_xlsx_main_table_data(transactions, tax_details, impostos_raw):
    """Cria os dados da tabela principal para o XLSX (mesmo conteúdo do PDF)."""
    data = []
    
    # ===== CABEÇALHO SIMPLIFICADO E ALINHADO CORRETAMENTE =====
    # Uma única linha de cabeçalho para garantir alinhamento perfeito
    headers = [
         "DATA DE\nENTRADA", "ESPÉCIE", "SÉRIE/\nSUB-SÉRIE", "NÚMERO", "DATA DO\nDOCUMENTO", 
        "CÓD. DO\nEMITENTE", "UF\nORIGEM", "VALOR\nCONTÁBIL", "CODIFICAÇÃO\nCONTÁBIL", 
        "CODIFICAÇÃO\nFISCAL", "ICMS\nCOD (*)", "ICMS BASE CÁLCULO\nVALOR OPERAÇÃO", 
        "ICMS\nALÍQ. %", "ICMS IMPOSTO\nCREDITADO", "IPI\nCOD (*)", 
        "IPI BASE DE CÁLCULO\nVALOR DA OPERAÇÃO", "IPI IMPOSTO\nCREDITADO", "OBSERVAÇÕES"
    ]
    data.append(headers)
    
    linha_vazia_modelo = [""] * 19  # ← Coluna A vazia + 18 de conteúdo
    
    for doc in transactions:
        segmentos = doc['items']
        segmentos_ordenados = sorted(segmentos, key=lambda x: x.get('cfop', 0))
        primeiro_segmento = segmentos_ordenados[0]
        
        vcon_total = sum(float(seg.get('valor_operacao', 0.0)) for seg in segmentos)
        
        # Obter CFOP para decidir se agrupa ou não
        cfop_primeiro_segmento = primeiro_segmento.get('cfop', 0)
        primeiro_digito_cfop = str(cfop_primeiro_segmento)[0] if cfop_primeiro_segmento else '0'
        deve_agrupar = primeiro_digito_cfop == '1'
        
        # Obter e ordenar todos os detalhes do primeiro segmento por sequencial
        detalhes_primeiro_segmento = sorted(impostos_raw.get(str(primeiro_segmento['codigo']), []), key=lambda x: x.get('SEQUENCIAL', 0))
        
        if deve_agrupar:
            # LÓGICA COM AGRUPAMENTO (CFOP iniciado com '1')
            impostos_agrupados = {}
            for detalhe in detalhes_primeiro_segmento:
                imposto_tipo = detalhe.get('IMPOSTO')
                if imposto_tipo not in impostos_agrupados:
                    impostos_agrupados[imposto_tipo] = {
                        'BASE_TOTAL': 0, 'ISENTAS_TOTAL': 0, 'OUTRAS_TOTAL': 0, 'VALOR_TOTAL': 0,
                        'ALIQUOTA': detalhe.get('ALIQUOTA', ''), 'REGISTROS': []
                    }
                
                try:
                    impostos_agrupados[imposto_tipo]['BASE_TOTAL'] += float(detalhe.get('BASE', 0))
                    impostos_agrupados[imposto_tipo]['ISENTAS_TOTAL'] += float(detalhe.get('ISENTAS', 0))
                    impostos_agrupados[imposto_tipo]['OUTRAS_TOTAL'] += float(detalhe.get('OUTRAS', 0))
                    impostos_agrupados[imposto_tipo]['VALOR_TOTAL'] += float(detalhe.get('VALOR', 0))
                    impostos_agrupados[imposto_tipo]['REGISTROS'].append(detalhe)
                except (ValueError, TypeError):
                    pass
            
            primeiro_icms_agrupado = impostos_agrupados.get(1)
            primeiro_ipi_agrupado = impostos_agrupados.get(2)
            
            # Determinar códigos e valores ICMS
            icms_cod = icms_base = icms_aliq = icms_imposto = ""
            if primeiro_icms_agrupado:
                detalhe_simulado = {
                    'BASE': primeiro_icms_agrupado['BASE_TOTAL'],
                    'ISENTAS': primeiro_icms_agrupado['ISENTAS_TOTAL'],
                    'OUTRAS': primeiro_icms_agrupado['OUTRAS_TOTAL'],
                    'VALOR': primeiro_icms_agrupado['VALOR_TOTAL']
                }
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                
                if codigo and valor_base > 0:
                    icms_cod = codigo
                    icms_base = format_currency(valor_base)
                    if mostrar_aliquota:
                        icms_aliq = format_currency(primeiro_icms_agrupado['ALIQUOTA']) if primeiro_icms_agrupado['ALIQUOTA'] else ""
                    if codigo == "1":
                        icms_imposto = format_currency(valor_imposto_creditado)
            
            # Determinar códigos e valores IPI
            ipi_cod = ipi_base = ipi_imposto = ""
            if primeiro_ipi_agrupado:
                detalhe_simulado = {
                    'BASE': primeiro_ipi_agrupado['BASE_TOTAL'],
                    'ISENTAS': primeiro_ipi_agrupado['ISENTAS_TOTAL'],
                    'OUTRAS': primeiro_ipi_agrupado['OUTRAS_TOTAL'],
                    'VALOR': primeiro_ipi_agrupado['VALOR_TOTAL']
                }
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                
                if codigo and valor_base > 0:
                    ipi_cod = codigo
                    ipi_base = format_currency(valor_base)
                    if codigo == "1":
                        ipi_imposto = format_currency(valor_imposto_creditado)
                    
            # Marca os detalhes usados no cabeçalho
            detalhes_usados = set()
            if primeiro_icms_agrupado and icms_cod:
                for registro in primeiro_icms_agrupado['REGISTROS']:
                    detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
            if primeiro_ipi_agrupado and ipi_cod:
                for registro in primeiro_ipi_agrupado['REGISTROS']:
                    detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
        else:
            # LÓGICA SEM AGRUPAMENTO (CFOP iniciado com '2', '3', etc.)
            icms_para_cabecalho = get_imposto_para_cabecalho(detalhes_primeiro_segmento, 1)
            ipi_para_cabecalho = get_imposto_para_cabecalho(detalhes_primeiro_segmento, 2)
            
            # Determinar códigos e valores ICMS
            icms_cod = icms_base = icms_aliq = icms_imposto = ""
            if icms_para_cabecalho:
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(icms_para_cabecalho)
                if codigo and valor_base > 0:
                    icms_cod = codigo
                    icms_base = format_currency(valor_base)
                    if mostrar_aliquota:
                        icms_aliq = format_currency(icms_para_cabecalho.get("ALIQUOTA", "")) if icms_para_cabecalho.get("ALIQUOTA") else ""
                    if codigo == "1":
                        icms_imposto = format_currency(valor_imposto_creditado)
            
            # Determinar códigos e valores IPI
            ipi_cod = ipi_base = ipi_imposto = ""
            if ipi_para_cabecalho:
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(ipi_para_cabecalho)
                if codigo and valor_base > 0:
                    ipi_cod = codigo
                    ipi_base = format_currency(valor_base)
                    if codigo == "1":
                        ipi_imposto = format_currency(valor_imposto_creditado)
                    
            # Marca os detalhes usados no cabeçalho
            detalhes_usados = set()
            if icms_para_cabecalho and icms_cod: 
                detalhes_usados.add((icms_para_cabecalho.get("NOTA"), icms_para_cabecalho.get("SEQUENCIAL")))
            if ipi_para_cabecalho and ipi_cod: 
                detalhes_usados.add((ipi_para_cabecalho.get("NOTA"), ipi_para_cabecalho.get("SEQUENCIAL")))

        # Verificar DIFALI para observações
        observacoes = ""
        tem_subtri = False
        
        # Primeiro, verifica se tem SUBTRI
        for segmento in segmentos:
            detalhes_segmento = impostos_raw.get(str(segmento['codigo']), [])
            for detalhe in detalhes_segmento:
                if detalhe.get('SIGLA', '').strip() == 'SUBTRI':
                    try:
                        valor_subtri = float(detalhe.get('VALOR', 0))
                        if valor_subtri > 0:
                            tem_subtri = True
                            break
                    except (ValueError, TypeError):
                        pass
            if tem_subtri:
                break
        
        # Só busca DIFALI se NÃO tiver SUBTRI
        if not tem_subtri:
            for segmento in segmentos:
                detalhes_segmento = impostos_raw.get(str(segmento['codigo']), [])
                for detalhe in detalhes_segmento:
                    if detalhe.get('SIGLA', '').strip() == 'DIFALI' and detalhe.get('VALOR'):
                        try:
                            valor_difali = float(detalhe.get('VALOR', 0))
                            if valor_difali > 0:
                                observacoes = f"Diferencial de Alíquota: {format_currency(valor_difali)}"
                                break
                        except (ValueError, TypeError):
                            pass
                if observacoes:
                    break

        # Criar linha de cabeçalho
        serie_sub = doc.get('serie', '')
        if doc.get('sub_serie') and doc['sub_serie'] != '0': 
            serie_sub += f"/{doc['sub_serie']}"
        
        linha_cabecalho = [
            format_date(doc['data_entrada']), doc['especie'], serie_sub, doc['numero'],
            format_date(doc['data_doc']), doc['cod_emitente'], doc['uf'], f"{vcon_total:.2f}",
            "", format_codificacao_fiscal(primeiro_segmento.get('cfop')), icms_cod, 
            icms_base, icms_aliq, icms_imposto, ipi_cod, ipi_base, ipi_imposto, observacoes
        ]
        data.append(linha_cabecalho)

        # Loop para gerar linhas de detalhe
        for i, segmento_atual in enumerate(segmentos_ordenados):
            # Se for um segmento adicional, gera a linha principal dele primeiro
            if i > 0:
                cfop_segmento_atual = segmento_atual.get('cfop', 0)
                primeiro_digito_cfop_seg = str(cfop_segmento_atual)[0] if cfop_segmento_atual else '0'
                deve_agrupar_seg = primeiro_digito_cfop_seg == '1'
                
                detalhes_atuais_ordenados = sorted(impostos_raw.get(str(segmento_atual['codigo']), []), key=lambda x: x.get('SEQUENCIAL', 0))
                
                if deve_agrupar_seg:
                    # Lógica com agrupamento para segmentos
                    impostos_agrupados_seg = {}
                    for detalhe in detalhes_atuais_ordenados:
                        imposto_tipo = detalhe.get('IMPOSTO')
                        if imposto_tipo not in impostos_agrupados_seg:
                            impostos_agrupados_seg[imposto_tipo] = {
                                'BASE_TOTAL': 0, 'ISENTAS_TOTAL': 0, 'OUTRAS_TOTAL': 0, 'VALOR_TOTAL': 0,
                                'ALIQUOTA': detalhe.get('ALIQUOTA', ''), 'REGISTROS': []
                            }
                        
                        try:
                            impostos_agrupados_seg[imposto_tipo]['BASE_TOTAL'] += float(detalhe.get('BASE', 0))
                            impostos_agrupados_seg[imposto_tipo]['ISENTAS_TOTAL'] += float(detalhe.get('ISENTAS', 0))
                            impostos_agrupados_seg[imposto_tipo]['OUTRAS_TOTAL'] += float(detalhe.get('OUTRAS', 0))
                            impostos_agrupados_seg[imposto_tipo]['VALOR_TOTAL'] += float(detalhe.get('VALOR', 0))
                            impostos_agrupados_seg[imposto_tipo]['REGISTROS'].append(detalhe)
                        except (ValueError, TypeError):
                            pass
                    
                    icms_principal_seg_agrupado = impostos_agrupados_seg.get(1)
                    ipi_principal_seg_agrupado = impostos_agrupados_seg.get(2)
                    
                    # Determinar códigos e valores ICMS para segmento
                    icms_cod_seg = icms_base_seg = icms_aliq_seg = icms_imposto_seg = ""
                    if icms_principal_seg_agrupado:
                        detalhe_simulado = {
                            'BASE': icms_principal_seg_agrupado['BASE_TOTAL'],
                            'ISENTAS': icms_principal_seg_agrupado['ISENTAS_TOTAL'],
                            'OUTRAS': icms_principal_seg_agrupado['OUTRAS_TOTAL'],
                            'VALOR': icms_principal_seg_agrupado['VALOR_TOTAL']
                        }
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                        
                        if codigo and valor_base > 0:
                            icms_cod_seg = codigo
                            icms_base_seg = format_currency(valor_base)
                            if mostrar_aliquota:
                                icms_aliq_seg = format_currency(icms_principal_seg_agrupado['ALIQUOTA']) if icms_principal_seg_agrupado['ALIQUOTA'] else ""
                            if codigo == "1":
                                icms_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Determinar códigos e valores IPI para segmento
                    ipi_cod_seg = ipi_base_seg = ipi_imposto_seg = ""
                    if ipi_principal_seg_agrupado:
                        detalhe_simulado = {
                            'BASE': ipi_principal_seg_agrupado['BASE_TOTAL'],
                            'ISENTAS': ipi_principal_seg_agrupado['ISENTAS_TOTAL'],
                            'OUTRAS': ipi_principal_seg_agrupado['OUTRAS_TOTAL'],
                            'VALOR': ipi_principal_seg_agrupado['VALOR_TOTAL']
                        }
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                        
                        if codigo and valor_base > 0:
                            ipi_cod_seg = codigo
                            ipi_base_seg = format_currency(valor_base)
                            if codigo == "1":
                                ipi_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Marca os detalhes usados
                    if icms_principal_seg_agrupado and icms_cod_seg:
                        for registro in icms_principal_seg_agrupado['REGISTROS']:
                            detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
                    if ipi_principal_seg_agrupado and ipi_cod_seg:
                        for registro in ipi_principal_seg_agrupado['REGISTROS']:
                            detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
                else:
                    # Lógica sem agrupamento para segmentos
                    icms_seg_para_cabecalho = get_imposto_para_cabecalho(detalhes_atuais_ordenados, 1)
                    ipi_seg_para_cabecalho = get_imposto_para_cabecalho(detalhes_atuais_ordenados, 2)
                    
                    # Determinar códigos e valores ICMS para segmento
                    icms_cod_seg = icms_base_seg = icms_aliq_seg = icms_imposto_seg = ""
                    if icms_seg_para_cabecalho:
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(icms_seg_para_cabecalho)
                        if codigo and valor_base > 0:
                            icms_cod_seg = codigo
                            icms_base_seg = format_currency(valor_base)
                            if mostrar_aliquota:
                                icms_aliq_seg = format_currency(icms_seg_para_cabecalho.get("ALIQUOTA", "")) if icms_seg_para_cabecalho.get("ALIQUOTA") else ""
                            if codigo == "1":
                                icms_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Determinar códigos e valores IPI para segmento
                    ipi_cod_seg = ipi_base_seg = ipi_imposto_seg = ""
                    if ipi_seg_para_cabecalho:
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(ipi_seg_para_cabecalho)
                        if codigo and valor_base > 0:
                            ipi_cod_seg = codigo
                            ipi_base_seg = format_currency(valor_base)
                            if codigo == "1":
                                ipi_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Marca os detalhes usados
                    if icms_seg_para_cabecalho and icms_cod_seg: 
                        detalhes_usados.add((icms_seg_para_cabecalho.get("NOTA"), icms_seg_para_cabecalho.get("SEQUENCIAL")))
                    if ipi_seg_para_cabecalho and ipi_cod_seg: 
                        detalhes_usados.add((ipi_seg_para_cabecalho.get("NOTA"), ipi_seg_para_cabecalho.get("SEQUENCIAL")))
                
                # Criar linha de segmento (começando com coluna A vazia)
                linha_segmento = linha_vazia_modelo.copy()
                linha_segmento[0] = ""  # Coluna A vazia
                linha_segmento[10] = format_codificacao_fiscal(segmento_atual.get('cfop'))  # Código fiscal
                linha_segmento[11] = icms_cod_seg
                linha_segmento[12] = icms_base_seg
                linha_segmento[13] = icms_aliq_seg
                linha_segmento[14] = icms_imposto_seg
                linha_segmento[15] = ipi_cod_seg
                linha_segmento[16] = ipi_base_seg
                linha_segmento[17] = ipi_imposto_seg
                data.append(linha_segmento)

            # Gerar linhas de detalhe esparsas
            def get_ordem_detalhes(codigo_segmento, impostos_raw):
                detalhes_segmento = impostos_raw.get(str(codigo_segmento), [])
                
                # Conta registros por tipo de imposto
                contadores = {}
                for detalhe in detalhes_segmento:
                    tipo = detalhe.get('IMPOSTO')
                    contadores[tipo] = contadores.get(tipo, 0) + 1
                
                # Ordena cada tipo conforme sua quantidade
                detalhes_ordenados = []
                
                # Processa ICMS (tipo 1) e IPI (tipo 2) primeiro
                for tipo in [1, 2]:
                    registros_tipo = [d for d in detalhes_segmento if d.get('IMPOSTO') == tipo]
                    if not registros_tipo:
                        continue
                    
                    qtd_registros = contadores.get(tipo, 0)
                    if qtd_registros == 2:
                        # Apenas 2 registros: ordem CRESCENTE
                        registros_ordenados = sorted(registros_tipo, key=lambda x: x.get('SEQUENCIAL', 0))
                    else:
                        # 1, 3+ registros: ordem DECRESCENTE
                        registros_ordenados = sorted(registros_tipo, key=lambda x: x.get('SEQUENCIAL', 0), reverse=True)
                    
                    detalhes_ordenados.extend(registros_ordenados)
                
                # Adiciona outros tipos sempre em ordem decrescente
                outros_tipos = [d for d in detalhes_segmento if d.get('IMPOSTO') not in [1, 2]]
                outros_ordenados = sorted(outros_tipos, key=lambda x: x.get('SEQUENCIAL', 0), reverse=True)
                detalhes_ordenados.extend(outros_ordenados)
                
                return detalhes_ordenados
            
            detalhes_atuais_ordenados_especial = get_ordem_detalhes(segmento_atual['codigo'], impostos_raw)
            
            for detalhe in detalhes_atuais_ordenados_especial:
                chave_detalhe = (detalhe.get("NOTA"), detalhe.get("SEQUENCIAL"))
                if chave_detalhe in detalhes_usados:
                    continue

                sigla = detalhe.get('SIGLA', '').strip().upper()
                imposto_id = detalhe.get('IMPOSTO')
                
                linha_detalhe = linha_vazia_modelo.copy()
                gerou_linha = False
                
                if sigla == 'SUBTRI':
                    try:
                        if float(detalhe.get("VALOR", 0.0)) > 0:
                            linha_detalhe[11] = "ST"  # ← MUDOU: índice +1
                            linha_detalhe[12] = format_currency(detalhe.get("BASE", ""))  # ← MUDOU: índice +1
                            linha_detalhe[14] = format_currency_or_blank(detalhe.get("VALOR", ""))  # ← MUDOU: índice +1
                            gerou_linha = True
                    except (ValueError, TypeError):
                        pass
                elif imposto_id == 1:  # ICMS
                    codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe)
                    if codigo and valor_base > 0:
                        linha_detalhe[11] = codigo  # ← MUDOU: índice +1
                        linha_detalhe[12] = format_currency(valor_base)  # ← MUDOU: índice +1
                        if mostrar_aliquota:
                            linha_detalhe[13] = format_currency(detalhe.get("ALIQUOTA", "")) if detalhe.get("ALIQUOTA") else ""  # ← MUDOU: índice +1
                        if codigo == "1":
                            linha_detalhe[14] = format_currency(valor_imposto_creditado)  # ← MUDOU: índice +1
                        gerou_linha = True
                elif imposto_id == 2:  # IPI
                    codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe)
                    if codigo and valor_base > 0:
                        linha_detalhe[15] = codigo  # ← MUDOU: índice +1
                        linha_detalhe[16] = format_currency(valor_base)  # ← MUDOU: índice +1
                        if codigo == "1":
                            linha_detalhe[17] = format_currency(valor_imposto_creditado)  # ← MUDOU: índice +1
                        gerou_linha = True
                
                if gerou_linha:
                    data.append(linha_detalhe)
                    detalhes_usados.add(chave_detalhe)

    return data

def create_xlsx_summaries_data(summaries_data):
    """Cria os dados dos demonstrativos para adicionar após a tabela principal (igual ao PDF)."""
    if not summaries_data:
        return []
    
    all_data = []
    
    # Verificar se há dados para mostrar
    tem_dados = (summaries_data.get('total_ipi') or 
                 summaries_data.get('total_icms') or 
                 summaries_data.get('total_subtri') or 
                 summaries_data.get('por_aliquota') or 
                 summaries_data.get('por_cfop') or 
                 summaries_data.get('por_estado'))
    
    if not tem_dados:
        return all_data
    
    # QUEBRA DE PÁGINA (linhas vazias para simular nova página)
    all_data.extend([[""] * 19 for _ in range(5)])  # Espaçamento para quebra de página
    
    # ===== 1. TOTAL IPI MENSAL =====
    all_data.append(["TOTAL IPI MENSAL"] + [""] * 18)
    all_data.append([""] * 19)
    all_data.append(["Código", "Valor"] + [""] * 17)
    all_data.append(["1", format_currency(summaries_data.get('total_ipi', {}).get('base', 0))] + [""] * 17)
    all_data.append(["2", format_currency(summaries_data.get('total_ipi', {}).get('valor', 0))] + [""] * 17)
    all_data.append(["3", format_currency(summaries_data.get('total_ipi', {}).get('isentas', 0))] + [""] * 17)
    all_data.append(["4", format_currency(summaries_data.get('total_ipi', {}).get('outras', 0))] + [""] * 17)
    all_data.append([""] * 19)  # Espaçamento
    
    # ===== 2. TOTAL ICMS MENSAL =====
    all_data.append(["TOTAL ICMS MENSAL"] + [""] * 18)
    all_data.append([""] * 19)
    all_data.append(["Código", "Valor"] + [""] * 17)
    all_data.append(["1", format_currency(summaries_data.get('total_icms', {}).get('base', 0))] + [""] * 17)
    all_data.append(["2", format_currency(summaries_data.get('total_icms', {}).get('valor', 0))] + [""] * 17)
    all_data.append(["3", format_currency(summaries_data.get('total_icms', {}).get('isentas', 0))] + [""] * 17)
    all_data.append(["4", format_currency(summaries_data.get('total_icms', {}).get('outras', 0))] + [""] * 17)
    all_data.append(["ST", format_currency(summaries_data.get('total_subtri', {}).get('valor', 0))] + [""] * 17)
    
    # Adicionar DIFAL se houver
    if summaries_data.get('total_difal', 0) > 0:
        all_data.append(["DIFAL", format_currency(summaries_data.get('total_difal', 0))] + [""] * 17)
    
    all_data.append([""] * 19)  # Espaçamento
    
    # ===== 3. DEMONSTRATIVO POR ALÍQUOTA =====
    if summaries_data.get('por_aliquota'):
        all_data.append(["DEMONSTRATIVO POR ALÍQUOTA"] + [""] * 18)
        all_data.append([""] * 19)
        all_data.append(["Alíquota %", "Valor Contábil", "Base Cálculo", "Imposto Creditado"] + [""] * 15)
        
        for aliquota in sorted(summaries_data['por_aliquota'].keys(), key=lambda x: float(x) if x.replace('.','').replace(',','').isdigit() else 0):
            dados = summaries_data['por_aliquota'][aliquota]
            all_data.append([
                f"{aliquota}%",
                format_currency(dados.get('valor_contabil', 0)),
                format_currency(dados.get('base', 0)),
                format_currency(dados.get('valor', 0))
            ] + [""] * 15)
        
        all_data.append([""] * 19)  # Espaçamento
    
    # ===== 4. DEMONSTRATIVO POR CFOP =====
    if summaries_data.get('por_cfop'):
        all_data.append(["DEMONSTRATIVO POR CFOP"] + [""] * 18)
        all_data.append([""] * 19)
        all_data.append(["CFOP", "Valor Contábil", "1-Base", "1-Valor", "2-Isentas", "3-Outras", "4-Base ST", "4-Imposto Creditado"] + [""] * 11)
        
        for cfop in sorted(summaries_data['por_cfop'].keys()):
            dados = summaries_data['por_cfop'][cfop]
            all_data.append([
                format_codificacao_fiscal(cfop),
                format_currency(dados.get('valor_contabil', 0)),
                format_currency(dados.get('tipo_1', {}).get('base', 0)),
                format_currency(dados.get('tipo_1', {}).get('valor', 0)),
                format_currency(dados.get('tipo_2', {}).get('valor', 0)),
                format_currency(dados.get('tipo_3', {}).get('valor', 0)),
                format_currency(dados.get('tipo_4', {}).get('base', 0)),
                format_currency(dados.get('tipo_4', {}).get('valor', 0))
            ] + [""] * 11)
        
        all_data.append([""] * 19)  # Espaçamento
    
    # ===== 5. DEMONSTRATIVO POR ESTADO =====
    if summaries_data.get('por_estado'):
        all_data.append(["DEMONSTRATIVO POR ESTADO"] + [""] * 18)
        all_data.append([""] * 19)
        all_data.append(["UF", "Valor Contábil", "1-Base", "1-Valor", "2-Isentas", "3-Outras", "4-Base ST", "4-Imposto Creditado"] + [""] * 11)
        
        for uf in sorted(summaries_data['por_estado'].keys()):
            dados = summaries_data['por_estado'][uf]
            all_data.append([
                uf,
                format_currency(dados.get('valor_contabil', 0)),
                format_currency(dados.get('tipo_1', {}).get('base', 0)),
                format_currency(dados.get('tipo_1', {}).get('valor', 0)),
                format_currency(dados.get('tipo_2', {}).get('valor', 0)),
                format_currency(dados.get('tipo_3', {}).get('valor', 0)),
                format_currency(dados.get('tipo_4', {}).get('base', 0)),
                format_currency(dados.get('tipo_4', {}).get('valor', 0))
            ] + [""] * 11)
    
    return all_data

def create_xlsx_file(filename, main_data, summaries_data, header_info):
    """Cria o arquivo XLSX 100% igual ao PDF - UMA ÚNICA ABA com paginação."""
    print(f"  - Criando arquivo XLSX: {filename}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Livro Registro P1"
    
    # ===== CABEÇALHO DA EMPRESA (igual ao PDF) =====
    empresa = header_info.get('empresa_nome', 'EMPRESA PADRAO')
    insc_est = header_info.get('ie', '')
    if not insc_est or insc_est == 'N/A':
        insc_est = ''
    
    cnpj_raw = str(header_info.get('cnpj', 'N/A'))
    cnpj = f"{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}/{cnpj_raw[8:12]}-{cnpj_raw[12:]}" if len(cnpj_raw) == 14 else cnpj_raw
    periodo = f"{format_date(header_info.get('periodo_inicio'))} a {format_date(header_info.get('periodo_fim'))}"
    
    # Linha 1: Título principal (começando na coluna B)
    ws.merge_cells('B1:S1')  # ← MUDOU: de A1:R1 para B1:S1
    ws['B1'] = "LIVRO REGISTRO DE ENTRADAS - RE - MODELO P1"
    ws['B1'].font = Font(bold=True, size=14)
    ws['B1'].alignment = Alignment(horizontal='center')
    
    # Linha 3: Título seção
    ws['B3'] = "REGISTRO DE ENTRADAS"
    ws['B3'].font = Font(bold=True, size=10)
    
    # Informações da empresa (linhas 4-7)
    ws['B4'] = f"EMPRESA: {empresa}"
    ws['B5'] = f"INSC.EST.: {insc_est}"
    ws['B6'] = f"CNPJ: {cnpj}"
    ws['B7'] = f"FOLHA: 001"
    ws['H7'] = f"MÊS OU PERÍODO/ANO: {periodo}"  # ← MUDOU: de G7 para H7
    
    # Códigos fiscais (lado direito - igual ao PDF)
    ws['M3'] = "(*) CÓDIGO DE VALORES FISCAIS"  # ← MUDOU: de L3 para M3
    ws['M3'].font = Font(bold=True, size=10)
    ws['M4'] = "1-OPERAÇÕES COM CRÉDITO DO IMPOSTO"  # ← MUDOU: de L4 para M4
    ws['M5'] = "2-OPER. SEM CRÉDITO DO IMPOSTO - ISENTAS/NÃO TRIBUTADAS"  # ← MUDOU: de L5 para M5
    ws['M6'] = "3-OPERAÇÕES SEM CRÉDITO DO IMPOSTO - OUTRAS"  # ← MUDOU: de L6 para M6
    ws['M7'] = "4-ICMS RETIDO POR SUBST.TRIBUTÁRIA"  # ← MUDOU: de L7 para M7
    
    # ===== TABELA PRINCIPAL (a partir da linha 9, coluna B) =====
    start_row = 9
    for row_idx, row_data in enumerate(main_data):
        for col_idx, cell_value in enumerate(row_data):
            cell = ws.cell(row=start_row + row_idx, column=col_idx + 2, value=cell_value)  # ← MUDOU: +2 para começar na coluna B
            
            # Formatação do cabeçalho da tabela (APENAS a primeira linha)
            if row_idx == 0:  # ← CORRIGIDO: apenas linha 0 (cabeçalho)
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            else:  # ← TODAS as outras linhas (dados) SEM fundo
                cell.font = Font(size=8)
                # Valores numéricos alinhados à direita (igual ao PDF)
                if col_idx in [8, 12, 13, 14, 16, 17]:  # Colunas de valores
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx == 18:  # Observações
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center')
    
    # ===== DEMONSTRATIVOS APÓS A TABELA (igual ao PDF) =====
    summaries_data_list = create_xlsx_summaries_data(summaries_data)
    
    if summaries_data_list:
        # Calcular linha de início dos demonstrativos
        start_summaries_row = start_row + len(main_data) + 1
        
        # SEPARAÇÃO SIMPLES (espaçamento manual ao invés de quebra de página programática)
        # A quebra de página será automática pelo Excel quando imprimir
        
        for row_idx, row_data in enumerate(summaries_data_list):
            for col_idx, cell_value in enumerate(row_data):
                if cell_value:  # Só preenche se não for vazio
                    cell = ws.cell(row=start_summaries_row + row_idx, column=col_idx + 2, value=cell_value)  # ← MUDOU: +2 para começar na coluna B
                    
                    # Formatação dos títulos dos demonstrativos
                    if cell_value in ["TOTAL IPI MENSAL", "TOTAL ICMS MENSAL", "DEMONSTRATIVO POR ALÍQUOTA", 
                                     "DEMONSTRATIVO POR CFOP", "DEMONSTRATIVO POR ESTADO"]:
                        cell.font = Font(bold=True, size=9)  # ← CORRIGIDO: 9pt ao invés de 12
                        cell.alignment = Alignment(horizontal='center')
                    # Formatação dos cabeçalhos das tabelas
                    elif cell_value in ["Código", "Valor", "Alíquota %", "Valor Contábil", "Base Cálculo", 
                                      "Imposto Creditado", "CFOP", "1-Base", "1-Valor", "2-Isentas", 
                                      "3-Outras", "4-Base ST", "4-Imposto Creditado", "UF"]:
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    # Dados dos demonstrativos
                    else:
                        cell.font = Font(size=9)
                        # Primeira coluna centralizada, demais à direita
                        if col_idx == 0:
                            cell.alignment = Alignment(horizontal='center')
                        else:
                            cell.alignment = Alignment(horizontal='right')
    
    # ===== AJUSTAR LARGURAS DAS COLUNAS (começando na coluna B) =====
    column_widths = [12, 8, 12, 12, 12, 12, 8, 12, 8, 12, 8, 15, 8, 15, 8, 15, 15, 20]
    # Coluna A fica com largura padrão (vazia)
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 2)].width = width  # ← MUDOU: +2 para começar na coluna B
    
    # ===== CONFIGURAÇÕES DE PÁGINA (igual ao PDF) =====
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Permitir múltiplas páginas na altura
    
    # Margens (igual ao PDF)
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 1.5
    ws.page_margins.bottom = 0.3
    
    # Repetir cabeçalho em todas as páginas (ajustado para coluna B)
    ws.print_title_rows = f'1:{start_row + 1}'  # Repete do cabeçalho até o cabeçalho da tabela
    
    # ===== REMOVER LINHAS DE GRADE =====
    ws.sheet_view.showGridLines = False
    
    # Salvar o arquivo
    wb.save(filename)
    print(f"  ✅ Arquivo XLSX criado: {filename}")
    print(f"      📋 UMA ÚNICA ABA com todo o conteúdo")
    print(f"      📄 Formato paisagem A4 (quebra automática na impressão)")
    print(f"      🎨 SEM linhas de grade e SEM fundo nas células de dados")
    print(f"      📝 Títulos com fonte 9pt (corrigido)")
    print(f"      🔤 Coluna A vazia, conteúdo inicia na coluna B")
    print(f"      ✅ Cabeçalho simplificado e alinhado corretamente")

def format_currency_or_blank(value):
    """Formata valor monetário, retornando string vazia se zero."""
    if value is None or float(value) == 0: 
        return ""
    return format_currency(value)

def format_currency(value):
    if value is None: return ""
    return locale.format_string('%.2f', Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), grouping=True)

def format_date(date_str):
    if not date_str or str(date_str).startswith("1900"): return ""
    try: return datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y') if isinstance(date_str, str) else date_str.strftime('%d/%m/%Y')
    except (ValueError, TypeError): return ""

def format_codificacao_fiscal(cfop):
    s_cfop = str(cfop)
    return s_cfop if len(s_cfop) < 4 else f"{s_cfop[0]}-{s_cfop[1:]}"

def draw_page_header(canvas, doc):
    canvas.saveState()
    empresa = HEADER_INFO.get('empresa_nome', 'EMPRESA PADRAO')
    insc_est = HEADER_INFO.get('ie', '')
    # Se IE estiver vazio, não mostrar N/A, deixar em branco mesmo
    if not insc_est or insc_est == 'N/A':
        insc_est = ''
    
    cnpj_raw = str(HEADER_INFO.get('cnpj', 'N/A'))
    cnpj = f"{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}/{cnpj_raw[8:12]}-{cnpj_raw[12:]}" if len(cnpj_raw) == 14 else cnpj_raw
    periodo = f"{format_date(HEADER_INFO.get('periodo_inicio'))} a {format_date(HEADER_INFO.get('periodo_fim'))}"
    page_num = str(canvas.getPageNumber()).zfill(3)

    canvas.setFont('Helvetica-Bold', 11)
    canvas.drawString(1*cm, 20.5*cm, "LIVRO REGISTRO DE ENTRADAS - RE - MODELO P1")
    canvas.line(1*cm, 20.3*cm, 28.7*cm, 20.3*cm)

    canvas.setFont('Helvetica-Bold', 9)
    canvas.drawString(1*cm, 19.8*cm, "REGISTRO DE ENTRADAS")
    canvas.setFont('Helvetica', 8)
    canvas.drawString(1*cm, 19.4*cm, f"EMPRESA: {empresa}")
    canvas.drawString(1*cm, 19.0*cm, f"INSC.EST.: {insc_est}")
    canvas.drawString(1*cm, 18.6*cm, f"CNPJ: {cnpj}")
    canvas.drawString(1*cm, 18.2*cm, f"FOLHA: {page_num}")
    canvas.drawString(6*cm, 18.2*cm, f"MÊS OU PERÍODO/ANO: {periodo}")

    canvas.setFont('Helvetica-Bold', 9)
    canvas.drawString(17*cm, 19.8*cm, "(*) CÓDIGO DE VALORES FISCAIS")
    canvas.setFont('Helvetica', 7)
    canvas.drawString(17*cm, 19.4*cm, "1-OPERAÇÕES COM CRÉDITO DO IMPOSTO")
    canvas.drawString(17*cm, 19.1*cm, "2-OPER. SEM CRÉDITO DO IMPOSTO - ISENTAS/NÃO TRIBUTADAS")
    canvas.drawString(17*cm, 18.8*cm, "3-OPERAÇÕES SEM CRÉDITO DO IMPOSTO - OUTRAS")
    canvas.drawString(17*cm, 18.5*cm, "4-ICMS RETIDO POR SUBST.TRIBUTÁRIA")
    
    canvas.line(1*cm, 17.8*cm, 28.7*cm, 17.8*cm)
    canvas.restoreState()

def determinar_codigo_fiscal_e_valores(detalhe):
    """
    Determina o código fiscal e os valores correspondentes baseado nos campos:
    1 = BASE > 0 (operações com crédito do imposto) → usa BASE + VALOR
    2 = ISENTAS > 0 (operações sem crédito - isentas/não tributadas) → usa ISENTAS
    3 = OUTRAS > 0 (operações sem crédito - outras) → usa OUTRAS
    
    Retorna: (codigo, valor_para_base, mostrar_aliquota, valor_imposto_creditado)
    """
    try:
        base = float(detalhe.get("BASE", 0))
        isentas = float(detalhe.get("ISENTAS", 0))
        outras = float(detalhe.get("OUTRAS", 0))
        valor_imposto = float(detalhe.get("VALOR", 0))
        
        if base > 0:
            return "1", base, True, valor_imposto  # Código 1 = mostra alíquota + imposto creditado
        elif isentas > 0:
            return "2", isentas, False, 0  # Código 2 = sem alíquota, sem imposto creditado
        elif outras > 0:
            return "3", outras, False, 0   # Código 3 = sem alíquota, sem imposto creditado
        else:
            return "", 0, False, 0
    except (ValueError, TypeError):
        return "", 0, False, 0

def get_imposto_para_cabecalho(detalhes, tipo_imposto):
    """
    NOVA LÓGICA: 
    - Se tem apenas 2 registros do mesmo tipo: pega o PRIMEIRO (sequência padrão)
    - Se tem 3 ou mais registros do mesmo tipo: pega o ÚLTIMO (sequência invertida)
    """
    registros_do_tipo = [d for d in detalhes if d.get('IMPOSTO') == tipo_imposto]
    if not registros_do_tipo:
        return None
    
    # Se tem apenas 2 registros: sequência padrão (pega o primeiro/menor sequencial)
    if len(registros_do_tipo) == 2:
        return min(registros_do_tipo, key=lambda x: x.get('SEQUENCIAL', 0))
    
    # Se tem 3 ou mais registros: "último primeiro" (pega o maior sequencial)
    return max(registros_do_tipo, key=lambda x: x.get('SEQUENCIAL', 0))

def create_main_table(transactions, tax_details, impostos_raw):
    """Cria a tabela principal usando a nova lógica corrigida: ÚLTIMO SEQUENCIAL para o cabeçalho."""
    data = [
        ["DATA DE\nENTRADA", "DOCUMENTOS FISCAIS", "", "", "", "", "", "VALOR\nCONTÁBIL", "CODIFICAÇÃO", "", "ICMS VALORES\nFISCAIS", "", "", "", "IPI VALORES\nFISCAIS", "", "" ,"OBSERVAÇÕES"],
        ["", "ESPÉCIE", "SÉRIE/\nSUB-SÉRIE", "NÚMERO", "DATA DO\nDOCUMENTO", "CÓD. DO\nEMITENTE", "UF\nORIGEM", "", "CONTÁBIL", "FISCAL", "COD (*)", "BASE CÁLCULO\nVALOR OPERAÇÃO", "ALÍQ.\n%", "IMPOSTO\nCREDITADO", "COD (*)", "BASE DE CÁLCULO\nVALOR DA OPERAÇÃO", "IMPOSTO\nCREDITADO", ""]
    ]
    
    resultado_final = []
    linha_vazia_modelo = {
        "DATA": "", "ESPECIE": "", "SERIE / SUB-SERIE": "", "NUMERO": "",
        "DATA DO DOCUMENTO": "", "COD DO EMITENTE": "", "UF ORIGEM": "", "VCON": "",
        "CODIGO CONTABIL": "", "CODIGO FISCAL": "",
        "ICMS_COD(*)": "", "ICMS_BASE_CALCULO_VALOR_OPERACAO": "", "ICMS_ALIQUOTA": "", "ICMS_IMPOSTO_CREDITADO": "",
        "IPI_COD(*)": "", "IPI_BASE_CALCULO_VALOR_OPERACAO": "", "IPI_IMPOSTO_CREDITADO": "",
        "OBSERVACOES": ""
    }
    
    for doc in transactions:
        segmentos = doc['items']
        
        # Ordena segmentos por CFOP para garantir ordem consistente
        segmentos_ordenados = sorted(segmentos, key=lambda x: x.get('cfop', 0))
        primeiro_segmento = segmentos_ordenados[0]
        
        vcon_total = sum(float(seg.get('valor_operacao', 0.0)) for seg in segmentos)
        
        # Obter CFOP para decidir se agrupa ou não
        cfop_primeiro_segmento = primeiro_segmento.get('cfop', 0)
        primeiro_digito_cfop = str(cfop_primeiro_segmento)[0] if cfop_primeiro_segmento else '0'
        deve_agrupar = primeiro_digito_cfop == '1'  # Agrupa apenas se CFOP começar com 1 (operações estaduais)
        
        # Obter e ordenar todos os detalhes do primeiro segmento por sequencial
        detalhes_primeiro_segmento = sorted(impostos_raw.get(str(primeiro_segmento['codigo']), []), key=lambda x: x.get('SEQUENCIAL', 0))
        
        if deve_agrupar:
            # LÓGICA COM AGRUPAMENTO (CFOP iniciado com '1' - operações estaduais)
            # Agrupar e somar impostos por tipo para o primeiro segmento
            impostos_agrupados = {}
            for detalhe in detalhes_primeiro_segmento:
                imposto_tipo = detalhe.get('IMPOSTO')
                if imposto_tipo not in impostos_agrupados:
                    impostos_agrupados[imposto_tipo] = {
                        'BASE_TOTAL': 0,
                        'ISENTAS_TOTAL': 0, 
                        'OUTRAS_TOTAL': 0,
                        'VALOR_TOTAL': 0,  # ← NOVO CAMPO PARA SOMAR VALORES
                        'ALIQUOTA': detalhe.get('ALIQUOTA', ''),
                        'REGISTROS': []
                    }
                
                try:
                    impostos_agrupados[imposto_tipo]['BASE_TOTAL'] += float(detalhe.get('BASE', 0))
                    impostos_agrupados[imposto_tipo]['ISENTAS_TOTAL'] += float(detalhe.get('ISENTAS', 0))
                    impostos_agrupados[imposto_tipo]['OUTRAS_TOTAL'] += float(detalhe.get('OUTRAS', 0))
                    impostos_agrupados[imposto_tipo]['VALOR_TOTAL'] += float(detalhe.get('VALOR', 0))  # ← SOMA VALORES
                    impostos_agrupados[imposto_tipo]['REGISTROS'].append(detalhe)
                except (ValueError, TypeError):
                    pass
            
            # Obter impostos principais agrupados
            primeiro_icms_agrupado = impostos_agrupados.get(1)
            primeiro_ipi_agrupado = impostos_agrupados.get(2)
            
            # Determinar códigos e valores ICMS com verificação de zero (usando valores agrupados)
            icms_cod = ""
            icms_base = ""
            icms_aliq = ""
            icms_imposto = ""
            if primeiro_icms_agrupado:
                # Simula um detalhe agrupado para usar a função de determinação
                detalhe_simulado = {
                    'BASE': primeiro_icms_agrupado['BASE_TOTAL'],
                    'ISENTAS': primeiro_icms_agrupado['ISENTAS_TOTAL'],
                    'OUTRAS': primeiro_icms_agrupado['OUTRAS_TOTAL'],
                    'VALOR': primeiro_icms_agrupado['VALOR_TOTAL']  # ← VALOR TOTAL AGRUPADO
                }
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                
                if codigo and valor_base > 0:
                    icms_cod = codigo
                    icms_base = format_currency(valor_base)
                    if mostrar_aliquota:
                        icms_aliq = format_currency(primeiro_icms_agrupado['ALIQUOTA']) if primeiro_icms_agrupado['ALIQUOTA'] else ""
                    if codigo == "1":  # ← SÓ PREENCHE IMPOSTO CREDITADO SE CÓDIGO = "1"
                        icms_imposto = format_currency(valor_imposto_creditado)
            
            # Determinar códigos e valores IPI com verificação de zero (usando valores agrupados)
            ipi_cod = ""
            ipi_base = ""
            ipi_imposto = ""
            if primeiro_ipi_agrupado:
                # Simula um detalhe agrupado para usar a função de determinação
                detalhe_simulado = {
                    'BASE': primeiro_ipi_agrupado['BASE_TOTAL'],
                    'ISENTAS': primeiro_ipi_agrupado['ISENTAS_TOTAL'],
                    'OUTRAS': primeiro_ipi_agrupado['OUTRAS_TOTAL'],
                    'VALOR': primeiro_ipi_agrupado['VALOR_TOTAL']  # ← VALOR TOTAL AGRUPADO
                }
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                
                if codigo and valor_base > 0:
                    ipi_cod = codigo
                    ipi_base = format_currency(valor_base)
                    if codigo == "1":  # ← SÓ PREENCHE IMPOSTO CREDITADO SE CÓDIGO = "1"
                        ipi_imposto = format_currency(valor_imposto_creditado)
                    
            # Marca os detalhes exatos que foram usados no cabeçalho (TODOS os registros dos tipos usados)
            detalhes_usados = set()
            if primeiro_icms_agrupado and icms_cod:
                for registro in primeiro_icms_agrupado['REGISTROS']:
                    detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
            if primeiro_ipi_agrupado and ipi_cod:
                for registro in primeiro_ipi_agrupado['REGISTROS']:
                    detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
        else:
            # LÓGICA SEM AGRUPAMENTO (CFOP iniciado com '2', '3', etc. - operações interestaduais/exterior)
            # *** MUDANÇA AQUI: Lógica especial baseada na quantidade de registros ***
            icms_para_cabecalho = get_imposto_para_cabecalho(detalhes_primeiro_segmento, 1)
            ipi_para_cabecalho = get_imposto_para_cabecalho(detalhes_primeiro_segmento, 2)
            
            # Determinar códigos e valores ICMS com verificação de zero
            icms_cod = ""
            icms_base = ""
            icms_aliq = ""
            icms_imposto = ""
            if icms_para_cabecalho:
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(icms_para_cabecalho)
                if codigo and valor_base > 0:
                    icms_cod = codigo
                    icms_base = format_currency(valor_base)
                    if mostrar_aliquota:
                        icms_aliq = format_currency(icms_para_cabecalho.get("ALIQUOTA", "")) if icms_para_cabecalho.get("ALIQUOTA") else ""
                    if codigo == "1":  # ← SÓ PREENCHE IMPOSTO CREDITADO SE CÓDIGO = "1"
                        icms_imposto = format_currency(valor_imposto_creditado)
            
            # Determinar códigos e valores IPI com verificação de zero
            ipi_cod = ""
            ipi_base = ""
            ipi_imposto = ""
            if ipi_para_cabecalho:
                codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(ipi_para_cabecalho)
                if codigo and valor_base > 0:
                    ipi_cod = codigo
                    ipi_base = format_currency(valor_base)
                    if codigo == "1":  # ← SÓ PREENCHE IMPOSTO CREDITADO SE CÓDIGO = "1"
                        ipi_imposto = format_currency(valor_imposto_creditado)
                    
            # Marca os detalhes exatos que foram usados no cabeçalho
            detalhes_usados = set()
            if icms_para_cabecalho and icms_cod: 
                detalhes_usados.add((icms_para_cabecalho.get("NOTA"), icms_para_cabecalho.get("SEQUENCIAL")))
            if ipi_para_cabecalho and ipi_cod: 
                detalhes_usados.add((ipi_para_cabecalho.get("NOTA"), ipi_para_cabecalho.get("SEQUENCIAL")))

        # Verificar se existe DIFALI para adicionar nas observações
        # REGRA: Só adiciona observação se NÃO tiver ST (Substituição Tributária)
        observacoes = ""
        tem_subtri = False
        
        # Primeiro, verifica se tem SUBTRI em qualquer segmento
        for segmento in segmentos:
            detalhes_segmento = impostos_raw.get(str(segmento['codigo']), [])
            for detalhe in detalhes_segmento:
                if detalhe.get('SIGLA', '').strip() == 'SUBTRI':
                    try:
                        valor_subtri = float(detalhe.get('VALOR', 0))
                        if valor_subtri > 0:
                            tem_subtri = True
                            break
                    except (ValueError, TypeError):
                        pass
            if tem_subtri:
                break
        
        # Só busca DIFALI se NÃO tiver SUBTRI
        if not tem_subtri:
            for segmento in segmentos:
                detalhes_segmento = impostos_raw.get(str(segmento['codigo']), [])
                for detalhe in detalhes_segmento:
                    if detalhe.get('SIGLA', '').strip() == 'DIFALI' and detalhe.get('VALOR'):
                        try:
                            valor_difali = float(detalhe.get('VALOR', 0))
                            if valor_difali > 0:
                                observacoes = f"Diferencial de\nAlíquota: {format_currency(valor_difali)}"
                                break  # Para no primeiro DIFALI > 0 encontrado
                        except (ValueError, TypeError):
                            pass
                if observacoes:  # Se já encontrou DIFALI, para de procurar
                    break

        # Criar linha de cabeçalho (comum para ambas as lógicas)
        serie_sub = doc.get('serie', '')
        if doc.get('sub_serie') and doc['sub_serie'] != '0': 
            serie_sub += f"/{doc['sub_serie']}"
        
        linha_cabecalho = {
            "DATA": format_date(doc['data_entrada']), "ESPECIE": doc['especie'],
            "SERIE / SUB-SERIE": serie_sub, "NUMERO": doc['numero'],
            "DATA DO DOCUMENTO": format_date(doc['data_doc']), "COD DO EMITENTE": doc['cod_emitente'],
            "UF ORIGEM": doc['uf'], "VCON": f"{vcon_total:.2f}",
            "CODIGO CONTABIL": "", "CODIGO FISCAL": format_codificacao_fiscal(primeiro_segmento.get('cfop')),
            "ICMS_COD(*)": icms_cod, 
            "ICMS_BASE_CALCULO_VALOR_OPERACAO": icms_base,
            "ICMS_ALIQUOTA": icms_aliq, 
            "ICMS_IMPOSTO_CREDITADO": icms_imposto,  # ← CAMPO AJUSTADO
            "IPI_COD(*)": ipi_cod, 
            "IPI_BASE_CALCULO_VALOR_OPERACAO": ipi_base,
            "IPI_IMPOSTO_CREDITADO": ipi_imposto,  # ← CAMPO AJUSTADO
            "OBSERVACOES": observacoes
        }
        resultado_final.append(linha_cabecalho)

        # --- Loop para gerar todas as linhas de detalhe ---
        for i, segmento_atual in enumerate(segmentos_ordenados):
            detalhes_atuais_ordenados = sorted(impostos_raw.get(str(segmento_atual['codigo']), []), key=lambda x: x.get('SEQUENCIAL', 0))
            
            # Se for um segmento adicional, gera a linha principal dele primeiro
            if i > 0:
                # Usa a mesma lógica de agrupamento baseada no CFOP do segmento atual
                cfop_segmento_atual = segmento_atual.get('cfop', 0)
                primeiro_digito_cfop_seg = str(cfop_segmento_atual)[0] if cfop_segmento_atual else '0'
                deve_agrupar_seg = primeiro_digito_cfop_seg == '1'
                
                if deve_agrupar_seg:
                
                    # LÓGICA COM AGRUPAMENTO PARA SEGMENTOS (CFOP iniciado com '1')
                    impostos_agrupados_seg = {}
                    for detalhe in detalhes_atuais_ordenados:
                        imposto_tipo = detalhe.get('IMPOSTO')
                        if imposto_tipo not in impostos_agrupados_seg:
                            impostos_agrupados_seg[imposto_tipo] = {
                                'BASE_TOTAL': 0,
                                'ISENTAS_TOTAL': 0,
                                'OUTRAS_TOTAL': 0,
                                'VALOR_TOTAL': 0,  # ← NOVO CAMPO
                                'ALIQUOTA': detalhe.get('ALIQUOTA', ''),
                                'REGISTROS': []
                            }
                        
                        try:
                            impostos_agrupados_seg[imposto_tipo]['BASE_TOTAL'] += float(detalhe.get('BASE', 0))
                            impostos_agrupados_seg[imposto_tipo]['ISENTAS_TOTAL'] += float(detalhe.get('ISENTAS', 0))
                            impostos_agrupados_seg[imposto_tipo]['OUTRAS_TOTAL'] += float(detalhe.get('OUTRAS', 0))
                            impostos_agrupados_seg[imposto_tipo]['VALOR_TOTAL'] += float(detalhe.get('VALOR', 0))  # ← SOMA VALORES
                            impostos_agrupados_seg[imposto_tipo]['REGISTROS'].append(detalhe)
                        except (ValueError, TypeError):
                            pass
                    
                    icms_principal_seg_agrupado = impostos_agrupados_seg.get(1)
                    ipi_principal_seg_agrupado = impostos_agrupados_seg.get(2)
                    
                    # Determinar códigos e valores ICMS para segmento (usando valores agrupados)
                    icms_cod_seg = ""
                    icms_base_seg = ""
                    icms_aliq_seg = ""
                    icms_imposto_seg = ""
                    if icms_principal_seg_agrupado:
                        detalhe_simulado = {
                            'BASE': icms_principal_seg_agrupado['BASE_TOTAL'],
                            'ISENTAS': icms_principal_seg_agrupado['ISENTAS_TOTAL'],
                            'OUTRAS': icms_principal_seg_agrupado['OUTRAS_TOTAL'],
                            'VALOR': icms_principal_seg_agrupado['VALOR_TOTAL']  # ← VALOR TOTAL
                        }
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                        
                        if codigo and valor_base > 0:
                            icms_cod_seg = codigo
                            icms_base_seg = format_currency(valor_base)
                            if mostrar_aliquota:
                                icms_aliq_seg = format_currency(icms_principal_seg_agrupado['ALIQUOTA']) if icms_principal_seg_agrupado['ALIQUOTA'] else ""
                            if codigo == "1":  # ← SÓ PREENCHE SE CÓDIGO = "1"
                                icms_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Determinar códigos e valores IPI para segmento (usando valores agrupados)
                    ipi_cod_seg = ""
                    ipi_base_seg = ""
                    ipi_imposto_seg = ""
                    if ipi_principal_seg_agrupado:
                        detalhe_simulado = {
                            'BASE': ipi_principal_seg_agrupado['BASE_TOTAL'],
                            'ISENTAS': ipi_principal_seg_agrupado['ISENTAS_TOTAL'],
                            'OUTRAS': ipi_principal_seg_agrupado['OUTRAS_TOTAL'],
                            'VALOR': ipi_principal_seg_agrupado['VALOR_TOTAL']  # ← VALOR TOTAL
                        }
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe_simulado)
                        
                        if codigo and valor_base > 0:
                            ipi_cod_seg = codigo
                            ipi_base_seg = format_currency(valor_base)
                            if codigo == "1":  # ← SÓ PREENCHE SE CÓDIGO = "1"
                                ipi_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Marca os detalhes que acabaram de ser usados para este segmento adicional (TODOS os registros dos tipos usados)
                    if icms_principal_seg_agrupado and icms_cod_seg:
                        for registro in icms_principal_seg_agrupado['REGISTROS']:
                            detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
                    if ipi_principal_seg_agrupado and ipi_cod_seg:
                        for registro in ipi_principal_seg_agrupado['REGISTROS']:
                            detalhes_usados.add((registro.get("NOTA"), registro.get("SEQUENCIAL")))
                else:
                    # LÓGICA SEM AGRUPAMENTO PARA SEGMENTOS (CFOP iniciado com '2', '3', etc.)
                    icms_seg_para_cabecalho = get_imposto_para_cabecalho(detalhes_atuais_ordenados, 1)
                    ipi_seg_para_cabecalho = get_imposto_para_cabecalho(detalhes_atuais_ordenados, 2)
                    
                    # Determinar códigos e valores ICMS para segmento
                    icms_cod_seg = ""
                    icms_base_seg = ""
                    icms_aliq_seg = ""
                    icms_imposto_seg = ""
                    if icms_seg_para_cabecalho:
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(icms_seg_para_cabecalho)
                        if codigo and valor_base > 0:
                            icms_cod_seg = codigo
                            icms_base_seg = format_currency(valor_base)
                            if mostrar_aliquota:
                                icms_aliq_seg = format_currency(icms_seg_para_cabecalho.get("ALIQUOTA", "")) if icms_seg_para_cabecalho.get("ALIQUOTA") else ""
                            if codigo == "1":  # ← SÓ PREENCHE SE CÓDIGO = "1"
                                icms_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Determinar códigos e valores IPI para segmento
                    ipi_cod_seg = ""
                    ipi_base_seg = ""
                    ipi_imposto_seg = ""
                    if ipi_seg_para_cabecalho:
                        codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(ipi_seg_para_cabecalho)
                        if codigo and valor_base > 0:
                            ipi_cod_seg = codigo
                            ipi_base_seg = format_currency(valor_base)
                            if codigo == "1":  # ← SÓ PREENCHE SE CÓDIGO = "1"
                                ipi_imposto_seg = format_currency(valor_imposto_creditado)
                    
                    # Marca os detalhes que acabaram de ser usados para este segmento adicional
                    if icms_seg_para_cabecalho and icms_cod_seg: 
                        detalhes_usados.add((icms_seg_para_cabecalho.get("NOTA"), icms_seg_para_cabecalho.get("SEQUENCIAL")))
                    if ipi_seg_para_cabecalho and ipi_cod_seg: 
                        detalhes_usados.add((ipi_seg_para_cabecalho.get("NOTA"), ipi_seg_para_cabecalho.get("SEQUENCIAL")))
                
                # Criar linha de segmento (comum para ambas as lógicas)
                linha_segmento = { **linha_vazia_modelo, **{
                    "CODIGO FISCAL": format_codificacao_fiscal(segmento_atual.get('cfop')),
                    "ICMS_COD(*)": icms_cod_seg, 
                    "ICMS_BASE_CALCULO_VALOR_OPERACAO": icms_base_seg, 
                    "ICMS_ALIQUOTA": icms_aliq_seg,
                    "ICMS_IMPOSTO_CREDITADO": icms_imposto_seg,  # ← CAMPO AJUSTADO
                    "IPI_COD(*)": ipi_cod_seg, 
                    "IPI_BASE_CALCULO_VALOR_OPERACAO": ipi_base_seg, 
                    "IPI_IMPOSTO_CREDITADO": ipi_imposto_seg,  # ← CAMPO AJUSTADO
                }}
                resultado_final.append(linha_segmento)

            # Agora itera sobre todos os detalhes do segmento atual para gerar as linhas de detalhe esparsas
            # NOVA LÓGICA: Ordem baseada na quantidade de registros por tipo
            def get_ordem_detalhes(codigo_segmento, impostos_raw):
                """
                Determina a ordem dos detalhes baseada na quantidade de registros de cada tipo:
                - Se tipo tem apenas 2 registros: ordem CRESCENTE (padrão)
                - Se tipo tem 3+ registros: ordem DECRESCENTE ("último primeiro")
                """
                detalhes_segmento = impostos_raw.get(str(codigo_segmento), [])
                
                # Conta registros por tipo de imposto
                contadores = {}
                for detalhe in detalhes_segmento:
                    tipo = detalhe.get('IMPOSTO')
                    contadores[tipo] = contadores.get(tipo, 0) + 1
                
                # Ordena cada tipo conforme sua quantidade
                detalhes_ordenados = []
                
                # Processa ICMS (tipo 1) e IPI (tipo 2) primeiro
                for tipo in [1, 2]:
                    registros_tipo = [d for d in detalhes_segmento if d.get('IMPOSTO') == tipo]
                    if not registros_tipo:
                        continue
                    
                    qtd_registros = contadores.get(tipo, 0)
                    if qtd_registros == 2:
                        # Apenas 2 registros: ordem CRESCENTE (padrão)
                        registros_ordenados = sorted(registros_tipo, key=lambda x: x.get('SEQUENCIAL', 0))
                    else:
                        # 1, 3+ registros: ordem DECRESCENTE ("último primeiro") 
                        registros_ordenados = sorted(registros_tipo, key=lambda x: x.get('SEQUENCIAL', 0), reverse=True)
                    
                    detalhes_ordenados.extend(registros_ordenados)
                
                # Adiciona outros tipos (9, etc.) sempre em ordem decrescente
                outros_tipos = [d for d in detalhes_segmento if d.get('IMPOSTO') not in [1, 2]]
                outros_ordenados = sorted(outros_tipos, key=lambda x: x.get('SEQUENCIAL', 0), reverse=True)
                detalhes_ordenados.extend(outros_ordenados)
                
                return detalhes_ordenados
            
            detalhes_atuais_ordenados_especial = get_ordem_detalhes(segmento_atual['codigo'], impostos_raw)
            
            for detalhe in detalhes_atuais_ordenados_especial:
                chave_detalhe = (detalhe.get("NOTA"), detalhe.get("SEQUENCIAL"))
                if chave_detalhe in detalhes_usados:
                    continue

                sigla = detalhe.get('SIGLA', '').strip().upper()
                imposto_id = detalhe.get('IMPOSTO')
                
                linha_detalhe = linha_vazia_modelo.copy()
                gerou_linha = False
                
                if sigla == 'SUBTRI':
                    try:
                        if float(detalhe.get("VALOR", 0.0)) > 0:
                            linha_detalhe["ICMS_COD(*)"] = "ST"
                            linha_detalhe["ICMS_BASE_CALCULO_VALOR_OPERACAO"] = format_currency(detalhe.get("BASE", ""))
                            linha_detalhe["ICMS_IMPOSTO_CREDITADO"] = format_currency_or_blank(detalhe.get("VALOR", ""))
                            gerou_linha = True
                    except (ValueError, TypeError):
                        pass
                elif imposto_id == 1:  # ICMS
                    codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe)
                    if codigo and valor_base > 0:
                        linha_detalhe["ICMS_COD(*)"] = codigo
                        linha_detalhe["ICMS_BASE_CALCULO_VALOR_OPERACAO"] = format_currency(valor_base)
                        if mostrar_aliquota:
                            linha_detalhe["ICMS_ALIQUOTA"] = format_currency(detalhe.get("ALIQUOTA", "")) if detalhe.get("ALIQUOTA") else ""
                        if codigo == "1":  # ← SÓ PREENCHE IMPOSTO CREDITADO SE CÓDIGO = "1"
                            linha_detalhe["ICMS_IMPOSTO_CREDITADO"] = format_currency(valor_imposto_creditado)
                        gerou_linha = True
                elif imposto_id == 2:  # IPI
                    codigo, valor_base, mostrar_aliquota, valor_imposto_creditado = determinar_codigo_fiscal_e_valores(detalhe)
                    if codigo and valor_base > 0:
                        linha_detalhe["IPI_COD(*)"] = codigo
                        linha_detalhe["IPI_BASE_CALCULO_VALOR_OPERACAO"] = format_currency(valor_base)
                        if codigo == "1":  # ← SÓ PREENCHE IMPOSTO CREDITADO SE CÓDIGO = "1"
                            linha_detalhe["IPI_IMPOSTO_CREDITADO"] = format_currency(valor_imposto_creditado)
                        # IPI nunca mostra alíquota na tabela final
                        gerou_linha = True
                
                if gerou_linha:
                    resultado_final.append(linha_detalhe)
                    detalhes_usados.add(chave_detalhe)

    # Converter resultado_final para o formato da tabela
    for item in resultado_final:
        row = [
            item.get("DATA", ""),
            item.get("ESPECIE", ""),
            item.get("SERIE / SUB-SERIE", ""),
            item.get("NUMERO", ""),
            item.get("DATA DO DOCUMENTO", ""),
            item.get("COD DO EMITENTE", ""),
            item.get("UF ORIGEM", ""),
            item.get("VCON", ""),
            item.get("CODIGO CONTABIL", ""),
            item.get("CODIGO FISCAL", ""),
            item.get("ICMS_COD(*)", ""),
            item.get("ICMS_BASE_CALCULO_VALOR_OPERACAO", ""),
            item.get("ICMS_ALIQUOTA", ""),
            item.get("ICMS_IMPOSTO_CREDITADO", ""),
            item.get("IPI_COD(*)", ""),
            item.get("IPI_BASE_CALCULO_VALOR_OPERACAO", ""),
            item.get("IPI_IMPOSTO_CREDITADO", ""),
            item.get("OBSERVACOES", "")
        ]
        data.append(row)

    # Criar e configurar a tabela
    col_widths = [45,30,30,40,50,40,30,45,30,30,30,55,22,50,22,55,50,75]  # Aumentada a coluna de observações
       # --- LARGURA DAS COLUNAS (em pontos. 1cm ≈ 28.35 pontos) ---
    # A ordem das larguras corresponde exatamente à ordem das colunas na tabela.
    # Índice | Largura | Coluna
    # --------------------------------------------------------------------------
    # 0      | 45      | DATA DE ENTRADA
    # 1      | 30      | ESPÉCIE
    # 2      | 30      | SÉRIE/SUB-SÉRIE
    # 3      | 30      | NÚMERO
    # 4      | 50      | DATA DO DOCUMENTO
    # 5      | 40      | CÓD. DO EMITENTE
    # 6      | 30      | UF ORIGEM
    # 7      | 45      | VALOR CONTÁBIL
    # 8      | 30      | CODIFICAÇÃO - CONTÁBIL
    # 9      | 30      | CODIFICAÇÃO - FISCAL
    # 10     | 30      | ICMS - COD (*)
    # 11     | 55      | ICMS - BASE CÁLCULO
    # 12     | 22      | ICMS - ALÍQ. %
    # 13     | 50      | ICMS - IMPOSTO CREDITADO
    # 14     | 22      | IPI - COD (*)
    # 15     | 55      | IPI - BASE DE CÁLCULO
    # 16     | 50      | IPI - IMPOSTO CREDITADO
    # 17     | 75      | OBSERVAÇÕES
    table = Table(data, colWidths=col_widths, repeatRows=2)
    style = TableStyle([
        ('GRID', (0,0), (-1,-1), 0.3, colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 6), ('LEADING', (0,0), (-1,0), 7),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica'), ('FONTSIZE', (0,1), (-1,1), 5), ('LEADING', (0,1), (-1,1), 6),
        ('FONTNAME', (0,2), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,2), (-1,-1), 6),
        ('SPAN', (1,0), (6,0)), ('SPAN', (8,0), (9,0)), ('SPAN', (10,0), (13,0)), ('SPAN', (14,0), (16,0)),
        ('SPAN', (0,0), (0,1)), ('SPAN', (7,0), (7,1)), ('SPAN', (17,0), (17,1)),
        ('ALIGN', (7,2), (7,-1), 'RIGHT'), ('ALIGN', (11,2), (13,-1), 'RIGHT'),
        ('ALIGN', (15,2), (16,-1), 'RIGHT'), ('ALIGN', (17,2), (17,-1), 'LEFT'),
        ('ALIGN', (17,2), (17,-1), 'LEFT'),  # Observações: alinhamento à esquerda, fonte normal
        ('VALIGN', (17,2), (17,-1), 'TOP'),  # Alinhamento vertical superior para observações
        ('WORDWRAP', (17,2), (17,-1), True),  # Quebra automática de linha nas observações
    ])
    table.setStyle(style)
    table.hAlign = 'CENTER'
    return table

def group_transactions_by_document(transactions):
    """Agrupa múltiplos lançamentos que pertencem à mesma nota fiscal."""
    grouped = defaultdict(lambda: {
        'items': [],
        'valor_contabil_total': Decimal(0)
    })

    for t in transactions:
        # A chave única para agrupar a nota fiscal
        doc_key = f"{t['numero']}|{t['cod_emitente']}|{t['serie']}|{t['data_doc']}"

        # Se for o primeiro lançamento desta nota, copia todos os dados do cabeçalho
        if not grouped[doc_key]['items']:
            # Copia todos os campos do lançamento original para o registro agrupado
            for key, value in t.items():
                if key not in ['items', 'valor_contabil', 'cfop']: # Evita copiar campos específicos do item
                    grouped[doc_key][key] = value

        # Adiciona o item (lançamento) com seu CFOP e valor específico
        grouped[doc_key]['items'].append({
            'cfop': t['cfop'],
            'valor_operacao': t['valor_contabil'],
            'codigo': t['codigo'] # Mantém o código original do lançamento para buscar impostos
        })
        # Soma o valor do lançamento ao total da nota
        grouped[doc_key]['valor_contabil_total'] += t['valor_contabil']

    return list(grouped.values())

def process_and_filter_transactions(data):
    """
    Filtra, agrupa e ordena as transações para o relatório final.
    """
    raw_transactions = data.get("transactions", [])
    
    modelos_permitidos = [
    '01', '02',  '04', '05', '06', '07', '08', '09', '10',
    '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
    '21', '22', '23', '24', '25', '26', '27', '28', '29', '30',
    '31', '32', '33', '34', '35', '36', '37', '38', '39', '40',
    '41', '42', '43', '44', '45', '46', '47', '48', '49', '50',
    '51', '52', '53', '54', '55', '56', '57', '58', '59', '60',
    '61', '62', '63', '64', '65', '66', '67', '68', '69', '70',
    '71', '72', '73', '74', '75', '76', '77', '78', '79', '80',
    '81', '82', '83', '84', '85', '86', '87', '88', '89', '90',
    '91', '93', '94', '95', '96', '97', '98', '99'
    ]
    
    # Filtro: apenas modelos permitidos
    filtered_transactions = [
        t for t in raw_transactions if str(t.get('modelo')).strip() in modelos_permitidos
    ]
    print(f"  - Registros brutos: {len(raw_transactions)}. Após filtro (modelos): {len(filtered_transactions)}.")

    grouped_transactions = group_transactions_by_document(filtered_transactions)
    print(f"  - Transações após agrupamento: {len(grouped_transactions)}.")

    default_date = datetime(1900, 1, 1)
    
    sorted_transactions = sorted(grouped_transactions, key=lambda t: (
        t.get('data_entrada') or default_date,
        t.get('data_doc') or default_date,
        t.get('nome_fornecedor') or '',
        int(t.get('numero') or 0),
    ))
    
    print("  - Transações filtradas, agrupadas e ordenadas com sucesso.")
    return sorted_transactions

def gerarLivroEntrada(codi_emp, data_inicio, data_fim, gerar_pdf=True, gerar_xlsx=False):
    """
    Função principal para gerar o Livro Registro de Entradas - Modelo P1
    
    Args:
        codi_emp (int): Código da empresa
        data_inicio (str): Data de início no formato 'YYYY-MM-DD'
        data_fim (str): Data de fim no formato 'YYYY-MM-DD'
        gerar_pdf (bool): Se deve gerar arquivo PDF
        gerar_xlsx (bool): Se deve gerar arquivo XLSX
    
    Returns:
        list: Uma lista contendo os nomes dos arquivos gerados com sucesso.
              Retorna uma lista vazia em caso de falha.
    """
    # Lista para armazenar os nomes dos arquivos gerados
    arquivos_gerados = []

    output_filename_pdf = f"LivroEntrada{codi_emp}.pdf"
    output_filename_xlsx = f"LivroEntrada{codi_emp}.xlsx"
    json_filename = f"Dados_Extraidos{codi_emp}.json"
    
    print("="*60)
    print("INICIANDO GERAÇÃO DO LIVRO REGISTRO DE ENTRADAS - MODELO P1")
    print("="*60)
    print(f"Empresa: {codi_emp}")
    print(f"Período: {data_inicio} até {data_fim}")
    if gerar_pdf:
        print(f"Arquivo PDF: {output_filename_pdf}")
    if gerar_xlsx:
        print(f"Arquivo XLSX: {output_filename_xlsx}")
    print(f"Arquivo JSON: {json_filename}")
    print()
    
    print("ETAPA 1: Extraindo dados brutos do banco de dados...")
    data, raw_data = extract_data_from_db(codi_emp, data_inicio, data_fim)
    
    if not data:
        print("❌ ERRO: Falha ao extrair dados do banco de dados.")
        return []

    print("\nETAPA 1.5: Salvando dados brutos do banco em JSON...")
    try:
        final_raw_data = {"header": data["header"], **raw_data}
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(final_raw_data, f, indent=4, default=str, ensure_ascii=False)
        print(f"  ✅ JSON com dados brutos salvo: '{json_filename}'")
        arquivos_gerados.append(json_filename) # Adiciona o JSON à lista de arquivos
    except Exception as ex:
        print(f"  ⚠️  ERRO ao salvar JSON: {ex}")
        import traceback
        traceback.print_exc()
    
    if not data.get("transactions"):
        print("⚠️  AVISO: Nenhuma transação encontrada para o período informado.")
        # Retorna a lista com o JSON que pode ter sido gerado, ou vazia se não foi.
        return arquivos_gerados

    print("\nETAPA 2: Processando (filtrando, agrupando e ordenando) transações em Python...")
    final_transactions = process_and_filter_transactions(data)
    
    if not final_transactions:
        print("⚠️  AVISO: Nenhuma transação válida para o relatório após a filtragem e agrupamento.")
        return arquivos_gerados

    print("\nETAPA 2.5: Calculando resumos a partir dos dados finais...")
    summaries = calculate_detailed_summaries(final_transactions, data.get("tax_details", {}))

    # Gerar PDF se solicitado
    if gerar_pdf:
        print("\nETAPA 3: Gerando relatório PDF...")
        try:
            doc = SimpleDocTemplate(output_filename_pdf, pagesize=landscape(A4), leftMargin=8, rightMargin=8, topMargin=4*cm, bottomMargin=8)
            story = []
            
            print("  - Criando tabela principal...")
            main_table = create_main_table(final_transactions, data.get("tax_details", {}), raw_data.get("impostos_raw", {}))
            story.append(main_table)
            
            print("  - Criando demonstrativos...")
            summary_flowables = create_summary_flowables(summaries)
            
            if summary_flowables:
                story.append(PageBreak())
                story.extend(summary_flowables)
            
            print("  - Construindo arquivo PDF...")
            doc.build(story, onFirstPage=draw_page_header, onLaterPages=draw_page_header)
            
            print(f"\n✅ SUCESSO: Relatório PDF gerado!\n📄 Arquivo PDF: '{output_filename_pdf}'")
            arquivos_gerados.append(output_filename_pdf)
        except Exception as ex:
            print(f"❌ ERRO ao gerar PDF: {ex}")
            import traceback
            traceback.print_exc()
            # Não retorna aqui, para tentar gerar o XLSX mesmo se o PDF falhar

    # Gerar XLSX se solicitado
    if gerar_xlsx:
        print("\nETAPA 4: Gerando relatório XLSX...")
        try:
            print("  - Preparando dados da tabela principal para XLSX...")
            main_data_xlsx = create_xlsx_main_table_data(final_transactions, data.get("tax_details", {}), raw_data.get("impostos_raw", {}))
            
            print("  - Criando arquivo XLSX 100% igual ao PDF...")
            create_xlsx_file(output_filename_xlsx, main_data_xlsx, summaries, data["header"])
            
            print(f"✅ SUCESSO: Relatório XLSX gerado!\n📄 Arquivo XLSX: '{output_filename_xlsx}'")
            arquivos_gerados.append(output_filename_xlsx)
        except Exception as ex:
            print(f"❌ ERRO ao gerar XLSX: {ex}")
            import traceback
            traceback.print_exc()

    # ===== RESUMO FINAL =====
    print(f"📊 Total de notas fiscais processadas: {len(final_transactions)}")
    print(f"📅 Período: {format_date(data['header'].get('periodo_inicio'))} a {format_date(data['header'].get('periodo_fim'))}")
    
    return arquivos_gerados